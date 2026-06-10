from django.db.models import Count, Q
from django.http import HttpResponse

from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.models import User
from .models import Lead, LeadActivity, Project
from .serializers import LeadSerializer, ProjectSerializer


# ─── helpers ────────────────────────────────────────────────────────────────

def _team_queryset():
    return (
        User.objects
        .filter(is_active=True, role__in=['STM', 'Sales Executive'])
        .annotate(lead_count=Count('assigned_leads'))
    )


def _team_data(qs):
    return [
        {
            'id':         m.id,
            'name':       m.name,
            'role':       m.role,
            'initials':   ''.join(n[0] for n in m.name.split()[:2]).upper() if m.name else '??',
            'lead_count': m.lead_count,
        }
        for m in qs
    ]


# ─── Dashboard ──────────────────────────────────────────────────────────────

class PreSalesDashboardView(APIView):
    """GET /api/presales/dashboard/"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        stats = {
            'total': Lead.objects.count(),
            'new':   Lead.objects.filter(status='New').count(),
            'cold':  Lead.objects.filter(status='Cold').count(),
            'warm':  Lead.objects.filter(status='Warm').count(),
            'lost':  Lead.objects.filter(status='Lost').count(),
        }

        recent_leads = (
            Lead.objects
            .select_related('project', 'assigned_to')
            .prefetch_related('activities')
            .order_by('-created_at')[:5]
        )

        team = _team_queryset().order_by('-lead_count')

        active_projects = (
            Project.objects
            .annotate(_lead_count=Count('leads'))
            .filter(status='Active')
            .order_by('-_lead_count')[:4]
        )
        for p in active_projects:
            p._lead_count = p._lead_count

        return Response({
            'stats':           stats,
            'recent_leads':    LeadSerializer(recent_leads, many=True).data,
            'team_queue':      _team_data(team),
            'active_projects': ProjectSerializer(active_projects, many=True).data,
        })


# ─── Team ───────────────────────────────────────────────────────────────────

class TeamMembersView(APIView):
    """GET /api/presales/team/"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(_team_data(_team_queryset().order_by('name')))


# ─── Projects ───────────────────────────────────────────────────────────────

class ProjectListCreateView(APIView):
    """GET /api/presales/projects/   POST /api/presales/projects/"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Project.objects.annotate(_lead_count=Count('leads'))

        type_   = request.query_params.get('type')
        status_ = request.query_params.get('status')
        search  = request.query_params.get('search')

        if type_:
            qs = qs.filter(type=type_)
        if status_:
            qs = qs.filter(status=status_)
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(location__icontains=search))

        for p in qs:
            p._lead_count = p._lead_count

        return Response(ProjectSerializer(qs, many=True).data)

    def post(self, request):
        serializer = ProjectSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProjectDetailView(APIView):
    """GET / PATCH / DELETE /api/presales/projects/<pk>/"""
    permission_classes = [IsAuthenticated]

    def _get(self, pk):
        try:
            p = Project.objects.annotate(_lead_count=Count('leads')).get(pk=pk)
            p._lead_count = p._lead_count
            return p
        except Project.DoesNotExist:
            return None

    def get(self, request, pk):
        project = self._get(pk)
        if not project:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(ProjectSerializer(project).data)

    def patch(self, request, pk):
        project = self._get(pk)
        if not project:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ProjectSerializer(project, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(ProjectSerializer(self._get(pk)).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        project = self._get(pk)
        if not project:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        project.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ─── Leads ──────────────────────────────────────────────────────────────────

class LeadListCreateView(APIView):
    """GET /api/presales/leads/   POST /api/presales/leads/"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Lead.objects.select_related('project', 'assigned_to').prefetch_related('activities')

        status_ = request.query_params.get('status')
        project = request.query_params.get('project')
        assignee = request.query_params.get('assigned_to')
        search  = request.query_params.get('search')

        if status_:
            qs = qs.filter(status=status_)
        if project:
            qs = qs.filter(project_id=project)
        if assignee:
            qs = qs.filter(assigned_to_id=assignee)
        if search:
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(phone__icontains=search) |
                Q(project__name__icontains=search),
            )

        return Response(LeadSerializer(qs, many=True).data)

    def post(self, request):
        serializer = LeadSerializer(data=request.data)
        if serializer.is_valid():
            lead = serializer.save(created_by=request.user)
            LeadActivity.objects.create(
                lead=lead,
                type='Enquiry',
                note=f'Lead created via {lead.get_source_display()}.',
                created_by=request.user,
            )
            return Response(
                LeadSerializer(
                    Lead.objects.select_related('project', 'assigned_to')
                    .prefetch_related('activities')
                    .get(pk=lead.pk)
                ).data,
                status=status.HTTP_201_CREATED,
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LeadDetailView(APIView):
    """GET / PATCH / DELETE /api/presales/leads/<pk>/"""
    permission_classes = [IsAuthenticated]

    def _get(self, pk):
        try:
            return (
                Lead.objects
                .select_related('project', 'assigned_to')
                .prefetch_related('activities')
                .get(pk=pk)
            )
        except Lead.DoesNotExist:
            return None

    def get(self, request, pk):
        lead = self._get(pk)
        if not lead:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(LeadSerializer(lead).data)

    def patch(self, request, pk):
        lead = self._get(pk)
        if not lead:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = LeadSerializer(lead, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(LeadSerializer(self._get(pk)).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        lead = self._get(pk)
        if not lead:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        lead.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ─── Lead Status Change ──────────────────────────────────────────────────────

class LeadStatusChangeView(APIView):
    """PATCH /api/presales/leads/<pk>/status/"""
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        try:
            lead = Lead.objects.get(pk=pk)
        except Lead.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        new_status = request.data.get('status')
        note       = request.data.get('note', '').strip()

        if new_status not in ['New', 'Cold', 'Warm', 'Lost']:
            return Response(
                {'detail': 'Status must be one of: New, Cold, Warm, Lost.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        old_status  = lead.status
        lead.status = new_status
        lead.save()

        activity_note = f'Status changed from {old_status} to {new_status}.'
        if note:
            activity_note += f'\nRemark: {note}'

        LeadActivity.objects.create(
            lead=lead,
            type='Status Change',
            note=activity_note,
            created_by=request.user,
        )

        return Response(
            LeadSerializer(
                Lead.objects.select_related('project', 'assigned_to')
                .prefetch_related('activities')
                .get(pk=lead.pk)
            ).data
        )


# ─── Lead Transfer ───────────────────────────────────────────────────────────

class LeadTransferView(APIView):
    """POST /api/presales/leads/<pk>/transfer/"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            lead = Lead.objects.get(pk=pk)
        except Lead.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        assignee_id = request.data.get('assignee_id')
        note        = request.data.get('note', '').strip()

        if assignee_id:
            try:
                assignee = User.objects.get(pk=assignee_id)
            except User.DoesNotExist:
                return Response({'detail': 'Assignee not found.'}, status=status.HTTP_404_NOT_FOUND)
        else:
            # Auto-assign: team member with fewest active leads
            assignee = (
                _team_queryset()
                .order_by('lead_count', 'name')
                .first()
            )
            if not assignee:
                return Response(
                    {'detail': 'No team members available for auto-assign.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        lead.assigned_to = assignee
        lead.status      = 'Warm'
        lead.save()

        activity_note = f'Lead transferred to {assignee.name} ({assignee.role}) and marked Warm.'
        if note:
            activity_note += f'\nRemark: {note}'

        LeadActivity.objects.create(
            lead=lead,
            type='Transfer',
            note=activity_note,
            created_by=request.user,
        )

        return Response(
            LeadSerializer(
                Lead.objects.select_related('project', 'assigned_to')
                .prefetch_related('activities')
                .get(pk=lead.pk)
            ).data
        )


# ─── Lead Follow-up ──────────────────────────────────────────────────────────

class LeadFollowupView(APIView):
    """PATCH /api/presales/leads/<pk>/followup/"""
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        from datetime import date as date_cls

        try:
            lead = Lead.objects.get(pk=pk)
        except Lead.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        followup_str = request.data.get('next_followup')

        if followup_str:
            try:
                parsed = date_cls.fromisoformat(str(followup_str))
            except ValueError:
                return Response(
                    {'detail': 'Invalid date format. Use YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            lead.next_followup = parsed
            note = f'Follow-up scheduled for {parsed.strftime("%-d %b %Y")}.'
        else:
            lead.next_followup = None
            note = 'Follow-up date cleared.'

        lead.save()

        LeadActivity.objects.create(
            lead=lead, type='Note', note=note, created_by=request.user,
        )

        return Response(
            LeadSerializer(
                Lead.objects.select_related('project', 'assigned_to')
                .prefetch_related('activities')
                .get(pk=lead.pk)
            ).data
        )


# ─── Excel Bulk Upload ────────────────────────────────────────────────────────

_HEADER_MAP = {
    'name':           'name',    'full name':     'name',
    'phone':          'phone',   'phone number':  'phone',   'mobile':   'phone',
    'email':          'email',   'email address': 'email',
    'project':        'project', 'project name':  'project',
    'source':         'source',  'lead source':   'source',
    'status':         'status',
    'budget':         'budget',  'budget range':  'budget',
    'notes':          'notes',   'note':          'notes',   'remarks':  'notes',
    'next followup':  'next_followup', 'follow up':   'next_followup',
    'followup date':  'next_followup', 'follow-up':   'next_followup',
}

_VALID_SOURCES  = {'Walk-in', 'Phone', 'Online', 'Reference', 'Email'}
_VALID_STATUSES = {'New', 'Cold', 'Warm', 'Lost'}


class LeadBulkUploadView(APIView):
    """POST /api/presales/leads/bulk-upload/"""
    permission_classes = [IsAuthenticated]
    parser_classes     = [MultiPartParser]

    def post(self, request):
        import openpyxl
        from datetime import date as date_cls

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        ext = file.name.lower().rsplit('.', 1)[-1] if '.' in file.name else ''
        if ext not in ('xlsx', 'xls'):
            return Response({'detail': 'Only .xlsx files are supported.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb   = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            return Response({'detail': f'Could not read file: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        if len(rows) < 2:
            return Response({'detail': 'File has no data rows.'}, status=status.HTTP_400_BAD_REQUEST)

        headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
        col_map = {}
        for i, h in enumerate(headers):
            key = _HEADER_MAP.get(h)
            if key and key not in col_map:
                col_map[key] = i

        if 'name' not in col_map or 'phone' not in col_map:
            return Response(
                {'detail': 'Excel must have "Name" and "Phone" columns.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        project_lookup = {p.name.lower(): p for p in Project.objects.all()}
        created_count  = 0
        failed_count   = 0
        errors         = []

        for row_num, row in enumerate(rows[1:], start=2):
            def cell(key, _row=row):
                idx = col_map.get(key)
                if idx is None or idx >= len(_row) or _row[idx] is None:
                    return ''
                return str(_row[idx]).strip()

            name  = cell('name')
            phone = cell('phone')

            if not name or not phone:
                failed_count += 1
                errors.append({'row': row_num, 'name': name or '—', 'error': 'Name and Phone are required.'})
                continue

            if Lead.objects.filter(phone=phone).exists():
                failed_count += 1
                errors.append({'row': row_num, 'name': name, 'error': f'Phone {phone} already exists.'})
                continue

            project     = project_lookup.get(cell('project').lower())
            lead_source = cell('source') or 'Walk-in'
            if lead_source not in _VALID_SOURCES:
                lead_source = 'Walk-in'
            lead_status = cell('status') or 'New'
            if lead_status not in _VALID_STATUSES:
                lead_status = 'New'

            followup = None
            followup_str = cell('next_followup')
            if followup_str:
                try:
                    followup = date_cls.fromisoformat(followup_str)
                except ValueError:
                    pass

            try:
                lead = Lead.objects.create(
                    name=name, phone=phone, email=cell('email'),
                    project=project, source=lead_source, status=lead_status,
                    budget=cell('budget'), notes=cell('notes'),
                    next_followup=followup, created_by=request.user,
                )
                LeadActivity.objects.create(
                    lead=lead, type='Enquiry',
                    note='Lead imported via Excel upload.',
                    created_by=request.user,
                )
                created_count += 1
            except Exception as e:
                failed_count += 1
                errors.append({'row': row_num, 'name': name, 'error': str(e)})

        return Response({
            'created': created_count,
            'failed':  failed_count,
            'total':   created_count + failed_count,
            'errors':  errors,
        })


class LeadUploadTemplateView(APIView):
    """GET /api/presales/leads/upload-template/  — public, no auth needed"""
    permission_classes = [AllowAny]

    def get(self, request):
        import openpyxl
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Leads'

        headers     = ['Name', 'Phone', 'Email', 'Project', 'Source',
                       'Status', 'Budget', 'Notes', 'Next Followup']
        header_fill = PatternFill(start_color='1E4080', end_color='1E4080', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill      = header_fill
            c.font      = header_font
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[c.column_letter].width = 20

        ws.append([
            'Rajesh Sharma', '+91 98765 43210', 'rajesh@example.com',
            'Vistara Heights', 'Walk-in', 'New', '60L - 70L',
            'Interested in 2BHK', '2026-07-01',
        ])

        notes_ws = wb.create_sheet('Valid Values')
        notes_ws.append(['Column',        'Valid Values'])
        notes_ws.append(['Source',        'Walk-in, Phone, Online, Reference, Email'])
        notes_ws.append(['Status',        'New, Cold, Warm, Lost'])
        notes_ws.append(['Project',       'Must match an existing project name exactly'])
        notes_ws.append(['Next Followup', 'YYYY-MM-DD format  e.g. 2026-07-15'])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="leads_upload_template.xlsx"'
        return resp
