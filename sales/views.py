import logging
import os
import hmac
import secrets
from datetime import datetime, time as dt_time, timedelta
import requests as http_requests
from django.conf import settings
from django.db import transaction
from django.db.models import Q, Count, OuterRef, Subquery
from django.utils import timezone
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny

logger = logging.getLogger(__name__)

from accounts.models import User
from accounts.permissions import is_platform_admin, scope_to_company
from sales.fields import phone_blind_index


def _resolve_company(request):
    """Return the company for the request, honouring ?company_id for platform admins."""
    cid = request.query_params.get('company_id') or request.data.get('company_id')
    if cid and is_platform_admin(request.user):
        Company = __import__('companies.models', fromlist=['Company']).Company
        return Company.objects.filter(pk=cid).first() or request.user.company
    return request.user.company
from .models import (
    Lead, LeadSource, Project, Plot, FollowUp, SiteVisit, Closure, LeadStatusHistory,
    DistributionSettings, UserAvailability, UserDistributionWeight, DistributionLog,
    SalesTeamMember, MetaWebhookConfig, MetaFormMapping,
    UserProjectAssignment, Booking, BackupSettings, BackupRecord, LeadTransfer, ChannelPartner,
)
from .serializers import (
    LeadListSerializer, LeadDetailSerializer, LeadCreateSerializer, LeadUpdateSerializer,
    LeadSourceSerializer, ProjectSerializer, PlotSerializer,
    FollowUpSerializer, SiteVisitSerializer, ClosureSerializer,
    LeadStatusHistorySerializer, BookingSerializer,
    BackupSettingsSerializer, BackupRecordSerializer, LeadTransferSerializer, ChannelPartnerSerializer,
)

PAGE_SIZE = 25


# The staff hierarchy, most senior first. Everything at or above Manager carries
# Manager's authority — a Director who could do less than the Manager reporting to
# them would be a permissions bug, so these are compared as a set rather than
# spelled out at each call site.
ROLE_HIERARCHY = ['Director', 'General Manager', 'Manager', 'Employee', 'Intern']
MANAGER_ROLES = ('Admin', 'Director', 'General Manager', 'Manager')


def is_manager_role(user):
    """True for Manager and anything senior to it (not Admin/staff — see below)."""
    return getattr(user, 'role', '') in ('Director', 'General Manager', 'Manager')


def is_admin_or_manager(user):
    return user.role in MANAGER_ROLES or user.is_staff


def _is_sales_admin(user):
    """True/hard Admin, platform staff, or a Sales Admin-Modules user — the same
    gate as Data Reset (see SalesDataResetView._is_admin), used for admin-only
    company master data like Channel Partners."""
    return bool(
        getattr(user, 'is_staff', False) or getattr(user, 'role', '') == 'Admin' or is_platform_admin(user)
        or 'Sales' in (getattr(user, 'admin_modules', None) or [])
    )


def has_sales_access(user):
    """Admin/Manager, or a plain employee who's been granted the Sales module —
    used for actions (like bulk lead import) that used to be manager-only but
    shouldn't be gated tighter than "can this person use Sales at all"."""
    return is_admin_or_manager(user) or 'Sales' in (getattr(user, 'modules', None) or [])


def _designation(user):
    return (getattr(user, 'designation', '') or '').lower()


def is_telecaller(user):
    d = _designation(user)
    return 'telecaller' in d or 'tele caller' in d


def is_stm(user):
    d = _designation(user)
    return 'stm' in d or 'sales team' in d or 'sales executive' in d


def is_cp(user):
    """CP Executive — an employee-level Channel Partner who sources & works their
    own leads (no Meta distribution). Scoped like an STM (by the lead's stm field)."""
    d = _designation(user)
    return 'cp executive' in d or 'channel partner' in d


def is_cp_manager(user):
    """A Manager whose designation starts with 'cp' (e.g. 'CP Cluster Head') —
    gets into the Channel Partner module ONLY, not the rest of Sales. Their lead
    visibility within it still comes from the existing Manager project-assignment
    mechanism (manager_project_ids/scope_leads_to_project) — no CP-specific
    scoping needed, it already applies to any Manager regardless of designation."""
    return getattr(user, 'role', '') == 'Manager' and _designation(user).startswith('cp')


def can_access_cp_module(user):
    """Who can reach the Channel Partner module: true/hard admins always can
    (see _is_hard_admin); a CP-designation Manager gets in via their
    designation. Mirrors web's canAccessChannelPartner(user) — keep in sync."""
    return bool(_is_hard_admin(user) or is_cp_manager(user))


def cp_lead_q(prefix=''):
    """A lead belongs to the Channel Partner module if EITHER it was added
    through the CP module itself (channel_partner FK set) OR it was added
    through the regular Sales flow with Source explicitly set to "Channel
    Partner" (a plain LeadSource, no specific partner attached). `prefix`
    lets callers scope a related model (e.g. 'lead__' for FollowUp/SiteVisit/
    Closure/Booking, which don't have these fields directly)."""
    return (
        Q(**{f'{prefix}channel_partner__isnull': False}) |
        Q(**{f'{prefix}source__name__iexact': 'Channel Partner'})
    )


# ── Hierarchy-based visibility ───────────────────────────────────────────────
# Data visibility is driven by the org tree (User.reporting_manager), NOT by
# designation strings. A user sees records owned (as STM or telecaller) by
# themselves or by anyone reporting to them, transitively. This scales to any
# designation/role without code changes — you only maintain reporting_manager.

def _sees_all_company(user, request=None, include_manager_role=True):
    """Users who see ALL company data: platform admins, staff, the Admin role,
    Managers, and top-of-tree department heads (report to no one but manage others).

    A Manager sees every project's leads, follow-ups, site visits and closures — the
    reporting chain does not limit them. Bookings are the deliberate exception: that
    surface is scoped by who is *named an approver* on a project, so the booking
    views call this with include_manager_role=False and a Manager falls through to
    the approver/reporting-chain rules below. MyTeamView also opts out, keeping its
    own `?scope=all` org-chart toggle as the way to widen that view.

    A Sales Admin-Modules user (a Manager granted 'Sales' in Admin Modules) gets
    full company visibility when `request` carries `?admin_view=1` — sent only by
    the web/app's mirrored "Admin" section pages (see isSalesModuleAdmin in
    sales/layout.js). Deliberately a distinct param name from the pre-existing
    `scope` (used by MyTeamView for its own unrelated org-chart toggle) to avoid
    colliding with it. Real admins (Chinmay, Prince, platform staff) are unaffected
    — they already return True unconditionally below."""
    if is_platform_admin(user) or user.is_staff or getattr(user, 'role', '') == 'Admin':
        return True
    if include_manager_role and is_manager_role(user):
        return True
    if (request is not None and request.query_params.get('admin_view') == '1'
            and 'Sales' in (getattr(user, 'admin_modules', None) or [])):
        return True
    # Top of the tree: reports to nobody, but has active reports under them.
    if user.reporting_manager_id is None and User.objects.filter(
        company=user.company, reporting_manager_id=user.id, is_active=True
    ).exists():
        return True
    return False


def _visible_user_ids(user):
    """Requester's own id + every user reporting to them, transitively, in the same
    company. Cycle-safe (tracked via the `ids` set) and depth-capped."""
    ids = {user.id}
    frontier = [user.id]
    for _ in range(50):  # safety cap on tree depth
        children = list(
            User.objects.filter(
                company=user.company, reporting_manager_id__in=frontier, is_active=True
            ).exclude(id__in=ids).values_list('id', flat=True)
        )
        if not children:
            break
        ids.update(children)
        frontier = children
    return ids


def _is_hard_admin(user):
    """A real company/platform administrator, as opposed to someone who merely reaches
    company-wide visibility through the org tree (a top-of-tree department head) or the
    Sales admin-modules flag. Only these are exempt from per-project approver scoping."""
    return is_platform_admin(user) or user.is_staff or getattr(user, 'role', '') == 'Admin'


def _approver_project_ids(user, company):
    """Ids of projects where `user` is a configured booking approver — they should
    see every booking for that project regardless of the STM's reporting chain."""
    return [
        p.id for p in Project.objects.filter(company=company).only('id', 'booking_approvers')
        if user.id in (p.booking_approvers or [])
    ]


def _can_approve_project(user, project, company):
    """Whether `user` is a configured approver for `project`'s bookings.

    Only the managers named on the project may approve it — being a manager is not
    itself authority to approve. A project that names nobody is approvable only by a
    real admin, rather than by everyone: the previous rule returned True whenever the
    list was empty, which let any manager approve bookings for the projects that had
    not been configured yet.

    Shared by approve/reject and cancel so the two cannot drift apart — cancel undoes
    an approval, frees the plots and deletes the signed LOI, so it needs at least the
    same authority as granting the approval did.
    """
    if _is_hard_admin(user):
        return True
    project_id = getattr(project, 'id', project)
    return project_id in _approver_project_ids(user, company)


def _cp_approver_project_ids(user, company):
    """Ids of projects where `user` is a configured CP booking approver — mirrors
    _approver_project_ids but reads cp_booking_approvers, the separate list that
    gates bookings whose lead came through a Channel Partner."""
    return [
        p.id for p in Project.objects.filter(company=company).only('id', 'cp_booking_approvers')
        if user.id in (p.cp_booking_approvers or [])
    ]


def _can_approve_cp_project(user, project, company):
    """Whether `user` is a configured CP approver for `project`'s Channel-Partner-
    sourced bookings — mirrors _can_approve_project exactly, against the separate
    cp_booking_approvers list."""
    if _is_hard_admin(user):
        return True
    project_id = getattr(project, 'id', project)
    return project_id in _cp_approver_project_ids(user, company)


def _is_cp_sourced(lead_id):
    """Whether the given lead (if any) is Channel-Partner-referred — decides which
    approver list (regular vs CP) gates a booking/closure tied to it. Uses the
    same definition as cp_lead_q (channel_partner FK OR Source = "Channel
    Partner") so a lead that shows up as CP in Leads/Dashboard also gets CP
    approval routing once it's booked — they used to disagree."""
    if not lead_id:
        return False
    return Lead.objects.filter(cp_lead_q(), id=lead_id).exists()


def _is_cp_sourced_booking(lead_id, booking_source=None):
    """Whether a booking counts as Channel-Partner-sourced for approval routing —
    either its lead is CP-attributed (the structured directory), or its own
    free-text Source field was set to "Channel Partner" (the older per-booking
    selector on the booking form itself, unconnected to the ChannelPartner
    directory but meant the same way: this deal came through a partner, so it
    should route to the CP approvers, not the project's regular ones)."""
    if (booking_source or '').strip().lower() == 'channel partner':
        return True
    return _is_cp_sourced(lead_id)


def _can_approve_booking(user, project_id, project, lead_id, company, booking_source=None):
    """Which approver list gates a booking or closure: a Channel-Partner-sourced
    one is approved by the project's CP approvers, everything else by its regular
    ones. No project at all → no approver-scoping check applies (matches the
    original `if b.project_id and not _can_approve_project(...)` shape at every
    call site this replaces)."""
    if not project_id:
        return True
    if _is_cp_sourced_booking(lead_id, booking_source):
        return _can_approve_cp_project(user, project, company)
    return _can_approve_project(user, project, company)


def can_assign_leads(user):
    """Telecallers, STMs & CP Executives cannot (re)assign leads — only everyone
    else (admins/managers/Sales CRM)."""
    return not (is_telecaller(user) or is_stm(user) or is_cp(user))


def _dist_type_for(user):
    """'telecaller' | 'stm' | None for a user based on their designation."""
    if is_telecaller(user):
        return 'telecaller'
    if is_stm(user):
        return 'stm'
    return None


# Self-marked availability stays active for this many hours, then auto-resets.
AVAILABILITY_TTL_HOURS = 12



def _role_signout(company, designation):
    """Configured sign-out time for a TC/STM role, or None (no settings / other role)."""
    s = DistributionSettings.objects.filter(company=company).first()
    if not s:
        return None
    d = (designation or '').lower()
    if 'telecaller' in d or 'tele caller' in d:
        return s.tc_signout_time
    if 'stm' in d or 'sales team' in d or 'sales executive' in d:
        return s.stm_signout_time
    return None


def _availability_expires_at(user):
    """ISO timestamp when the user's availability auto-expires today = the role's
    sign-out time. None if no sign-out is configured (caller falls back to the TTL)."""
    signout = _role_signout(getattr(user, 'company', None), getattr(user, 'designation', ''))
    if signout is None:
        return None
    from zoneinfo import ZoneInfo
    from datetime import datetime as _dt
    tz = ZoneInfo('Asia/Kolkata')
    return _dt.combine(timezone.now().astimezone(tz).date(), signout, tzinfo=tz).isoformat()


def _availability_active(avail, user=None):
    """True if marked available *today* and it's still before the role's configured
    sign-out time — availability auto-expires at sign-out. Falls back to a 12h TTL
    when the company has no distribution sign-out configured."""
    if not avail or not avail.is_available or not avail.checked_in_at:
        return False
    from zoneinfo import ZoneInfo
    now_ist = timezone.now().astimezone(ZoneInfo('Asia/Kolkata'))
    if avail.date != now_ist.date():          # a stale prior-day record is expired
        return False
    u = user or avail.user
    signout = _role_signout(u.company, u.designation)
    if signout is None:                        # no sign-out configured → legacy 12h TTL
        return (timezone.now() - avail.checked_in_at) < timedelta(hours=AVAILABILITY_TTL_HOURS)
    return now_ist.time() < signout            # auto-expires at sign-out


# Only Manager is confined to assigned projects. Director and General Manager sit
# above the project line and always see the whole company, so they are never scoped
# and are never offered a project assignment.
PROJECT_SCOPED_ROLES = ('Manager',)


def manager_project_ids(user):
    """Projects a Manager is confined to, or None if they are not project-scoped.

    A Manager who has projects assigned sees leads, site visits and closures for
    those projects only, instead of the whole company.

    Assignment is the opt-in: a Manager with no project assigned keeps company-wide
    visibility, so introducing this does not blank out anyone's screens — you scope a
    manager by assigning them projects. Admins, platform staff, Directors and General
    Managers are never scoped.

    Bookings deliberately do not use this: a manager may book a plot on any project.
    """
    if is_platform_admin(user) or getattr(user, 'is_staff', False) or getattr(user, 'role', '') == 'Admin':
        return None
    if getattr(user, 'role', '') not in PROJECT_SCOPED_ROLES:
        return None
    pids = list(
        UserProjectAssignment.objects.filter(user=user).values_list('project_id', flat=True)
    )
    return pids or None


def scope_leads_to_project(qs, user, lead_prefix=''):
    """Narrow a lead-side queryset to the manager's assigned projects, if any."""
    pids = manager_project_ids(user)
    if pids is None:
        return qs
    return qs.filter(**{f'{lead_prefix}project__in': pids})


def scope_leads_to_role(qs, user, lead_prefix='', request=None):
    """Restrict a Lead-related queryset by org hierarchy: a user sees leads OWNED (as
    STM or telecaller) by themselves or by anyone reporting to them, transitively.
    Admins / staff / top-of-tree heads see all company data. `lead_prefix` lets callers
    scope related models (e.g. 'lead__' for SiteVisit / Closure). Pass `request` through
    so a Sales Admin-Modules user gets full data when their Admin section explicitly
    asks for it via `?scope=company` (see _sees_all_company).

    A Sales department manager (Sales in manager_modules — the same flag Club 1000
    uses for its manager-level access) additionally sees the whole unassigned pool
    (no stm, no telecaller) alongside their own team's owned leads — they need that
    visibility to review and distribute leads, not just to see what's already theirs."""
    # Frontline callers (telecaller / STM / CP) are ALWAYS restricted to leads owned
    # by them (or their own reports) — never the unassigned pool, never full-company
    # data — regardless of any manager_modules flag or reporting-line quirk. Without
    # this a telecaller who also carries the 'Sales' manager flag (mis-config) would
    # see every unrouted lead in the company as "My Leads".
    if is_telecaller(user) or is_stm(user) or is_cp(user):
        ids = _visible_user_ids(user)
        return qs.filter(
            Q(**{f'{lead_prefix}stm__in': ids}) | Q(**{f'{lead_prefix}telecaller__in': ids})
        )
    if _sees_all_company(user, request):
        # A manager assigned to specific projects sees only those projects' leads.
        return scope_leads_to_project(qs, user, lead_prefix)
    ids = _visible_user_ids(user)
    own_filter = Q(**{f'{lead_prefix}stm__in': ids}) | Q(**{f'{lead_prefix}telecaller__in': ids})
    if 'Sales' in (getattr(user, 'manager_modules', None) or []):
        own_filter |= Q(**{f'{lead_prefix}stm__isnull': True}) & Q(**{f'{lead_prefix}telecaller__isnull': True})
    return qs.filter(own_filter)


def _lead_in_scope(request, lead_id):
    """True if the given lead belongs to the requester's company (or requester is platform admin)."""
    if not lead_id:
        return False
    return scope_to_company(Lead.objects.filter(pk=lead_id), request.user).exists()


def _project_in_scope(request, project_id):
    """True if the given project belongs to the requester's company (or requester is platform admin)."""
    if not project_id:
        return False
    return scope_to_company(Project.objects.filter(pk=project_id), request.user).exists()


class StatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.core.cache import cache

        # Dashboard stats are ~5 COUNT/aggregate queries; cache briefly per
        # (user, company) so repeated dashboard loads don't re-hit Postgres.
        # 20s TTL keeps numbers near-live. Shared (consistent) once Redis is on.
        company_id = request.query_params.get('company_id')
        date_from  = request.query_params.get('date_from')
        date_to    = request.query_params.get('date_to')

        # Include date range, admin_view AND cp_only in cache key — otherwise a Sales
        # Admin-Modules user's team-scoped dashboard and their Admin-section (full
        # company) dashboard, or the regular Sales dashboard and the Channel Partner
        # one, would collide on the same key and serve each other's stale data.
        admin_view = request.query_params.get('admin_view') == '1'
        cp_only = request.query_params.get('cp_only') == 'true'
        cache_key = f'sales_stats:{request.user.id}:{company_id or "own"}:{date_from or ""}:{date_to or ""}:{"admin" if admin_view else "own"}:{"cp" if cp_only else "all"}'
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached)

        today = timezone.localdate()
        leads_qs = scope_to_company(Lead.objects.all(), request.user)

        # Channel Partner leads have their own module and dashboard — the main
        # Sales dashboard's counts never include them, and the CP dashboard's
        # counts are ONLY them (see LeadListView for the matching All Leads
        # behaviour, and cp_lead_q for what counts as a CP lead).
        if cp_only:
            leads_qs = leads_qs.filter(cp_lead_q())
        else:
            leads_qs = leads_qs.exclude(cp_lead_q())

        # Telecallers / STMs only see stats for leads assigned to them.
        leads_qs = scope_leads_to_role(leads_qs, request.user, request=request)

        # Platform admin: filter by a specific company (used by admin company picker)
        if company_id and is_platform_admin(request.user):
            leads_qs   = leads_qs.filter(company_id=company_id)
            sv_filter  = {'lead__company_id': company_id}
            cl_filter  = {'company_id': company_id}
            prj_filter = {'company_id': company_id}
        else:
            sv_filter = cl_filter = prj_filter = {}

        # Role/company-scoped leads WITHOUT the created_at window — used for the
        # SQL funnel count (leads that *became* warm within the window).
        leads_scope = leads_qs

        # Apply optional date range filter
        if date_from:
            leads_qs = leads_qs.filter(created_at__date__gte=date_from)
        if date_to:
            leads_qs = leads_qs.filter(created_at__date__lte=date_to)

        # Single aggregate query instead of 6 separate COUNTs
        agg = leads_qs.aggregate(
            total_leads=Count('id'),
            new_leads=Count('id', filter=Q(status='new')),
            leads_today=Count('id', filter=Q(created_at__date=today)),
            called_count=Count('id', filter=~Q(telecaller_status='') & Q(telecaller_status__isnull=False)),
        )

        # Status-bucket counts (hot/warm/callback/not_reachable/cold, and the STM
        # equivalents) are counted by the date the lead's status actually CHANGED
        # to that value (LeadStatusHistory), not by created_at. A lead received on
        # an earlier date but marked warm today must land in "today"'s warm count —
        # counting against the created_at-filtered leads_qs hid it entirely. Same
        # fix already applied to sql_count below and to StatsTrendView's per-day
        # warm/hot/cold rows; this brings the stat-card tiles in line with those.
        def _status_transition_count(field, value):
            qs = LeadStatusHistory.objects.filter(
                lead__in=leads_scope, field_changed=field, new_value=value)
            if date_from:
                qs = qs.filter(created_at__date__gte=date_from)
            if date_to:
                qs = qs.filter(created_at__date__lte=date_to)
            return qs.values('lead').distinct().count()

        hot_count           = _status_transition_count('telecaller_status', 'hot')
        warm_count          = _status_transition_count('telecaller_status', 'warm')
        callback_count      = _status_transition_count('telecaller_status', 'callback')
        not_reachable_count = _status_transition_count('telecaller_status', 'not_reachable')
        cold_count          = _status_transition_count('telecaller_status', 'cold')

        # STM-pipeline hot/warm/cold: a lead that got a site-visit outcome of Hot
        # still sits at stm_status='sv_done' (the outcome doesn't overwrite it —
        # see SiteVisitDetailView), so it counts toward Hot here too, dated by
        # when that visit was completed rather than a stm_status transition.
        _latest_sv_for_lead = SiteVisit.objects.filter(
            lead=OuterRef('pk'), status='completed',
        ).order_by('-visited_at')

        def _sv_outcome_lead_ids(value):
            qs = leads_scope.filter(stm_status='sv_done').annotate(
                _sv_outcome=Subquery(_latest_sv_for_lead.values('outcome')[:1]),
                _sv_visited=Subquery(_latest_sv_for_lead.values('visited_at')[:1]),
            ).filter(_sv_outcome=value)
            if date_from:
                qs = qs.filter(_sv_visited__date__gte=date_from)
            if date_to:
                qs = qs.filter(_sv_visited__date__lte=date_to)
            return set(qs.values_list('id', flat=True))

        def _effective_stm_count(value):
            direct_hist = LeadStatusHistory.objects.filter(
                lead__in=leads_scope, field_changed='stm_status', new_value=value)
            if date_from:
                direct_hist = direct_hist.filter(created_at__date__gte=date_from)
            if date_to:
                direct_hist = direct_hist.filter(created_at__date__lte=date_to)
            direct_ids = set(direct_hist.values_list('lead_id', flat=True))
            return len(direct_ids | _sv_outcome_lead_ids(value))

        # STM-pipeline counts (by stm_status) for the STM/CP dashboard.
        stm_hot_count           = _effective_stm_count('hot')
        stm_warm_count          = _effective_stm_count('warm')
        stm_cold_count          = _effective_stm_count('cold')
        stm_sv_scheduled_count  = _status_transition_count('stm_status', 'sv_scheduled')
        # The Site Visits tile reports visits that actually HAPPENED — a scheduled,
        # no-show or cancelled visit is not one. Counting every row made the tile read
        # 48 where only 26 had been done. Dated by when the visit happened, not when
        # the row was created, so a date range means "visited in this period" the same
        # way Closures means "closed in this period".
        sv_qs = scope_to_company(SiteVisit.objects.all(), request.user, 'lead__company')
        sv_qs = sv_qs.filter(status='completed', visited_at__isnull=False)
        cl_qs = scope_to_company(Closure.objects.all(), request.user, 'company')
        # Same CP/non-CP split as leads_qs above — these are independent queries
        # against SiteVisit/Closure, not derived from leads_qs, so they need the
        # same cp_lead_q filter applied directly or the CP dashboard's Site
        # Visits/Closures tiles would silently show the whole company's numbers.
        if cp_only:
            sv_qs = sv_qs.filter(cp_lead_q(prefix='lead__'))
            cl_qs = cl_qs.filter(cp_lead_q(prefix='lead__'))
        else:
            sv_qs = sv_qs.exclude(cp_lead_q(prefix='lead__'))
            cl_qs = cl_qs.exclude(cp_lead_q(prefix='lead__'))
        if not _sees_all_company(request.user, request):
            _ids = _visible_user_ids(request.user)
            sv_qs = sv_qs.filter(Q(stm__in=_ids) | Q(referred_by_telecaller__in=_ids))
            cl_qs = cl_qs.filter(Q(stm__in=_ids) | Q(referred_by_telecaller__in=_ids))
        if date_from:
            sv_qs = sv_qs.filter(visited_at__date__gte=date_from)
            cl_qs = cl_qs.filter(closure_date__gte=date_from)
        if date_to:
            sv_qs = sv_qs.filter(visited_at__date__lte=date_to)
            cl_qs = cl_qs.filter(closure_date__lte=date_to)
        # Follow-up calls: a completed follow-up IS a call that was made, counted on
        # the day it was completed so the dashboard's date filter applies to it the
        # same way it does to everything else. Scoped by assignee exactly as the
        # Follow-Ups screen is, so the tile and that list agree.
        fu_qs = scope_to_company(FollowUp.objects.all(), request.user, 'lead__company')
        if not _sees_all_company(request.user, request):
            fu_qs = fu_qs.filter(assigned_to__in=_visible_user_ids(request.user))
        if company_id and is_platform_admin(request.user):
            fu_qs = fu_qs.filter(lead__company_id=company_id)
        fu_done = fu_qs.filter(status='completed', completed_at__isnull=False)
        if date_from:
            fu_done = fu_done.filter(completed_at__date__gte=date_from)
        if date_to:
            fu_done = fu_done.filter(completed_at__date__lte=date_to)
        followup_call_count = fu_done.count()

        # Still-open follow-ups, for the Pending / Overdue tiles. A pending row has
        # no completed_at, so these are dated on scheduled_at instead: the range then
        # reads as "due in this window" and the tiles agree with the Follow-Ups
        # screen's own Pending/Overdue chips, which count the same way. Overdue is
        # that same set narrowed to rows already past their slot.
        fu_open = fu_qs.filter(status='pending')
        if date_from:
            fu_open = fu_open.filter(scheduled_at__date__gte=date_from)
        if date_to:
            fu_open = fu_open.filter(scheduled_at__date__lte=date_to)
        followup_pending_count = fu_open.count()
        followup_overdue_count = fu_open.filter(scheduled_at__lt=timezone.now()).count()

        # "To Call" backlog: assigned leads this user has not actioned yet. Same rule
        # as the All Leads "To Call" tab (work=pending), so the tile and that tab can
        # never disagree. Both status columns are blank-not-null, so '' is the whole
        # of "not yet worked".
        if is_telecaller(request.user):
            to_call_count = leads_qs.filter(telecaller_status='').count()
        elif is_stm(request.user) or is_cp(request.user):
            to_call_count = leads_qs.filter(stm_status='').count()
        else:
            to_call_count = leads_qs.filter(status='new').count()

        cl_scoped = cl_qs.filter(**cl_filter)
        sv_scoped = sv_qs.filter(**sv_filter)
        active_projects_qs = scope_to_company(Project.objects.filter(is_active=True), request.user).filter(**prj_filter)
        if cp_only:
            # The CP dashboard's "Active Projects" means projects with actual CP
            # activity, not every active project company-wide — leads_scope is
            # already split to the CP pool above (cp_lead_q), not date-windowed
            # so this stays a stable "currently active in CP" count.
            active_projects_qs = active_projects_qs.filter(id__in=leads_scope.values('project_id'))
        sv_done, closures, active_projects = (
            sv_scoped.count(),
            cl_scoped.count(),
            active_projects_qs.count(),
        )
        # Post-visit outcome breakdown of the same completed-visits window above.
        sv_hot_count  = sv_scoped.filter(outcome='hot').count()
        sv_warm_count = sv_scoped.filter(outcome='warm').count()
        sv_cold_count = sv_scoped.filter(outcome='cold').count()
        sv_not_interested_count = sv_scoped.filter(outcome='not_interested').count()

        # SQL funnel: distinct leads that are effectively warm (stm_status → warm,
        # or still sv_done with a Warm visit outcome) in the window — same
        # definition as stm_warm_count above.
        sql_count = stm_warm_count

        # Avg closure timeline: mean days from lead arrival (created_at) to closure_date.
        _diffs = [
            (cdate - created.date()).days
            for created, cdate in cl_scoped.values_list('lead__created_at', 'closure_date')
            if created and cdate
        ]
        avg_closure_days = round(sum(_diffs) / len(_diffs), 1) if _diffs else None
        # No .only() here: LeadListSerializer reads ~11 more fields (meta_*, statuses,
        # is_duplicate, …); deferring them caused a per-field query per lead (N+1).
        recent = leads_qs.select_related('project', 'source', 'telecaller', 'stm').order_by('-created_at')[:8]
        payload = {
            'total_leads':        agg['total_leads'],
            'new_leads':          agg['new_leads'],
            'leads_today':        agg['leads_today'],
            'called_count':       agg['called_count'],
            'to_call_count':      to_call_count,
            'followup_call_count': followup_call_count,
            'followup_pending_count': followup_pending_count,
            'followup_overdue_count': followup_overdue_count,
            # Every call made in the window: new leads worked plus follow-up calls.
            'total_called_count': agg['called_count'] + followup_call_count,
            'hot_count':          hot_count,
            'warm_count':         warm_count,
            'callback_count':     callback_count,
            'not_reachable_count':not_reachable_count,
            'cold_count':         cold_count,
            'stm_hot_count':          stm_hot_count,
            'stm_warm_count':         stm_warm_count,
            'stm_cold_count':         stm_cold_count,
            'stm_sv_scheduled_count': stm_sv_scheduled_count,
            'sv_done':            sv_done,
            'sv_hot_count':  sv_hot_count,
            'sv_warm_count': sv_warm_count,
            'sv_cold_count': sv_cold_count,
            'sv_not_interested_count': sv_not_interested_count,
            'closures':           closures,
            'sql_count':          sql_count,
            'avg_closure_days':   avg_closure_days,
            'active_projects':    active_projects,
            'recent_leads':       LeadListSerializer(recent, many=True).data,
        }
        cache.set(cache_key, payload, timeout=20)
        return Response(payload)


class StatsTrendView(APIView):
    """Daily MQL and SV counts for the last 30 days (or within date_from/date_to)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models.functions import TruncDate
        from datetime import date

        company_id = request.query_params.get('company_id')
        date_from  = request.query_params.get('date_from')
        date_to    = request.query_params.get('date_to')

        today = timezone.localdate()
        if not date_from:
            date_from = str(today - timedelta(days=29))
        if not date_to:
            date_to = str(today)

        leads_qs = scope_to_company(Lead.objects.all(), request.user)
        leads_qs = scope_leads_to_role(leads_qs, request.user, request=request)
        if company_id and is_platform_admin(request.user):
            leads_qs = leads_qs.filter(company_id=company_id)

        # MQL: of the leads that ARRIVED on each day, how many have had their
        # telecaller status set — "called" means a new lead came in and its status
        # was changed. Grouped by created_at so the chart counts exactly the leads
        # the Called/MQL tile counts over the same date filter.
        #
        # This was grouped by updated_at, which is neither: that column moves on any
        # edit, so a lead touched for an unrelated reason counted as a call and a
        # lead edited on several days counted on each one. On a live telecaller it
        # read 50 against a tile of 25.
        mql_rows = (
            leads_qs
            .filter(
                created_at__date__gte=date_from,
                created_at__date__lte=date_to,
                telecaller_status__isnull=False,
            )
            .exclude(telecaller_status='')
            .annotate(day=TruncDate('created_at'))
            .values('day')
            .annotate(count=Count('id'))
            .order_by('day')
        )

        sv_qs = scope_to_company(SiteVisit.objects.all(), request.user, 'lead__company')
        if not _sees_all_company(request.user, request):
            ids = _visible_user_ids(request.user)
            sv_qs = sv_qs.filter(Q(stm__in=ids) | Q(referred_by_telecaller__in=ids))
        if company_id and is_platform_admin(request.user):
            sv_qs = sv_qs.filter(lead__company_id=company_id)

        # SV: site visits created per day
        sv_rows = (
            sv_qs
            .filter(created_at__date__gte=date_from, created_at__date__lte=date_to)
            .annotate(day=TruncDate('created_at'))
            .values('day')
            .annotate(count=Count('id'))
            .order_by('day')
        )

        # Warm/SQL: count by the date the lead actually BECAME warm — the status-history
        # entry where telecaller_status changed to 'warm' — not when the lead arrived or
        # was last edited (updated_at). Scoped to the same visible leads.
        warm_rows = (
            LeadStatusHistory.objects
            .filter(
                lead__in=leads_qs,
                field_changed='telecaller_status',
                new_value='warm',
                created_at__date__gte=date_from,
                created_at__date__lte=date_to,
            )
            .annotate(day=TruncDate('created_at'))
            .values('day')
            .annotate(count=Count('id'))
            .order_by('day')
        )

        # Closures per day (by closure_date) — for the STM/CP reports charts.
        cl_qs = scope_to_company(Closure.objects.all(), request.user, 'company')
        if not _sees_all_company(request.user, request):
            ids = _visible_user_ids(request.user)
            cl_qs = cl_qs.filter(Q(stm__in=ids) | Q(referred_by_telecaller__in=ids))
        if company_id and is_platform_admin(request.user):
            cl_qs = cl_qs.filter(company_id=company_id)
        # booking/total amounts are EncryptedDecimalField (cannot Sum() in the DB),
        # so aggregate count + amount per day in Python for the closures chart tooltip.
        closure_map = {}
        for c in (cl_qs
                  .filter(closure_date__gte=date_from, closure_date__lte=date_to)
                  .only('closure_date', 'total_amount', 'booking_amount')):
            key = str(c.closure_date)
            amt = c.total_amount or c.booking_amount or 0
            entry = closure_map.setdefault(key, {'count': 0, 'amount': 0.0})
            entry['count'] += 1
            entry['amount'] += float(amt)
        closures_ser = [
            {'date': k, 'count': v['count'], 'amount': v['amount']}
            for k, v in sorted(closure_map.items())
        ]

        # STM pipeline trends — count by the day the lead's stm_status changed to
        # hot/warm/cold (status-history), for the STM/CP dashboard & reports charts.
        def _stm_status_trend(val):
            return (
                LeadStatusHistory.objects
                .filter(lead__in=leads_qs, field_changed='stm_status', new_value=val,
                        created_at__date__gte=date_from, created_at__date__lte=date_to)
                .annotate(day=TruncDate('created_at'))
                .values('day').annotate(count=Count('id')).order_by('day')
            )
        stm_hot_rows  = _stm_status_trend('hot')
        stm_warm_rows = _stm_status_trend('warm')
        stm_cold_rows = _stm_status_trend('cold')

        def _ser(rows):
            return [{'date': str(r['day']), 'count': r['count']} for r in rows]

        return Response({
            'mql':      _ser(mql_rows),
            'sv':       _ser(sv_rows),
            'warm':     _ser(warm_rows),
            'closures': closures_ser,
            'stm_hot':  _ser(stm_hot_rows),
            'stm_warm': _ser(stm_warm_rows),
            'stm_cold': _ser(stm_cold_rows),
            'date_from': date_from,
            'date_to':   date_to,
        })


def _leads_by_status_transition(leads_scope, field, value, date_from, date_to):
    """Lead ids whose `field` changed to `value` within [date_from, date_to] —
    mirrors StatsView's transition-count logic so a dashboard tile's date range
    and a click-through list filter mean the same thing."""
    hist = LeadStatusHistory.objects.filter(
        lead__in=leads_scope, field_changed=field, new_value=value)
    if date_from:
        hist = hist.filter(created_at__date__gte=date_from)
    if date_to:
        hist = hist.filter(created_at__date__lte=date_to)
    return set(hist.values_list('lead_id', flat=True))


def _leads_by_sv_outcome(leads_scope, value, date_from, date_to):
    """Lead ids still at stm_status='sv_done' whose latest completed visit outcome
    matches `value`, dated by when that visit happened — mirrors StatsView's
    sv-outcome union for the effective hot/warm/cold count."""
    latest_sv = SiteVisit.objects.filter(
        lead=OuterRef('pk'), status='completed',
    ).order_by('-visited_at')
    qs = leads_scope.filter(stm_status='sv_done').annotate(
        _sv_outcome=Subquery(latest_sv.values('outcome')[:1]),
        _sv_visited=Subquery(latest_sv.values('visited_at')[:1]),
    ).filter(_sv_outcome=value)
    if date_from:
        qs = qs.filter(_sv_visited__date__gte=date_from)
    if date_to:
        qs = qs.filter(_sv_visited__date__lte=date_to)
    return set(qs.values_list('id', flat=True))


class LeadListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Defer heavy text blobs not needed for list view
        qs = scope_to_company(
            Lead.objects.select_related('project', 'source', 'telecaller', 'stm'),
            request.user,
        ).defer(
            'telecaller_remarks', 'stm_remarks', 'requirement',
            'preferred_location', 'budget_min', 'budget_max',
        )

        # Channel Partner leads are their own module — the main Sales "All Leads"
        # never shows them, only the Channel Partner section itself does (via
        # cp_only/channel_partner_id below). Keeps the two pools disjoint in both
        # directions: a CP-sourced lead never lands in a regular Sales view, and a
        # regular lead never leaks into the CP one. See cp_lead_q for what counts
        # as a CP lead (channel_partner FK set OR Source explicitly "Channel Partner").
        if not (request.query_params.get('cp_only') == 'true' or request.query_params.get('channel_partner_id')):
            qs = qs.exclude(cp_lead_q())

        # The visit's Hot/Warm/Cold outcome (most recent completed visit) — shown
        # alongside "sv done" so the list reads e.g. "SV Done · Hot" instead of
        # just the generic stage, without overwriting stm_status itself.
        latest_completed_sv_outcome = SiteVisit.objects.filter(
            lead=OuterRef('pk'), status='completed',
        ).order_by('-visited_at').values('outcome')[:1]
        qs = qs.annotate(sv_outcome=Subquery(latest_completed_sv_outcome))

        # Telecallers / STMs only see leads assigned to them.
        qs = scope_leads_to_role(qs, request.user, request=request)

        # Filters
        search = request.query_params.get('search', '').strip()
        if search:
            # name/phone/email are encrypted, so the database can't match on them.
            # A full 10-digit number goes through the blind index (indexed, exact).
            # Anything else -- a partial number, a name fragment -- is matched in
            # Python over the already company/role-scoped rows, then fed back as an
            # id filter so every downstream filter, sort and page still works.
            # Measured at ~8us per decrypted value; a 11k-lead company costs ~250ms
            # and only on an explicit search.
            digits = ''.join(c for c in search if c.isdigit())
            if len(digits) >= 10 and not any(c.isalpha() for c in search):
                qs = qs.filter(phone_key=phone_blind_index(digits))
            else:
                needle = search.lower()
                hits = [
                    pk for pk, nm, ph, em in qs.values_list('id', 'name', 'phone', 'email')
                    if needle in (nm or '').lower()
                    or needle in (ph or '').lower()
                    or needle in (em or '').lower()
                ]
                qs = qs.filter(id__in=hits)

        if request.query_params.get('status'):
            qs = qs.filter(status=request.query_params['status'])
        date_from_param = request.query_params.get('date_from')
        date_to_param = request.query_params.get('date_to')
        # A dashboard tile like "Warm/SQL" counts leads that BECAME that status
        # within the selected date range (see StatsView), not leads currently
        # sitting at that status — so its click-through must filter the same way,
        # or a lead received earlier but marked warm today would be missing from
        # one and present in the other. Only kicks in when a date range is
        # actually given; a bare status filter (no date) still means "currently
        # this status", which is its own useful, unrelated view.
        skip_created_at_filter = False
        telecaller_status_filter = request.query_params.get('telecaller_status')
        TRANSITION_TC_STATUSES = {'hot', 'warm', 'callback', 'not_reachable', 'cold'}
        if telecaller_status_filter:
            if (date_from_param or date_to_param) and telecaller_status_filter in TRANSITION_TC_STATUSES:
                ids = _leads_by_status_transition(
                    qs, 'telecaller_status', telecaller_status_filter, date_from_param, date_to_param)
                qs = qs.filter(id__in=ids)
                skip_created_at_filter = True
            else:
                qs = qs.filter(telecaller_status=telecaller_status_filter)
        stm_status_filter = request.query_params.get('stm_status')
        if stm_status_filter:
            if stm_status_filter in ('hot', 'warm', 'cold'):
                if date_from_param or date_to_param:
                    direct_ids = _leads_by_status_transition(
                        qs, 'stm_status', stm_status_filter, date_from_param, date_to_param)
                    sv_ids = _leads_by_sv_outcome(qs, stm_status_filter, date_from_param, date_to_param)
                    qs = qs.filter(id__in=(direct_ids | sv_ids))
                    skip_created_at_filter = True
                else:
                    # A lead still sitting at "sv done" whose visit outcome was Hot
                    # counts as Hot too — the outcome doesn't overwrite stm_status,
                    # but it should still surface here alongside leads reclassified
                    # to hot/warm/cold directly.
                    qs = qs.filter(
                        Q(stm_status=stm_status_filter)
                        | Q(stm_status='sv_done', sv_outcome=stm_status_filter)
                    )
            elif stm_status_filter == 'sv_scheduled' and (date_from_param or date_to_param):
                ids = _leads_by_status_transition(
                    qs, 'stm_status', stm_status_filter, date_from_param, date_to_param)
                qs = qs.filter(id__in=ids)
                skip_created_at_filter = True
            else:
                qs = qs.filter(stm_status=stm_status_filter)
        project_id = request.query_params.get('project_id')
        if project_id == 'none':
            qs = qs.filter(project__isnull=True)   # unmapped leads (no project)
        elif project_id:
            qs = qs.filter(project_id=project_id)
        if request.query_params.get('source_id'):
            qs = qs.filter(source_id=request.query_params['source_id'])
        if request.query_params.get('channel_partner_id'):
            qs = qs.filter(channel_partner_id=request.query_params['channel_partner_id'])
        elif request.query_params.get('cp_only') == 'true':
            # The Channel Partner section's "CP Leads" tab — every lead referred
            # by any channel partner, OR added via the regular Sales flow with
            # Source set to "Channel Partner" (see cp_lead_q).
            qs = qs.filter(cp_lead_q())
        if request.query_params.get('telecaller_id'):
            qs = qs.filter(telecaller_id=request.query_params['telecaller_id'])
        if request.query_params.get('stm_id'):
            qs = qs.filter(stm_id=request.query_params['stm_id'])
        if request.query_params.get('is_duplicate') == 'true':
            qs = qs.filter(is_duplicate=True)
        if not skip_created_at_filter:
            if date_from_param:
                qs = qs.filter(created_at__date__gte=date_from_param)
            if date_to_param:
                qs = qs.filter(created_at__date__lte=date_to_param)
        if request.query_params.get('campaign'):
            qs = qs.filter(meta_campaign_name__icontains=request.query_params['campaign'])

        # Platform admin: filter by a specific company (used by admin company picker)
        if request.query_params.get('company_id') and is_platform_admin(request.user):
            qs = qs.filter(company_id=request.query_params['company_id'])

        # Work split for telecaller / STM portals: separate the leads they still have
        # to call ('pending') from the ones they've already actioned ('called'),
        # keyed off their own status field. Admins/managers fall back to overall status.
        work = request.query_params.get('work')
        if work == 'pending':
            if is_telecaller(request.user):
                qs = qs.filter(telecaller_status='')
            elif is_stm(request.user) or is_cp(request.user):
                qs = qs.filter(stm_status='')
            else:
                qs = qs.filter(status='new')
        elif work == 'called':
            if is_telecaller(request.user):
                qs = qs.exclude(telecaller_status='')
            elif is_stm(request.user) or is_cp(request.user):
                qs = qs.exclude(stm_status='')
            else:
                qs = qs.exclude(status='new')

        # Optional ordering override (default is newest-first from the model Meta).
        # 'pending' lists use oldest-first (FIFO) so fresh leads queue at the bottom
        # and never push down the lead currently being worked.
        ordering = request.query_params.get('ordering')
        if ordering in ('created_at', '-created_at', 'updated_at', '-updated_at', 'stm_assigned_at', '-stm_assigned_at'):
            qs = qs.order_by(ordering)

        total = qs.count()
        page = int(request.query_params.get('page', 1))
        offset = (page - 1) * PAGE_SIZE
        leads = qs[offset: offset + PAGE_SIZE]

        return Response({
            'count': total,
            'results': LeadListSerializer(leads, many=True).data,
        })

    def post(self, request):
        # Any authenticated Sales user (incl. telecallers) may add a lead.
        # Consistent with PATCH (lead update), which has no admin/manager gate.
        # Only admins/managers may assign a telecaller/STM on create; strip those
        # fields for callers (they self-source) so they can't assign to others.
        data = {k: v for k, v in request.data.items()}
        can_assign = can_assign_leads(request.user)
        if not can_assign:
            data.pop('telecaller', None)
            data.pop('stm', None)
        ser = LeadCreateSerializer(data=data)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)

        company = request.user.company

        # If a project is supplied it must belong to the requester's company.
        project = ser.validated_data.get('project')
        if project and not _project_in_scope(request, project.id):
            return Response({'detail': 'Invalid project for your company.'}, status=status.HTTP_400_BAD_REQUEST)

        # Duplicate check — match last 10 digits regardless of +91 prefix. Scoped to
        # the creator's own bucket (telecaller→their leads, STM/CP→their leads) so a
        # CP's lead is only a duplicate of another CP lead, not of someone else's.
        # Admins/managers keep the company-wide check.
        phone = ser.validated_data['phone']
        clean = ''.join(c for c in phone if c.isdigit())[-10:]
        dup_qs = (
            scope_leads_to_role(scope_to_company(Lead.objects.all(), request.user), request.user)
            .filter(phone_key=phone_blind_index(clean))
            if clean else Lead.objects.none()
        )
        existing = dup_qs.first()

        # Self-sourced (manually added) leads are assigned to their creator so they
        # land in that person's pipeline, and are marked actioned (status defaults to
        # 'warm' if none given) so they appear in the "Called" bucket, not "To Call" —
        # the creator already has the contact, there's nothing to call fresh.
        extra = {}
        if not can_assign:
            # Callers self-source: own the lead + mark actioned → their "Called" bucket.
            if is_cp(request.user) or is_stm(request.user):
                extra['stm'] = request.user
                if not ser.validated_data.get('stm_status'):
                    extra['stm_status'] = 'warm'
            elif is_telecaller(request.user):
                extra['telecaller'] = request.user
                if not ser.validated_data.get('telecaller_status'):
                    # 'callback' (not 'warm') so a blank status lands in the telecaller's
                    # "Called" bucket WITHOUT auto-transferring — 'warm' is a deliberate
                    # transfer-to-STM action handled below.
                    extra['telecaller_status'] = 'callback'
        else:
            # Admin/manager assigned via the form → stamp assignment time. Status is
            # left empty so the lead lands in the assignee's "To Call" bucket.
            if ser.validated_data.get('telecaller'):
                extra['telecaller_assigned_at'] = timezone.now()
            if ser.validated_data.get('stm'):
                extra['stm_assigned_at'] = timezone.now()

        lead = ser.save(
            company=company,
            is_duplicate=bool(existing),
            duplicate_of=existing if existing else None,
            **extra,
        )
        if existing:
            existing.duplicate_count += 1
            existing.save(update_fields=['duplicate_count'])

        # Optional backdate — "when did this lead actually come in" (e.g. a walk-in
        # logged a day later). created_at is auto_now_add, so it can't be set via the
        # serializer; overwrite it directly afterward, same as the bulk importer does.
        lead_date = _imp_dt(data.get('lead_date'))
        if lead_date:
            lead.created_at = lead_date
            lead.save(update_fields=['created_at'])

        _record_lead_created(lead, by=request.user)
        # A telecaller/STM/admin can set an initial TC/STM Status in the SAME request
        # that creates the lead (e.g. a telecaller working a live WhatsApp chat calls
        # it warm right away instead of adding it blank and PATCHing afterward). The
        # PATCH handler logs every status transition to LeadStatusHistory — this path
        # only ever logged 'created' (+ 'warm_transfer' below), never the underlying
        # telecaller_status/stm_status/status transitions themselves, so a lead that
        # went warm at creation was invisible to every date-ranged status filter and
        # dashboard tile (all of them read LeadStatusHistory, not the live field).
        old_status = lead.status
        creation_history = []
        if lead.telecaller_status:
            creation_history.append(LeadStatusHistory(
                lead=lead, changed_by=request.user,
                field_changed='telecaller_status', old_value='', new_value=lead.telecaller_status,
            ))
        if lead.stm_status:
            creation_history.append(LeadStatusHistory(
                lead=lead, changed_by=request.user,
                field_changed='stm_status', old_value='', new_value=lead.stm_status,
            ))
        if creation_history:
            LeadStatusHistory.objects.bulk_create(creation_history)
        # Notify the assignee when an admin/manager hand-picks them on create.
        if can_assign:
            from notifications import notify
            if lead.telecaller_id:
                notify(lead.telecaller, 'new_lead', 'New Lead Assigned',
                       f'{lead.name} has been assigned to you.', {'lead_id': lead.id})
            if lead.stm_id:
                notify(lead.stm, 'new_lead', 'New Lead Assigned',
                       f'{lead.name} has been assigned to you.', {'lead_id': lead.id})
        # Telecaller marked the new lead "warm" → warm-transfer into the STM pipeline
        # (mirrors the PATCH behaviour): overall status = warm_transferred, then
        # auto-assign an STM. Applies whether warm came from a caller or an admin form.
        if lead.telecaller_status == 'warm' and lead.status != 'warm_transferred':
            lead.status = 'warm_transferred'
            lead.save(update_fields=['status'])
            LeadStatusHistory.objects.create(
                lead=lead, changed_by=request.user,
                field_changed='status', old_value=old_status, new_value=lead.status,
            )
            LeadStatusHistory.objects.create(
                lead=lead, changed_by=request.user,
                field_changed='warm_transfer', old_value='', new_value='Transferred to STM',
            )
        # A lead that starts directly in the STM pipeline (self-sourced by an STM/CP,
        # or given an initial STM Status on the create form — e.g. a Channel Partner
        # lead) has Overall mirror STM Status immediately, same as every later PATCH
        # already does (see LeadDetailView.patch).
        if lead.stm_status and lead.status != lead.stm_status:
            old_status = lead.status
            lead.status = lead.stm_status
            lead.save(update_fields=['status'])
            LeadStatusHistory.objects.create(
                lead=lead, changed_by=request.user,
                field_changed='status', old_value=old_status, new_value=lead.status,
            )
        if lead.status == 'warm_transferred' and lead.stm_id is None:
            _run_distribution(lead.company, 'stm')
        # Auto-distribute to a telecaller only when the lead is still unassigned
        # (admin didn't pick one and it isn't self-sourced / warm-transferred).
        elif not lead.telecaller_id and not lead.stm_id:
            _run_distribution(company, 'telecaller')
        lead.refresh_from_db()

        return Response(LeadDetailSerializer(lead).data, status=status.HTTP_201_CREATED)


class LeadDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_lead(self, request, pk):
        try:
            qs = scope_to_company(
                Lead.objects.select_related('project', 'source', 'telecaller', 'stm'),
                request.user,
            )
            # Telecallers / STMs can only open leads assigned to them.
            qs = scope_leads_to_role(qs, request.user, request=request)
            return qs.get(pk=pk)
        except Lead.DoesNotExist:
            return None

    def get(self, request, pk):
        lead = self._get_lead(request, pk)
        if not lead:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        data = LeadDetailSerializer(lead).data
        # Most recent 30 events, returned oldest→newest so the timeline reads in order.
        # Tie-break by id keeps same-second events in their logical creation order
        # (e.g. status change → warm transfer → STM assigned).
        recent = list(lead.history.order_by('-created_at', '-id')[:30])
        recent.reverse()
        data['history'] = LeadStatusHistorySerializer(recent, many=True).data
        data['follow_ups'] = FollowUpSerializer(lead.follow_ups.all(), many=True).data
        data['site_visits'] = SiteVisitSerializer(lead.site_visits.all(), many=True).data
        return Response(data)

    def patch(self, request, pk):
        lead = self._get_lead(request, pk)
        if not lead:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        old_status       = lead.status
        old_tc_status    = lead.telecaller_status
        old_stm_status   = lead.stm_status
        old_tc_id        = lead.telecaller_id
        old_stm_id       = lead.stm_id
        old_tc_name      = lead.telecaller.name if lead.telecaller else ''
        old_stm_name     = lead.stm.name        if lead.stm        else ''
        old_tc_remarks   = lead.telecaller_remarks
        old_stm_remarks  = lead.stm_remarks

        # Field-level write restrictions (mirrors the portal UI):
        #  - Telecallers may only write telecaller (TC) fields.
        #  - STMs may only write STM fields.
        #  - Neither may (re)assign leads. Admins/managers/Sales CRM may edit everything.
        data = {k: v for k, v in request.data.items()}
        if not can_assign_leads(request.user):
            for f in ('telecaller', 'stm'):
                data.pop(f, None)
        if is_telecaller(request.user):
            for f in ('stm', 'stm_status', 'stm_remarks'):
                data.pop(f, None)
        elif is_stm(request.user):
            for f in ('telecaller', 'telecaller_status', 'telecaller_remarks'):
                data.pop(f, None)

        ser = LeadUpdateSerializer(lead, data=data, partial=True)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        lead = ser.save()

        # A lead is "warm" when EITHER the telecaller sets TC Status = warm OR the
        # overall status is set to warm_transferred. Keep both in sync so the TC Status
        # column always shows 'warm' and Overall always shows 'warm_transferred',
        # then hand the lead to the STM pipeline. (TC's warm ≠ STM status — stm_status
        # stays blank.)
        warm_now = (
            (old_tc_status != 'warm' and lead.telecaller_status == 'warm') or
            (old_status != 'warm_transferred' and lead.status == 'warm_transferred')
        )
        if warm_now:
            sync = []
            if lead.status != 'warm_transferred':
                lead.status = 'warm_transferred'; sync.append('status')
            if lead.telecaller_status != 'warm':
                lead.telecaller_status = 'warm'; sync.append('telecaller_status')
            if sync:
                lead.save(update_fields=sync)

        # Once the lead is with sales, the Overall Status mirrors the STM's status
        # exactly (assigned → on TC assignment; warm_transferred → on TC warm; then
        # whatever the STM sets — cold, sv_scheduled, sv_done, closed, …).
        if lead.stm_status and old_stm_status != lead.stm_status:
            if lead.status != lead.stm_status:
                lead.status = lead.stm_status
                lead.save(update_fields=['status'])

        history_entries = []
        if old_status != lead.status:
            history_entries.append(LeadStatusHistory(
                lead=lead, changed_by=request.user,
                field_changed='status', old_value=old_status, new_value=lead.status,
            ))
        if old_tc_status != lead.telecaller_status:
            history_entries.append(LeadStatusHistory(
                lead=lead, changed_by=request.user,
                field_changed='telecaller_status', old_value=old_tc_status, new_value=lead.telecaller_status,
            ))
        if old_stm_status != lead.stm_status:
            history_entries.append(LeadStatusHistory(
                lead=lead, changed_by=request.user,
                field_changed='stm_status', old_value=old_stm_status, new_value=lead.stm_status,
            ))
        # Remarks are free text, not a status transition — logged so the STM (or anyone
        # else) can see exactly what the telecaller wrote and when, once the lead is
        # transferred to them. Same for STM's own remarks, for symmetry.
        # new_value is capped at 100 chars in the DB, but remarks can run much longer —
        # the full text goes in `remarks` (a TextField), new_value just holds a preview.
        if old_tc_remarks != lead.telecaller_remarks and lead.telecaller_remarks:
            history_entries.append(LeadStatusHistory(
                lead=lead, changed_by=request.user, field_changed='telecaller_remarks',
                old_value='', new_value=lead.telecaller_remarks[:100], remarks=lead.telecaller_remarks,
            ))
        if old_stm_remarks != lead.stm_remarks and lead.stm_remarks:
            history_entries.append(LeadStatusHistory(
                lead=lead, changed_by=request.user, field_changed='stm_remarks',
                old_value='', new_value=lead.stm_remarks[:100], remarks=lead.stm_remarks,
            ))
        if old_tc_id != lead.telecaller_id:
            new_tc_name = lead.telecaller.name if lead.telecaller else ''
            history_entries.append(LeadStatusHistory(
                lead=lead, changed_by=request.user,
                field_changed='telecaller', old_value=old_tc_name, new_value=new_tc_name,
            ))
            if lead.telecaller:
                from notifications import notify
                notify(lead.telecaller, 'new_lead', 'New Lead Assigned',
                       f'{lead.name} has been assigned to you.', {'lead_id': lead.id})
        if old_stm_id != lead.stm_id:
            new_stm_name = lead.stm.name if lead.stm else ''
            history_entries.append(LeadStatusHistory(
                lead=lead, changed_by=request.user,
                field_changed='stm', old_value=old_stm_name, new_value=new_stm_name,
            ))
            if lead.stm:
                from notifications import notify
                notify(lead.stm, 'new_lead', 'New Lead Assigned',
                       f'{lead.name} has been assigned to you.', {'lead_id': lead.id})
        if warm_now:
            history_entries.append(LeadStatusHistory(
                lead=lead, changed_by=request.user,
                field_changed='warm_transfer', old_value='', new_value='Transferred to STM',
            ))
        if history_entries:
            LeadStatusHistory.objects.bulk_create(history_entries)

        # Auto-assign whenever the lead is in the warm bucket and has no STM yet —
        # whether it got there via TC Status = warm OR by setting Overall Status
        # to 'warm_transferred' directly. Window-gated; no-op if no STM available.
        if lead.status == 'warm_transferred' and lead.stm_id is None:
            _run_distribution(lead.company, 'stm')
            lead.refresh_from_db()

        return Response(LeadDetailSerializer(lead).data)

    def delete(self, request, pk):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        lead = self._get_lead(request, pk)
        if not lead:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        lead.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BulkDeleteLeadsView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'No IDs provided.'}, status=status.HTTP_400_BAD_REQUEST)
        deleted, _ = scope_to_company(Lead.objects.filter(id__in=ids), request.user).delete()
        return Response({'deleted': deleted})


def _sync_plots(project):
    existing_count = project.plots.count()
    target = project.total_plots or 0
    # Only auto-create numbered plots if NO plots exist yet.
    # This prevents re-triggering on PATCH (e.g. after bulk typed-plot creation).
    if target > 0 and existing_count == 0:
        Plot.objects.bulk_create([
            Plot(project=project, number=str(i))
            for i in range(1, target + 1)
        ])


class ProjectListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        projects = scope_to_company(
            Project.objects.annotate(lead_count=Count('leads')).prefetch_related('plots'),
            request.user,
        )
        if request.query_params.get('active_only') == 'true':
            projects = projects.filter(is_active=True)
        if request.query_params.get('company_id') and is_platform_admin(request.user):
            projects = projects.filter(company_id=request.query_params['company_id'])
        return Response(ProjectSerializer(projects, many=True).data)

    def post(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        ser = ProjectSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        project = ser.save(company=request.user.company)
        _sync_plots(project)
        project = Project.objects.annotate(lead_count=Count('leads')).prefetch_related('plots').get(pk=project.pk)
        return Response(ProjectSerializer(project).data, status=status.HTTP_201_CREATED)


class ProjectDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            project = scope_to_company(
                Project.objects.annotate(lead_count=Count('leads')).prefetch_related('plots'),
                request.user,
            ).get(pk=pk)
        except Project.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(ProjectSerializer(project).data)

    def patch(self, request, pk):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        # Who approves a project's bookings is an administrative setting, not a field any
        # Manager may edit — otherwise an approver restricted to one project could simply
        # add themselves to another and approve it. Mirrors the same gate the Approver
        # Setup panel uses on the client (role Admin / staff / Sales admin-module).
        if (('booking_approvers' in request.data or 'cp_booking_approvers' in request.data) and not (
            _is_hard_admin(request.user) or 'Sales' in (getattr(request.user, 'admin_modules', None) or [])
        )):
            return Response(
                {'detail': 'Only an administrator can change booking approvers.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        try:
            project = scope_to_company(Project.objects.all(), request.user).get(pk=pk)
        except Project.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        ser = ProjectSerializer(project, data=request.data, partial=True)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        project = ser.save()
        # _sync_plots intentionally NOT called on PATCH — plots are managed via /plots/bulk/
        project = Project.objects.annotate(lead_count=Count('leads')).prefetch_related('plots').get(pk=project.pk)
        return Response(ProjectSerializer(project).data)

    def delete(self, request, pk):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            project = scope_to_company(Project.objects.all(), request.user).get(pk=pk)
        except Project.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        project.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


PLOT_HOLD_TIMEOUT = timedelta(minutes=10)


def _release_expired_holds(plots_qs):
    """Self-healing: flip stale soft-holds (a rep selected the unit on the picker but
    never submitted) back to available before reading. Only touches held_by-tracked
    holds — never a hard hold backed by a real pending Booking (held_by is cleared at
    submission time), and never an admin's manual hold via PlotDetailView.patch (which
    never sets held_by). A hold pinned by a saved draft is also exempt — the rep is
    still mid-way through the form, not just browsing; it only frees on submit,
    discard, or an explicit release."""
    cutoff = timezone.now() - PLOT_HOLD_TIMEOUT
    candidates = list(plots_qs.filter(status='hold', held_by__isnull=False, held_at__lt=cutoff)
                               .values_list('id', 'project_id'))
    if not candidates:
        return
    candidate_ids = {pid for pid, _ in candidates}
    project_ids   = {proj for _, proj in candidates}
    pinned = set()
    for b in Booking.objects.filter(status='draft', project_id__in=project_ids).only('plot_id', 'plot_ids'):
        if b.plot_id:
            pinned.add(b.plot_id)
        pinned.update(b.plot_ids or [])
    to_expire = candidate_ids - pinned
    if to_expire:
        Plot.objects.filter(id__in=to_expire).update(status='available', held_by=None, held_at=None)


class PlotListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        project_id = request.query_params.get('project')
        if not project_id or not str(project_id).isdigit():
            return Response({'detail': 'A valid numeric project query param is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not _project_in_scope(request, project_id):
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        _release_expired_holds(Plot.objects.filter(project_id=project_id))
        plots = Plot.objects.filter(project_id=project_id)
        return Response(PlotSerializer(plots, many=True).data)


class PlotDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            plot = scope_to_company(Plot.objects.all(), request.user, 'project__company').get(pk=pk)
        except Plot.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        ser = PlotSerializer(plot, data=request.data, partial=True)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        return Response(PlotSerializer(ser.save()).data)


class PlotHoldView(APIView):
    """A rep selecting units on the plot map — soft-reserve them immediately so no
    other rep can also select the same unit while this one is getting an LOI signed.
    Self-releases after PLOT_HOLD_TIMEOUT if never submitted (see _release_expired_holds)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ids = [int(x) for x in (request.data.get('plot_ids') or []) if str(x).isdigit()]
        held, failed = [], []
        with transaction.atomic():
            for pid in ids:
                try:
                    plot = scope_to_company(Plot.objects.select_for_update(), request.user, 'project__company').get(pk=pid)
                except Plot.DoesNotExist:
                    failed.append({'id': pid, 'reason': 'not_found'})
                    continue
                _release_expired_holds(Plot.objects.filter(pk=pid))
                plot.refresh_from_db()
                if plot.status != 'available':
                    reason = 'held_by_other' if (plot.held_by_id and plot.held_by_id != request.user.id) else plot.status
                    failed.append({'id': pid, 'number': plot.number, 'reason': reason})
                    continue
                plot.status, plot.held_by, plot.held_at = 'hold', request.user, timezone.now()
                plot.save(update_fields=['status', 'held_by', 'held_at'])
                held.append(pid)
        return Response({'held': held, 'failed': failed})


class PlotReleaseView(APIView):
    """Release units this rep soft-held but didn't end up booking (deselected, hit
    Clear, or picked something else). No-ops on plots not held by this user — in
    particular a plot that's since become a real booking's hard hold (held_by cleared
    at submission) is silently skipped rather than accidentally freed."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ids = [int(x) for x in (request.data.get('plot_ids') or []) if str(x).isdigit()]
        n = Plot.objects.filter(pk__in=ids, held_by=request.user, status='hold') \
                         .update(status='available', held_by=None, held_at=None)
        return Response({'released': n})


class LeadSourceListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        sources = scope_to_company(LeadSource.objects.filter(is_active=True), request.user)
        if request.query_params.get('company_id') and is_platform_admin(request.user):
            sources = sources.filter(company_id=request.query_params['company_id'])
        return Response(LeadSourceSerializer(sources, many=True).data)

    def post(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        ser = LeadSourceSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        return Response(LeadSourceSerializer(ser.save(company=request.user.company)).data, status=status.HTTP_201_CREATED)


class BackfillDuplicatesView(APIView):
    """One-time endpoint to mark existing duplicate leads based on last 10 phone digits."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=403)
        from collections import defaultdict
        # Stream rows with .iterator() so the whole Lead table is never materialised in
        # memory at once (prevents OOM on large tenants). Only id/phone are accumulated.
        leads = (
            scope_to_company(Lead.objects.all(), request.user)
            .only('id', 'phone', 'created_at')
            .order_by('created_at')
            .iterator(chunk_size=2000)
        )
        phone_map = defaultdict(list)
        for l in leads:
            clean = ''.join(c for c in (l.phone or '') if c.isdigit())[-10:]
            if clean:
                phone_map[clean].append(l.id)
        marked = 0
        for clean, ids in phone_map.items():
            if len(ids) > 1:
                original_id = ids[0]
                dup_ids = ids[1:]
                Lead.objects.filter(id__in=dup_ids).update(is_duplicate=True, duplicate_of_id=original_id)
                Lead.objects.filter(id=original_id).update(duplicate_count=len(dup_ids))
                marked += len(dup_ids)
        return Response({'marked_duplicates': marked})


class LeadSourceDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            source = scope_to_company(LeadSource.objects.all(), request.user).get(pk=pk)
        except LeadSource.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        source.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ChannelPartnerListCreateView(APIView):
    """Admin-only directory of external referral partners (CP Details) — distinct
    from a 'CP Executive' employee, who manages the relationship with these."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Read access is open to anyone with Sales access (not just admins/CP
        # managers) — an STM booking a unit through the regular Sales module
        # needs this list too, to pick a Channel Partner Name from the directory
        # rather than typing it freehand. Mirrors LeadSourceListView.get, which
        # is similarly unrestricted; only creating/editing/deleting entries
        # (below) stays admin/CP-manager only.
        qs = scope_to_company(ChannelPartner.objects.all(), request.user).annotate(lead_count=Count('leads'))
        if request.query_params.get('company_id') and is_platform_admin(request.user):
            qs = qs.filter(company_id=request.query_params['company_id'])
        if request.query_params.get('category'):
            qs = qs.filter(category=request.query_params['category'])
        search = request.query_params.get('search', '').strip()
        if search:
            # name/contact_no/firm_name — matched in Python like Lead's search,
            # since name/contact_no are encrypted and can't be filtered in SQL.
            needle = search.lower()
            hits = [
                pk for pk, nm, ph, firm in qs.values_list('id', 'name', 'contact_no', 'firm_name')
                if needle in (nm or '').lower() or needle in (ph or '').lower() or needle in (firm or '').lower()
            ]
            qs = qs.filter(id__in=hits)
        return Response(ChannelPartnerSerializer(qs, many=True).data)

    def post(self, request):
        if not can_access_cp_module(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        ser = ChannelPartnerSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        cp = ser.save(company=_resolve_company(request), created_by=request.user)
        # Optional backdate — "when did this partnership actually start" — same
        # override-after-create trick as a Lead's lead_date (created_at is
        # auto_now_add, so it can't be set via the serializer).
        date_added = _imp_dt(request.data.get('date_added'))
        if date_added:
            cp.created_at = date_added
            cp.save(update_fields=['created_at'])
        return Response(ChannelPartnerSerializer(cp).data, status=status.HTTP_201_CREATED)


class ChannelPartnerDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        if not can_access_cp_module(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            cp = scope_to_company(ChannelPartner.objects.all(), request.user).get(pk=pk)
        except ChannelPartner.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        ser = ChannelPartnerSerializer(cp, data=request.data, partial=True)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        ser.save()
        return Response(ser.data)

    def delete(self, request, pk):
        if not can_access_cp_module(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            cp = scope_to_company(ChannelPartner.objects.all(), request.user).get(pk=pk)
        except ChannelPartner.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        cp.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class FollowUpListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = scope_to_company(
            FollowUp.objects.select_related('lead', 'assigned_to'),
            request.user, 'lead__company',
        )
        if not _sees_all_company(request.user, request):
            qs = qs.filter(assigned_to__in=_visible_user_ids(request.user))
        else:
            # A follow-up has no project of its own — scope through its lead.
            qs = scope_leads_to_project(qs, request.user, 'lead__')
        if request.query_params.get('company_id') and is_platform_admin(request.user):
            qs = qs.filter(lead__company_id=request.query_params['company_id'])
        if request.query_params.get('lead_id'):
            qs = qs.filter(lead_id=request.query_params['lead_id'])
        if request.query_params.get('status'):
            qs = qs.filter(status=request.query_params['status'])
        if request.query_params.get('cp_only') == 'true':
            qs = qs.filter(cp_lead_q(prefix='lead__'))
        return Response(FollowUpSerializer(qs, many=True).data)

    def post(self, request):
        ser = FollowUpSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        if not _lead_in_scope(request, request.data.get('lead')):
            return Response({'detail': 'Invalid lead for your company.'}, status=status.HTTP_400_BAD_REQUEST)
        followup = ser.save(created_by=request.user)
        if followup.assigned_to and followup.assigned_to_id != request.user.id:
            from notifications import notify
            when = followup.scheduled_at.strftime('%d %b %I:%M %p') if followup.scheduled_at else ''
            notify(followup.assigned_to, 'followup', 'New Follow-Up',
                   (f'{followup.lead.name} · {when}').strip(' ·'),
                   {'lead_id': followup.lead_id, 'followup_id': followup.id})
        return Response(FollowUpSerializer(followup).data, status=status.HTTP_201_CREATED)


class FollowUpDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        try:
            followup = scope_to_company(FollowUp.objects.all(), request.user, 'lead__company').get(pk=pk)
        except FollowUp.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        ser = FollowUpSerializer(followup, data=request.data, partial=True)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        return Response(FollowUpSerializer(ser.save()).data)


class SiteVisitListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = scope_to_company(
            SiteVisit.objects.select_related('lead', 'project', 'stm'),
            request.user, 'lead__company',
        )
        if not _sees_all_company(request.user, request):
            _ids = _visible_user_ids(request.user)
            qs = qs.filter(Q(stm__in=_ids) | Q(referred_by_telecaller__in=_ids))
        else:
            # A manager assigned to specific projects sees only those projects' visits.
            qs = scope_leads_to_project(qs, request.user)
        # Platform admin viewing a specific company (?company_id) — honour the filter.
        # A site visit has no company of its own; it belongs to its lead's company, the
        # same path scope_to_company uses above. Filtering company_id directly raised
        # FieldError, so this endpoint 500'd for any platform admin with a company
        # selected — which is what left My Conversions showing 0 site visits.
        cid = request.query_params.get('company_id')
        if cid and is_platform_admin(request.user):
            qs = qs.filter(lead__company_id=cid)
        if request.query_params.get('lead_id'):
            qs = qs.filter(lead_id=request.query_params['lead_id'])
        if request.query_params.get('cp_only') == 'true':
            qs = qs.filter(cp_lead_q(prefix='lead__'))
        return Response(SiteVisitSerializer(qs, many=True).data)

    def post(self, request):
        ser = SiteVisitSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        if not _lead_in_scope(request, request.data.get('lead')):
            return Response({'detail': 'Invalid lead for your company.'}, status=status.HTTP_400_BAD_REQUEST)
        sv = ser.save()
        sched = sv.scheduled_at.strftime('%d %b %I:%M %p') if sv.scheduled_at else ''
        # A visit can be created already-completed (the sv_done fallback when no
        # scheduled visit exists yet) — label it as such, with the outcome, rather
        # than always saying "Scheduled" regardless of its actual status.
        if sv.status == 'completed':
            label = f'Completed · {sv.get_outcome_display()}' if sv.outcome else 'Completed'
        else:
            label = f'Scheduled · {sched}' if sched else 'Scheduled'
        LeadStatusHistory.objects.create(
            lead=sv.lead, changed_by=request.user, field_changed='site_visit',
            old_value='', new_value=label[:100],
            remarks='Site visit scheduled' if sv.status != 'completed' else 'Site visit completed',
        )
        from notifications import notify
        for who in (sv.stm, sv.referred_by_telecaller):
            if who and who.id != request.user.id:
                notify(who, 'sv', 'Site Visit Scheduled',
                       (f'{sv.lead.name} · {sched}').strip(' ·'),
                       {'lead_id': sv.lead_id, 'sv_id': sv.id})
        return Response(SiteVisitSerializer(sv).data, status=status.HTTP_201_CREATED)


class SiteVisitDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        try:
            sv = scope_to_company(SiteVisit.objects.all(), request.user, 'lead__company').get(pk=pk)
        except SiteVisit.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        old_status = sv.status
        # Marking a visit Done must record what came of it — an outcome and remarks,
        # not just a status flip — so the pipeline can tell an interested walk-in
        # from a dead one. Checked against the merged (existing + incoming) values so
        # a client that already set these on an earlier PATCH isn't forced to resend.
        if request.data.get('status') == 'completed' and old_status != 'completed':
            new_outcome = request.data.get('outcome', sv.outcome)
            new_remarks = request.data.get('remarks', sv.remarks)
            if not new_outcome or not str(new_remarks or '').strip():
                return Response(
                    {'detail': 'Outcome and remarks are required to mark a site visit as done.'},
                    status=status.HTTP_400_BAD_REQUEST)
        ser = SiteVisitSerializer(sv, data=request.data, partial=True)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        sv = ser.save()
        if sv.status != old_status:
            # Include the outcome in the logged transition — "Completed · Hot" —
            # so the lead's history timeline shows what came of the visit, not
            # just that it happened.
            new_value = sv.get_status_display()
            if sv.status == 'completed' and sv.outcome:
                new_value = f'Completed · {sv.get_outcome_display()}'
            LeadStatusHistory.objects.create(
                lead=sv.lead, changed_by=request.user, field_changed='site_visit',
                old_value=old_status, new_value=new_value,
                remarks='Site visit updated',
            )
            if sv.status == 'completed':
                # Telecaller who referred the lead + the STM both hear that the SV is done.
                from notifications import notify
                for who in (sv.referred_by_telecaller, sv.stm):
                    if who and who.id != request.user.id:
                        notify(who, 'sv_done', 'Site Visit Done',
                               f"{sv.lead.name}'s site visit is complete.",
                               {'lead_id': sv.lead_id, 'sv_id': sv.id})
        return Response(SiteVisitSerializer(sv).data)


class ClosureListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = scope_to_company(
            Closure.objects.select_related('lead', 'project', 'stm'),
            request.user, 'company',
        )
        if not _sees_all_company(request.user, request):
            _ids = _visible_user_ids(request.user)
            qs = qs.filter(Q(stm__in=_ids) | Q(referred_by_telecaller__in=_ids))
        else:
            # A manager assigned to specific projects sees only those projects' closures.
            qs = scope_leads_to_project(qs, request.user)
        # Platform admin viewing a specific company (?company_id) — honour the filter.
        cid = request.query_params.get('company_id')
        if cid and is_platform_admin(request.user):
            qs = qs.filter(company_id=cid)
        if request.query_params.get('cp_only') == 'true':
            qs = qs.filter(cp_lead_q(prefix='lead__'))
        return Response(ClosureSerializer(qs, many=True).data)

    def post(self, request):
        ser = ClosureSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        if not _lead_in_scope(request, request.data.get('lead')):
            return Response({'detail': 'Invalid lead for your company.'}, status=status.HTTP_400_BAD_REQUEST)
        closure = ser.save()
        parts = [closure.get_status_display()]
        unit = f'{closure.unit_type} {closure.unit_no}'.strip()
        if unit:
            parts.append(unit)
        if closure.total_amount:
            parts.append(f'₹{closure.total_amount:g}')
        LeadStatusHistory.objects.create(
            lead=closure.lead, changed_by=request.user, field_changed='closure',
            old_value='', new_value=' · '.join(parts)[:100], remarks='Closure recorded',
        )
        if closure.stm:
            from notifications import notify_many, reporting_chain
            notify_many(reporting_chain(closure.stm), 'closure', 'New Closure',
                        (f'{closure.stm.name} closed {closure.lead.name} · {unit}').strip(' ·'),
                        {'lead_id': closure.lead_id, 'closure_id': closure.id})
        return Response(ClosureSerializer(closure).data, status=status.HTTP_201_CREATED)


class TelecallerListView(APIView):
    """Users for lead assignment. Filters by User.designation icontains crm_role param.
    Falls back to all Sales-module users if no designation match found."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        crm_role = request.query_params.get('crm_role')
        cid      = request.query_params.get('company_id')
        if is_platform_admin(request.user):
            if cid:
                from companies.models import Company as Co
                co = Co.objects.filter(pk=cid).first()
                base_qs = User.objects.filter(company=co, is_active=True) if co else User.objects.none()
            else:
                base_qs = User.objects.filter(is_active=True)
        else:
            base_qs = User.objects.filter(company=request.user.company, is_active=True)
        sales_qs = base_qs.filter(modules__contains=['Sales']).order_by('name')

        if crm_role in ('telecaller', 'stm'):
            users = base_qs.filter(designation__icontains=crm_role).order_by('name')
            if not users.exists():
                users = sales_qs
        elif crm_role == 'cp':
            # CP executives (channel partners) — for CP managers assigning leads.
            users = base_qs.filter(
                Q(designation__icontains='cp executive') | Q(designation__icontains='channel partner')
            ).order_by('name')
            if not users.exists():
                users = sales_qs
        elif crm_role == 'cp_module':
            # Everyone with access to the Channel Partner module — admins/staff/
            # Sales Admin-Modules users, and CP-designation Managers (see
            # is_cp_manager) — for the "who owns this CP lead" filter. Not a
            # designation substring, so filtered in Python like the other
            # cross-cutting permission checks in this file.
            users = sorted(
                (u for u in base_qs if _is_sales_admin(u) or is_cp_manager(u)),
                key=lambda u: u.name or '',
            )
        elif crm_role == 'sales_cp':
            # Same-company employees with Sales module access — for the CP
            # module's "Assign STM" dropdown, letting a CP Cluster Head hand a
            # CP lead straight to a Sales-side person, no approval step.
            # Company-scoped like every other list here (base_qs already
            # honours the caller's own company, or ?company_id for a platform
            # admin) — no separate CP-access requirement on the assignee.
            users = sales_qs
            project_id = request.query_params.get('project_id')
            if project_id:
                # Strictly the STMs assigned to THIS project (Team Users →
                # Assign, the same UserProjectAssignment used by
                # manager_project_ids) — not the opt-in "unassigned = every
                # project" fallback used elsewhere, since most employees have
                # no assignment row at all and that made this list barely
                # narrow down. Only an explicit assignment gets someone listed.
                assigned_ids = set(
                    UserProjectAssignment.objects.filter(project_id=project_id).values_list('user_id', flat=True)
                )
                users = [u for u in users if u.id in assigned_ids]
        else:
            users = sales_qs

        data = [
            {'id': u.id, 'name': u.name, 'user_code': u.user_code, 'role': u.role, 'designation': u.designation}
            for u in users
        ]
        return Response(data)


class CompanyUsersSlimView(APIView):
    """Lightweight user list for Sales CRM — only fields the UI needs, no heavy JSONField serialization."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company_id = request.query_params.get('company_id')
        company = (
            __import__('companies.models', fromlist=['Company']).Company.objects.filter(pk=company_id).first()
            if company_id and is_platform_admin(request.user)
            else request.user.company
        )
        users = (
            User.objects
            .filter(company=company, is_active=True)
            .exclude(role='Admin')
            .only('id', 'name', 'user_code', 'designation', 'role', 'phone', 'email')
            .order_by('name')
        )
        data = [{
            'id':          u.id,
            'name':        u.name,
            'user_code':   u.user_code,
            'designation': u.designation,
            'role':        u.role,
            'phone':       u.phone,
            'email':       u.email,
        } for u in users]
        return Response(data)


# ── Sales Team Members ──────────────────────────────────────────────────────
# models already imported at top of file


class SalesTeamView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cid = request.query_params.get('company_id')
        if is_platform_admin(request.user):
            if cid:
                from companies.models import Company as Co
                company = Co.objects.filter(pk=cid).first()
                users = User.objects.filter(company=company, is_active=True, department__icontains='sales') if company else User.objects.none()
            else:
                users = User.objects.filter(is_active=True, department__icontains='sales')
        else:
            users = User.objects.filter(company=request.user.company, is_active=True, department__icontains='sales')
        users = users.order_by('name')

        data = [{
            'id':          u.id,
            'name':        u.name,
            'email':       u.email,
            'phone':       u.phone,
            'user_code':   u.user_code,
            'designation': u.designation,
            'role':        u.role,
        } for u in users]
        return Response(data)

    def post(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        user_id  = request.data.get('user_id')
        crm_role = request.data.get('crm_role', 'telecaller')
        try:
            user = User.objects.get(pk=user_id, company=request.user.company)
        except User.DoesNotExist:
            return Response({'detail': 'User not found in your company.'}, status=status.HTTP_404_NOT_FOUND)
        member, created = SalesTeamMember.objects.get_or_create(user=user, defaults={'crm_role': crm_role})
        if not created:
            member.crm_role  = crm_role
            member.is_active = True
            member.save()
        return Response({'id': member.id, 'user_id': user.id, 'name': user.name, 'crm_role': member.crm_role, 'designation': user.designation, 'user_code': user.user_code}, status=status.HTTP_201_CREATED)


class SalesTeamMemberDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            m = SalesTeamMember.objects.get(pk=pk, user__company=request.user.company)
        except SalesTeamMember.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if 'crm_role' in request.data:
            m.crm_role = request.data['crm_role']
        if 'is_active' in request.data:
            m.is_active = request.data['is_active']
        m.save()
        return Response({'id': m.id, 'crm_role': m.crm_role, 'is_active': m.is_active})

    def delete(self, request, pk):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            m = SalesTeamMember.objects.get(pk=pk, user__company=request.user.company)
        except SalesTeamMember.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        m.is_active = False
        m.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Distribution Settings ─────────────────────────────────────────────────────
class DistributionSettingsView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_or_create(self, company):
        obj, _ = DistributionSettings.objects.get_or_create(company=company)
        return obj

    def get(self, request):
        company = _resolve_company(request)
        s = self._get_or_create(company)
        managers = list(
            User.objects.filter(company=company, is_active=True, role__in=MANAGER_ROLES)
            .exclude(role='Admin')
            .order_by('name').values('id', 'name', 'designation')
        )
        return Response({
            'tc_signin_time':   str(s.tc_signin_time)[:5],
            'tc_signout_time':  str(s.tc_signout_time)[:5],
            'stm_signin_time':  str(s.stm_signin_time)[:5],
            'stm_signout_time': str(s.stm_signout_time)[:5],
            'managers': managers,   # for the per-project booking-approver picker
        })

    def put(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        s = self._get_or_create(_resolve_company(request))
        for field in ('tc_signin_time', 'tc_signout_time', 'stm_signin_time', 'stm_signout_time'):
            if field in request.data:
                setattr(s, field, request.data[field])
        s.save()
        return Response({'detail': 'Saved.'})


# ── Availability ──────────────────────────────────────────────────────────────
class AvailabilityView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from datetime import date as date_cls
        today = request.query_params.get('date', str(date_cls.today()))
        company = _resolve_company(request)
        desig_map = {'TELECALLER': 'telecaller', 'STM': 'stm'}
        users = (
            User.objects
            .filter(company=company, is_active=True)
            .exclude(role='Admin')
            .filter(designation__in=['TELECALLER', 'STM'])
            .only('id', 'name', 'designation')
            .order_by('name')
        )
        avail_map = {}
        checkin_map = {}
        for a in UserAvailability.objects.filter(user__company=request.user.company, date=today).select_related('user', 'user__company'):
            active = _availability_active(a)
            avail_map[a.user_id] = active
            if active and a.checked_in_at:
                checkin_map[a.user_id] = a.checked_in_at.isoformat()
        # Assigned projects per user (for the availability label).
        proj_map: dict[int, list] = {}
        for uid, pname in (
            UserProjectAssignment.objects
            .filter(user__in=users)
            .values_list('user_id', 'project__name')
        ):
            proj_map.setdefault(uid, []).append(pname)
        data = []
        for u in users:
            data.append({
                'user_id':      u.id,
                'name':         u.name,
                'role':         desig_map.get(u.designation.upper(), u.designation.lower()),
                'is_available': avail_map.get(u.id, False),
                'checked_in_at': checkin_map.get(u.id),
                'projects':     proj_map.get(u.id, []),
            })
        return Response(data)

    def post(self, request):
        """Admin toggles any user's availability for today (by user_id)."""
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        from datetime import date as date_cls
        user_id      = request.data.get('user_id')
        is_available = request.data.get('is_available', True)
        today        = str(date_cls.today())
        company      = _resolve_company(request)
        try:
            user = User.objects.get(pk=user_id, company=company)
        except User.DoesNotExist:
            return Response({'detail': 'User not found.'}, status=404)
        obj, _ = UserAvailability.objects.update_or_create(
            user=user, date=today,
            defaults={'is_available': is_available, 'checked_in_at': timezone.now() if is_available else None},
        )
        # Marking available flushes the unassigned bucket to this role (window-gated).
        dist_type = _dist_type_for(user)
        if obj.is_available and dist_type:
            _run_distribution(user.company, dist_type)
        return Response({'user_id': user.id, 'is_available': obj.is_available})


class AvailabilityHistoryView(APIView):
    """Sign-in history day by day — who marked available and at what time.

    Reports what was recorded on each date rather than reusing _availability_active(),
    which expires any prior-day record by design: correct for today's board, but a
    history row must still show that someone signed in on the 3rd. Project labels are
    likewise left off, since assignments are current state and would misrepresent what
    a person was on back then.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from datetime import date as date_cls
        company = _resolve_company(request)
        today = date_cls.today()
        date_from = request.query_params.get('date_from') or str(today - timedelta(days=29))
        date_to   = request.query_params.get('date_to')   or str(today)

        desig_map = {'TELECALLER': 'telecaller', 'STM': 'stm'}
        rows = (
            UserAvailability.objects
            .filter(user__company=company, date__gte=date_from, date__lte=date_to,
                    user__designation__in=['TELECALLER', 'STM'])
            .select_related('user')
            .order_by('-date', 'user__name')
        )
        days = {}
        for a in rows:
            d = days.setdefault(str(a.date), {'date': str(a.date), 'telecallers': [], 'stms': []})
            entry = {
                'user_id':       a.user_id,
                'name':          a.user.name,
                'is_available':  a.is_available,
                'checked_in_at': a.checked_in_at.isoformat() if a.checked_in_at else None,
            }
            role = desig_map.get((a.user.designation or '').upper())
            d['stms' if role == 'stm' else 'telecallers'].append(entry)
        out = []
        for d in days.values():
            d['telecaller_count'] = sum(1 for x in d['telecallers'] if x['is_available'])
            d['stm_count']        = sum(1 for x in d['stms'] if x['is_available'])
            out.append(d)
        return Response(out)


class MyAvailabilityView(APIView):
    """Self-service availability for telecallers / STMs.
    Marking available stays active for AVAILABILITY_TTL_HOURS, then auto-resets."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from datetime import date as date_cls
        today = str(date_cls.today())
        avail = UserAvailability.objects.filter(user=request.user, date=today).first()
        active = _availability_active(avail, request.user)
        expires_at = None
        if active:
            # Auto-expires at the role sign-out time; fall back to the TTL if unset.
            expires_at = _availability_expires_at(request.user)
            if expires_at is None and avail and avail.checked_in_at:
                expires_at = (avail.checked_in_at + timedelta(hours=AVAILABILITY_TTL_HOURS)).isoformat()
        return Response({
            'is_available':  active,
            'checked_in_at': avail.checked_in_at.isoformat() if (avail and avail.checked_in_at) else None,
            'expires_at':    expires_at,
            'ttl_hours':     AVAILABILITY_TTL_HOURS,
        })

    def post(self, request):
        from datetime import date as date_cls
        if not (is_telecaller(request.user) or is_stm(request.user)):
            return Response({'detail': 'Only telecallers and STMs can mark their own availability.'},
                            status=status.HTTP_403_FORBIDDEN)
        is_available = request.data.get('is_available', True)
        today = str(date_cls.today())
        obj, _ = UserAvailability.objects.update_or_create(
            user=request.user, date=today,
            defaults={'is_available': is_available, 'checked_in_at': timezone.now() if is_available else None},
        )
        active = _availability_active(obj, request.user)
        # Marking available flushes the unassigned bucket to this user's role (window-gated).
        if active:
            _run_distribution(request.user.company, _dist_type_for(request.user))
        expires_at = None
        if active:
            expires_at = _availability_expires_at(request.user)
            if expires_at is None and obj.checked_in_at:
                expires_at = (obj.checked_in_at + timedelta(hours=AVAILABILITY_TTL_HOURS)).isoformat()
        return Response({'is_available': active, 'expires_at': expires_at, 'ttl_hours': AVAILABILITY_TTL_HOURS})


# ── Distribution Weights ──────────────────────────────────────────────────────
class DistributionWeightView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company = _resolve_company(request)
        users = (
            User.objects
            .filter(company=company, is_active=True, designation__in=['TELECALLER', 'STM'])
            .only('id', 'name', 'designation')
        )
        weight_map = {
            w.user_id: w.weight
            for w in UserDistributionWeight.objects.filter(user__company=company)
        }
        return Response([
            {'user_id': u.id, 'name': u.name, 'role': u.designation.upper(), 'weight': weight_map.get(u.id, 1)}
            for u in users
        ])

    def patch(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        company = _resolve_company(request)
        updates = request.data.get('updates', [])  # [{user_id, weight}]
        for item in updates:
            uid = item.get('user_id')
            w   = max(1, int(item.get('weight', 1)))
            try:
                user = User.objects.get(pk=uid, company=company)
                UserDistributionWeight.objects.update_or_create(user=user, defaults={'weight': w})
            except User.DoesNotExist:
                pass
        return Response({'detail': 'Weights saved.'})


# ── Distribution ─────────────────────────────────────────────────────────────
def _window_state(company, dist_type):
    """Return 'open' | 'before_signin' | 'after_signout' for the company's
    sign-in/sign-out window (IST). No settings → treated as 'open'."""
    from zoneinfo import ZoneInfo
    settings = DistributionSettings.objects.filter(company=company).first()
    if not settings:
        return 'open'
    field_prefix = 'tc' if dist_type == 'telecaller' else 'stm'
    now_ist = timezone.now().astimezone(ZoneInfo('Asia/Kolkata')).strftime('%H:%M')
    signin  = str(getattr(settings, f'{field_prefix}_signin_time'))[:5]
    signout = str(getattr(settings, f'{field_prefix}_signout_time'))[:5]
    if now_ist < signin:
        return 'before_signin'
    if now_ist >= signout:
        return 'after_signout'
    return 'open'


def _telecaller_project_ids(company):
    """Projects that actually have a telecaller on them.

    A project with nobody from telecalling assigned has no telecaller stage, so its
    leads go straight to an STM. Derived from the assignments rather than a setting
    someone has to remember to flip: assign a telecaller and the project rejoins the
    telecaller flow by itself; remove the last one and it leaves.
    """
    return set(
        UserProjectAssignment.objects.filter(
            user__company=company,
            user__is_active=True,
            user__designation__iexact='TELECALLER',
        ).values_list('project_id', flat=True)
    )


def _pending_direct_to_stm(company):
    """New leads whose project has no telecaller assigned to it."""
    return Lead.objects.filter(
        company=company, status='new',
        telecaller__isnull=True, stm__isnull=True,
        project__isnull=False,
    ).exclude(project_id__in=_telecaller_project_ids(company))


def _run_distribution(company, dist_type, triggered_by=None, gate='full'):
    """Distribute, then hand telecaller-less projects straight to an STM.

    A project with no telecaller assigned has no telecalling stage, so its leads must
    reach an STM. That has to happen even when the telecaller pass returned early --
    window closed, nobody marked available -- because those leads never needed a
    telecaller in the first place.
    """
    result = _distribute(company, dist_type, triggered_by, gate)
    if dist_type != 'telecaller' or not _pending_direct_to_stm(company).exists():
        return result

    direct = _distribute(company, 'stm', triggered_by, gate)
    merged = dict(result)
    merged['distributed'] = result.get('distributed', 0) + direct.get('distributed', 0)
    merged['assignments'] = {**result.get('assignments', {}), **direct.get('assignments', {})}
    if direct.get('distributed'):
        merged.pop('message', None)
    elif direct.get('message'):
        merged['message'] = ' '.join(filter(None, [result.get('message'), direct['message']]))
    return merged


def _distribute(company, dist_type, triggered_by=None, gate='full'):
    """Weighted, project-aware, window-gated assignment of the current unassigned
    bucket to available telecallers/STMs. Reusable by both the manual Distribute
    button and the automatic triggers (lead created / marked available / warm).

    gate='full'    → only runs when the window is 'open' (auto-assignment).
    gate='signout' → runs unless 'after_signout' (manual admin override).

    triggered_by=None marks the assignment as automatic ("System") in history.
    Returns the same dict shape the API has always returned.
    """
    from datetime import date as date_cls

    desig = 'TELECALLER' if dist_type == 'telecaller' else 'STM'

    state = _window_state(company, dist_type)
    if state == 'after_signout':
        return {'distributed': 0, 'message': f'Distribution window closed for {desig}. Leads remain unassigned.'}
    if gate == 'full' and state != 'open':
        return {'distributed': 0, 'message': f'Outside {desig} distribution window. Leads remain unassigned.'}

    today = str(date_cls.today())

    # Users marked available today. Availability auto-expires at the role's sign-out
    # time, which the window gate above already enforces (distribution never runs
    # after sign-out), so a same-day check-in stays valid through the whole window.
    avail_ids = set(
        UserAvailability.objects.filter(
            user__company=company,
            user__designation__iexact=desig,
            date=today,
            is_available=True,
        ).values_list('user_id', flat=True)
    )
    if not avail_ids:
        return {'distributed': 0, 'message': f'No {desig}s have marked available today.'}

    members = list(User.objects.filter(pk__in=avail_ids, is_active=True).only('id', 'name'))
    if not members:
        return {'distributed': 0, 'message': f'No active {desig} users available.'}

    weight_map = {
        w.user_id: w.weight
        for w in UserDistributionWeight.objects.filter(user__in=members)
    }

    with transaction.atomic():
        # Lock unassigned leads row-by-row so concurrent distribution calls
        # (auto + manual firing simultaneously) can't grab the same leads.
        company_leads = Lead.objects.filter(company=company)
        if dist_type == 'telecaller':
            # stm__isnull=True too — a lead an STM (or CP) already self-sourced has
            # stm set but stays status='new'/telecaller=NULL (nothing else moves it
            # off 'new' at create time), so without this it silently qualified as
            # "unassigned" and got swept into telecaller distribution the next time
            # ANY unrelated lead-create triggered this company-wide run — handing an
            # STM's own lead to a telecaller entirely by accident.
            # Only projects that have a telecaller assigned. The rest are handled by
            # the STM pass below; leaving them here would park them as "skipped".
            qs = (company_leads
                  .filter(telecaller__isnull=True, stm__isnull=True, status='new',
                          project_id__in=_telecaller_project_ids(company))
                  .select_for_update(skip_locked=True).order_by('created_at'))
        else:
            # Warm-transferred leads, plus new leads whose project has no telecaller
            # assigned to it at all.
            qs = (company_leads
                  .filter(Q(status='warm_transferred', stm__isnull=True)
                          | (Q(status='new', stm__isnull=True, telecaller__isnull=True,
                               project__isnull=False)
                             & ~Q(project_id__in=_telecaller_project_ids(company))))
                  .select_for_update(skip_locked=True).order_by('created_at'))

        leads = list(qs)
        if not leads:
            return {'distributed': 0, 'message': 'No unassigned leads found.'}

        # Today's existing assignment counts (for fair weighted continuation across runs).
        today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        if dist_type == 'telecaller':
            count_qs = Lead.objects.filter(
                telecaller__in=members, telecaller_assigned_at__gte=today_start
            ).values('telecaller_id').annotate(n=Count('id'))
            counts = {row['telecaller_id']: row['n'] for row in count_qs}
        else:
            count_qs = Lead.objects.filter(
                stm__in=members, stm_assigned_at__gte=today_start
            ).values('stm_id').annotate(n=Count('id'))
            counts = {row['stm_id']: row['n'] for row in count_qs}
        for m in members:
            counts.setdefault(m.id, 0)

        # Project assignments (STRICT): a member only receives leads of the project(s)
        # assigned to them. A member with NO project assigned receives NOTHING — and a
        # lead with no project can't be routed to anyone.
        proj_map = {}
        for uid, pid in UserProjectAssignment.objects.filter(
            user__in=members
        ).values_list('user_id', 'project_id'):
            proj_map.setdefault(uid, set()).add(pid)

        member_ids   = [m.id for m in members]
        id_to_member = {m.id: m for m in members}
        user_leads   = {m.id: [] for m in members}
        now = timezone.now()
        skipped = 0

        # Pre-bucket eligible members by project so each lead is matched in O(1)
        # instead of scanning every member (O(L×M) → O(L+M)). Members are added in
        # member_ids order, so the weighted-min tie-break stays identical to before.
        proj_to_uids = {}
        for uid in member_ids:
            for pid in proj_map.get(uid, ()):
                proj_to_uids.setdefault(pid, []).append(uid)

        for lead in leads:
            eligible = proj_to_uids.get(lead.project_id) if lead.project_id is not None else None
            if not eligible:
                skipped += 1
                continue
            best = min(eligible, key=lambda uid: counts[uid] / (weight_map.get(uid, 1)))
            user_leads[best].append(lead.pk)
            counts[best] += 1

        assignments = []
        history_rows = []
        note = 'Auto-assigned' if triggered_by is None else 'Manually assigned'
        for uid, pks in user_leads.items():
            if not pks:
                continue
            if dist_type == 'telecaller':
                Lead.objects.filter(pk__in=pks).update(
                    telecaller_id=uid, status='assigned', telecaller_assigned_at=now,
                )
            else:
                Lead.objects.filter(pk__in=pks).update(stm_id=uid, stm_assigned_at=now)
                # A lead that skipped the telecaller stage arrives still 'new';
                # mark it assigned. Not 'warm_transferred' — nobody transferred it,
                # and that status feeds the warm/SQL funnel.
                Lead.objects.filter(pk__in=pks, status='new').update(status='assigned')
            for pk in pks:
                history_rows.append(LeadStatusHistory(
                    lead_id=pk, changed_by=triggered_by,
                    field_changed=dist_type, old_value='', new_value=id_to_member[uid].name,
                    remarks=note,
                ))
            assignments.append({'name': id_to_member[uid].name, 'count': len(pks)})
            from notifications import notify
            notify(id_to_member[uid], 'new_lead', 'New Leads Assigned',
                   f'{len(pks)} new lead{"s" if len(pks) > 1 else ""} assigned to you.')

        if history_rows:
            LeadStatusHistory.objects.bulk_create(history_rows)

        distributed = sum(a['count'] for a in assignments)
        if distributed:
            DistributionLog.objects.create(
                company=company,
                dist_type=dist_type,
                triggered_by=triggered_by,
                leads_distributed=distributed,
                details={'assignments': assignments, 'auto': triggered_by is None},
            )

    resp = {'distributed': distributed, 'assignments': {a['name']: a['count'] for a in assignments}}
    if skipped:
        resp['message'] = f'{skipped} lead(s) left unassigned — no available {desig} is assigned to their project.'
    return resp


def _record_lead_created(lead, by=None):
    """Add the opening 'Lead created' entry to a lead's history (with its source)."""
    src = lead.source.name if lead.source_id else 'manual'
    campaign = lead.meta_campaign_name or ''
    new_value = (f'{src} · {campaign}' if campaign else src)[:100]
    LeadStatusHistory.objects.create(
        lead=lead, changed_by=by, field_changed='created',
        old_value='', new_value=new_value, remarks='Lead created',
    )


class DistributeView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        dist_type = request.data.get('dist_type', request.data.get('type', 'telecaller'))
        company   = _resolve_company(request)
        # Manual admin trigger: weight-based, allowed before sign-in, blocked after sign-out.
        resp = _run_distribution(company, dist_type, triggered_by=request.user, gate='signout')
        return Response(resp)


class DistributionLogView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        logs = scope_to_company(
            DistributionLog.objects.select_related('triggered_by'), request.user
        )
        if request.query_params.get('company_id') and is_platform_admin(request.user):
            logs = logs.filter(company_id=request.query_params['company_id'])
        logs = logs[:30]
        data = [{
            'id':                  log.id,
            'dist_type':           log.dist_type,
            'leads_distributed':   log.leads_distributed,
            'triggered_by_name':   log.triggered_by.name if log.triggered_by else 'System',
            'details':             log.details,
            'created_at':          log.created_at,
        } for log in logs]
        return Response(data)

    def delete(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        qs = scope_to_company(DistributionLog.objects.all(), request.user)
        if request.query_params.get('company_id') and is_platform_admin(request.user):
            qs = qs.filter(company_id=request.query_params['company_id'])
        qs.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Bulk Import ───────────────────────────────────────────────────────────────
# ── Lead-import helpers (flexible cell parsing for the lifecycle template) ──────
def _imp_dt(val):
    """Parse a cell into an aware datetime. Accepts ISO, yyyy-mm-dd, dd-mm-yyyy, dd/mm/yyyy."""
    from datetime import datetime as _dt, time as _time
    from django.utils.dateparse import parse_datetime, parse_date
    import re as _re
    s = str(val or '').strip()
    if not s:
        return None
    dt = parse_datetime(s)
    if dt:
        dt = timezone.make_aware(dt) if timezone.is_naive(dt) else dt
        # Midnight (incl. Excel date cells) → noon so the calendar date is timezone-stable.
        if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
            dt = dt.replace(hour=12)
        return dt
    d = parse_date(s)
    if d:
        # Anchor date-only values at noon so the calendar date is stable across timezones.
        return timezone.make_aware(_dt.combine(d, _time(12, 0)))
    m = _re.match(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})$', s)
    if m:
        dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yy < 100:
            yy += 2000
        try:
            return timezone.make_aware(_dt(yy, mm, dd, 12, 0))
        except ValueError:
            return None
    return None


def _imp_date(val):
    dt = _imp_dt(val)
    return dt.date() if dt else None


def _imp_int(val):
    s = str(val or '').replace(',', '').strip()
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _imp_dec(val):
    s = str(val or '').replace(',', '').replace('₹', '').strip()
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _imp_purpose(val, valid):
    """Comma/pipe/semicolon-separated purpose values ('Investment, End Use') → the
    matching canonical option keys, dropping anything that doesn't match."""
    if not val:
        return []
    import re as _re
    parts = [p.strip().lower().replace(' ', '_') for p in _re.split(r'[,;|/]+', str(val)) if p.strip()]
    return [p for p in parts if p in valid]


# Canonical row keys the importer understands (the full-pipeline columns).
IMPORT_COLUMNS = [
    'name', 'phone', 'alt_phone', 'email', 'project', 'source', 'campaign', 'adset', 'ad_name',
    'requirement', 'budget_min', 'budget_max', 'preferred_location', 'city', 'address', 'purpose', 'budget_bucket',
    'lead_date', 'overall_status',
    'telecaller_code', 'telecaller_status', 'telecaller_remarks',
    'stm_code', 'stm_status', 'stm_remarks',
    'sv_scheduled_date', 'sv_visited_date', 'sv_status', 'sv_referred_by_code', 'sv_remarks',
    'closure_date', 'closure_status', 'unit_no', 'unit_type', 'booking_amount', 'total_amount', 'closure_remarks',
]
# Header → canonical-key aliases (the per-row loop reads 'creative', not 'ad_name').
_IMP_ALIASES = {
    'name': {'name', 'full_name', 'fullname', 'customer_name', 'lead_name', 'first_name'},
    'phone': {'phone', 'phone_number', 'phonenumber', 'mobile', 'mobile_number', 'contact', 'cell'},
    'alt_phone': {'alt_phone', 'alternate_phone', 'phone_2', 'secondary_phone', 'other_phone'},
    'email': {'email', 'e_mail', 'email_address'},
    'campaign': {'campaign', 'campaign_name', 'meta_campaign', 'utm_campaign', 'ad_campaign'},
    'adset': {'adset', 'adset_name', 'ad_set', 'ad_group_name', 'adgroup'},
    'creative': {'creative', 'ad_name', 'creative_name', 'ad_creative', 'advertisement_name'},
    'lead_date': {'lead_date', 'date', 'created', 'created_at', 'submission_date', 'timestamp'},
    # Backward-compat: these columns used to hold a raw numeric id (pre-user_code
    # rename) — a template downloaded before the rename, or a habitually-typed old
    # header, should still auto-map instead of silently dropping the column.
    'telecaller_code': {'telecaller_code', 'telecaller_id'},
    'stm_code': {'stm_code', 'stm_id'},
    'sv_referred_by_code': {'sv_referred_by_code', 'sv_referred_by_id'},
}
_IMP_CANON = set(IMPORT_COLUMNS) | {'creative'}


def _imp_canon_key(header):
    import re as _re
    k = _re.sub(r'[\s\-]+', '_', str(header or '').strip().lower())
    for field, aliases in _IMP_ALIASES.items():
        if k in aliases:
            return field
    return k if k in _IMP_CANON else None


def _imp_parse_file(f):
    """Parse an uploaded .xlsx/.csv into a list of row dicts keyed by canonical column names."""
    import io
    fname = (getattr(f, 'name', '') or '').lower()
    headers, raw_rows = [], []
    if fname.endswith('.csv') or fname.endswith('.txt'):
        import csv
        data = f.read()
        text = data.decode('utf-8-sig', errors='ignore') if isinstance(data, bytes) else data
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        raw_rows = [dict(r) for r in reader]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb['Leads'] if 'Leads' in wb.sheetnames else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        headers = [('' if h is None else str(h).strip()) for h in (next(it, []) or [])]
        for r in it:
            raw_rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    colmap = {h: _imp_canon_key(h) for h in headers}
    rows = []
    for rr in raw_rows:
        out = {}
        for h, v in rr.items():
            c = colmap.get(h)
            if c and v is not None and str(v).strip() != '':
                out[c] = v
        if out.get('name') or out.get('phone'):
            rows.append(out)
    return rows


class BulkImportLeadsView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not has_sales_access(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        from .models import LEAD_STATUS, TC_STATUS, STM_STATUS, SV_STATUS, CLOSURE_STATUS, BUDGET_BUCKETS

        rows       = request.data.get('leads', [])
        project_id = request.data.get('project_id')   # default project for every row
        source_id  = request.data.get('source_id')    # default source for every row
        company    = request.user.company
        # An STM only works the STM stage — telecaller assignment isn't theirs to set,
        # so any telecaller_code/status/remarks in the file is ignored for their
        # uploads (mirrors the template omitting those columns for an STM login).
        uploader_is_stm = is_stm(request.user)

        # App/web may upload the spreadsheet itself (multipart) instead of pre-parsed
        # JSON rows — parse it server-side into the same canonical row dicts.
        if not rows and request.FILES.get('file'):
            try:
                rows = _imp_parse_file(request.FILES['file'])
            except Exception as e:
                return Response({'detail': 'Could not read the file: %s' % e}, status=status.HTTP_400_BAD_REQUEST)

        if not rows:
            return Response({'detail': 'No leads provided.'}, status=status.HTTP_400_BAD_REQUEST)

        # A supplied default project/source must belong to the requester's company.
        if project_id and not _project_in_scope(request, project_id):
            return Response({'detail': 'Invalid project for your company.'}, status=status.HTTP_400_BAD_REQUEST)
        if source_id and not scope_to_company(LeadSource.objects.filter(pk=source_id), request.user).exists():
            return Response({'detail': 'Invalid source for your company.'}, status=status.HTTP_400_BAD_REQUEST)

        # Allowed status values + company-scoped lookup tables (resolved once).
        LEAD_ST = {k for k, _ in LEAD_STATUS}
        TC_ST   = {k for k, _ in TC_STATUS}
        STM_ST  = {k for k, _ in STM_STATUS}
        SV_ST   = {k for k, _ in SV_STATUS}
        CL_ST   = {k for k, _ in CLOSURE_STATUS}
        BB_ST   = {k for k, _ in BUDGET_BUCKETS}
        PURPOSE_VALID = {'investment', 'end_use', 'other'}
        proj_by_name = {p.name.strip().lower(): p.id for p in scope_to_company(Project.objects.all(), request.user)}
        src_by_name  = {s.name.strip().lower(): s.id for s in scope_to_company(LeadSource.objects.all(), request.user)}
        uq = User.objects.filter(is_active=True)
        if company:
            uq = uq.filter(company=company)
        uq = list(uq)
        code_to_id = {u.user_code.strip().lower(): u.id for u in uq if u.user_code}
        valid_user_ids = {u.id for u in uq}

        def _uid(v):
            s = str(v or '').strip()
            if not s:
                return None
            hit = code_to_id.get(s.lower())
            if hit:
                return hit
            # Backward-compat: a file from before the user_code rename may still
            # carry a raw numeric id in the cell (only the header changed).
            i = _imp_int(s)
            return i if i in valid_user_ids else None

        imported = 0
        duplicates = 0
        errors = 0
        bare_new = 0
        failed = []
        warnings = []  # non-fatal: row still imports, but a code/value didn't resolve

        # Build existing dup set (last-10-digits) scoped to this company — O(n) once.
        company_leads = scope_to_company(Lead.objects.all(), request.user)
        # Read the blind index rather than decrypting every stored phone.
        existing_keys = set(company_leads.values_list('phone_key', flat=True))
        existing_keys.discard('')

        to_create = []   # Lead objects
        meta      = []   # parallel per-row dict carrying lead_date + SV/closure raw data
        for i, row in enumerate(rows):
            name  = str(row.get('name', '')).strip()
            phone = str(row.get('phone', '')).strip()
            if not name or not phone:
                errors += 1
                failed.append({'row': i + 1, 'name': name, 'phone': phone, 'reason': 'Missing name or phone'})
                continue

            clean = ''.join(c for c in phone if c.isdigit())[-10:]
            # existing_keys holds blind-index hashes, so hash before comparing.
            clean_key = phone_blind_index(clean) if clean else ''
            is_dup = bool(clean_key) and clean_key in existing_keys

            rproj = proj_by_name.get(str(row.get('project', '')).strip().lower()) or project_id or None
            rsrc  = src_by_name.get(str(row.get('source', '')).strip().lower()) or source_id or None
            tc_id  = None if uploader_is_stm else _uid(row.get('telecaller_code'))
            stm_id = _uid(row.get('stm_code'))
            sv_ref_id = _uid(row.get('sv_referred_by_code'))
            code_checks = [
                ('STM Code', row.get('stm_code'), stm_id),
                ('SV Referred By Code', row.get('sv_referred_by_code'), sv_ref_id),
            ]
            if not uploader_is_stm:
                code_checks.insert(0, ('Telecaller Code', row.get('telecaller_code'), tc_id))
            for label, raw_val, resolved in code_checks:
                if str(raw_val or '').strip() and not resolved:
                    warnings.append({'row': i + 1, 'name': name, 'field': label, 'value': str(raw_val).strip(),
                                      'reason': "didn't match any user's code — left unassigned"})
            # An STM uploading their own leads (e.g. a walk-in sign-in sheet, no STM
            # Code column filled in) self-sources them, same as the single "Add Lead"
            # flow already does — checked after the warning above so a genuinely wrong
            # code still surfaces its warning rather than silently becoming "assign to
            # me". Otherwise a row with neither STM nor telecaller falls through to
            # telecaller auto-distribution, handing the STM's own lead to someone else.
            if uploader_is_stm and not stm_id:
                stm_id = request.user.id

            tc_status  = '' if uploader_is_stm else str(row.get('telecaller_status', '')).strip().lower()
            tc_status  = tc_status if tc_status in TC_ST else ''
            stm_status = str(row.get('stm_status', '')).strip().lower()
            stm_status = stm_status if stm_status in STM_ST else ''

            budget_bucket = str(row.get('budget_bucket', '')).strip().lower().replace(' ', '_')
            budget_bucket = budget_bucket if budget_bucket in BB_ST else ''
            purpose = _imp_purpose(row.get('purpose'), PURPOSE_VALID)

            lead_dt = _imp_dt(row.get('lead_date'))

            # SV / closure presence
            sv_sched = _imp_dt(row.get('sv_scheduled_date'))
            sv_vis   = _imp_dt(row.get('sv_visited_date'))
            sv_stat  = str(row.get('sv_status', '')).strip().lower()
            sv_stat  = sv_stat if sv_stat in SV_ST else ''
            has_sv   = bool(sv_sched or sv_vis or sv_stat or str(row.get('sv_remarks', '')).strip())
            cl_date  = _imp_date(row.get('closure_date'))

            # Overall lead status: explicit wins; otherwise derive from the furthest stage reached.
            overall = str(row.get('overall_status', '')).strip().lower()
            if overall not in LEAD_ST:
                if cl_date:
                    overall = 'closed'
                elif has_sv:
                    overall = 'sv_done' if sv_stat == 'completed' else 'sv_scheduled'
                elif stm_id:
                    overall = 'warm_transferred'
                elif tc_id:
                    overall = 'assigned'
                else:
                    overall = 'new'

            to_create.append(Lead(
                company=company,
                name=name,
                phone=phone,
                # bulk_create skips save(), so set the lookup key here or these rows
                # would be invisible to duplicate detection and phone search.
                phone_key=clean_key,
                alt_phone=str(row.get('alt_phone', '')).strip(),
                email=str(row.get('email', '')).strip(),
                project_id=rproj,
                source_id=rsrc,
                meta_campaign_name=str(row.get('campaign', '')).strip(),
                meta_adset_name=str(row.get('adset', '')).strip(),
                meta_ad_name=str(row.get('creative', '')).strip(),
                requirement=str(row.get('requirement', '')).strip(),
                preferred_location=str(row.get('preferred_location', '')).strip(),
                budget_min=_imp_int(row.get('budget_min')),
                budget_max=_imp_int(row.get('budget_max')),
                city=str(row.get('city', '')).strip(),
                address=str(row.get('address', '')).strip(),
                purpose=purpose,
                budget_bucket=budget_bucket,
                status=overall,
                telecaller_id=tc_id,
                telecaller_status=tc_status,
                telecaller_remarks='' if uploader_is_stm else str(row.get('telecaller_remarks', '')).strip(),
                telecaller_assigned_at=(lead_dt or timezone.now()) if tc_id else None,
                stm_id=stm_id,
                stm_status=stm_status,
                stm_remarks=str(row.get('stm_remarks', '')).strip(),
                stm_assigned_at=(lead_dt or timezone.now()) if stm_id else None,
                is_duplicate=is_dup,
            ))
            meta.append({
                'lead_dt': lead_dt,
                'has_sv': has_sv, 'sv_sched': sv_sched, 'sv_vis': sv_vis, 'sv_stat': sv_stat or 'scheduled',
                'sv_ref': sv_ref_id, 'sv_remarks': str(row.get('sv_remarks', '')).strip(),
                'cl_date': cl_date, 'cl_status': (str(row.get('closure_status', '')).strip().lower() if str(row.get('closure_status', '')).strip().lower() in CL_ST else 'booked'),
                'unit_no': str(row.get('unit_no', '')).strip(), 'unit_type': str(row.get('unit_type', '')).strip(),
                'booking_amount': _imp_dec(row.get('booking_amount')), 'total_amount': _imp_dec(row.get('total_amount')),
                'cl_remarks': str(row.get('closure_remarks', '')).strip(),
            })

            if is_dup:
                duplicates += 1
            else:
                imported += 1
                if clean_key:
                    existing_keys.add(clean_key)  # catch in-batch duplicates too
            # Only a genuinely untouched lead (no telecaller AND no STM) should be swept
            # into telecaller auto-distribution — mirrors the single-lead-create check
            # above (`not lead.telecaller_id and not lead.stm_id`). A row that names an
            # STM but not a telecaller has already skipped/passed that stage; sweeping
            # it in anyway is what was handing STM-assigned leads to a telecaller too.
            if not tc_id and not stm_id and overall == 'new':
                bare_new += 1

        with transaction.atomic():
            created = Lead.objects.bulk_create(to_create)

            # Honour historical lead_date by overriding the auto_now_add created_at.
            dated = []
            for lead, m in zip(created, meta):
                if m['lead_dt']:
                    lead.created_at = m['lead_dt']
                    dated.append(lead)
            if dated:
                Lead.objects.bulk_update(dated, ['created_at'])

            # Materialise Site Visits + Closures linked to each freshly created lead.
            svs, closures = [], []
            for lead, m in zip(created, meta):
                if m['has_sv']:
                    svs.append(SiteVisit(
                        lead=lead, project_id=lead.project_id,
                        scheduled_at=m['sv_sched'], visited_at=m['sv_vis'], status=m['sv_stat'],
                        stm_id=lead.stm_id, referred_by_telecaller_id=(m['sv_ref'] or lead.telecaller_id),
                        remarks=('[Imported] ' + m['sv_remarks']).strip(),
                    ))
                if m['cl_date']:
                    # Historical closure (no Booking/LOI) — tagged so it's distinguishable
                    # from closures produced by the booking form.
                    closures.append(Closure(
                        company_id=lead.company_id, lead=lead,
                        client_name=lead.name or '', client_phone=lead.phone or '',
                        project_id=lead.project_id, stm_id=lead.stm_id,
                        referred_by_telecaller_id=lead.telecaller_id, status=m['cl_status'],
                        closure_date=m['cl_date'], unit_no=m['unit_no'], unit_type=m['unit_type'],
                        booking_amount=m['booking_amount'], total_amount=m['total_amount'],
                        remarks=('[Imported] ' + m['cl_remarks']).strip(),
                    ))
            if svs:
                SiteVisit.objects.bulk_create(svs)
            if closures:
                Closure.objects.bulk_create(closures)

        # Auto-assign only the genuinely bare/new bucket (rows that carried an STM/TC
        # or a later stage are already placed and must not be redistributed).
        if bare_new:
            _run_distribution(company, 'telecaller')
        return Response({
            'imported': imported, 'duplicates': duplicates, 'errors': errors, 'failed': failed,
            'warnings': warnings,
            'site_visits': len([m for m in meta if m['has_sv']]),
            'closures': len([m for m in meta if m['cl_date']]),
        })


class LeadImportTemplateView(APIView):
    """Generates the Full-Pipeline import template (.xlsx) server-side with dropdowns,
    a styled table, coloured required/closure headers and a Reference sheet — so the
    mobile app (which can't build a rich xlsx on-device) downloads the same template
    the web generates."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not has_sales_access(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        import openpyxl
        from io import BytesIO
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        company = request.user.company
        projects = list(scope_to_company(Project.objects.all(), request.user).values_list('name', flat=True))
        sources  = list(scope_to_company(LeadSource.objects.all(), request.user).values_list('name', flat=True))
        uq = User.objects.filter(is_active=True).exclude(role='Admin')
        if company:
            uq = uq.filter(company=company)
        users = list(uq.values('id', 'name', 'user_code', 'designation', 'role', 'phone').order_by('name'))

        # An STM only works the STM stage — telecaller assignment isn't theirs to set,
        # so the template doesn't even offer those columns for an STM login (uploads
        # ignore them regardless — see BulkImportLeadsView — this just avoids handing
        # out a template with fields that'll silently be dropped).
        uploader_is_stm = is_stm(request.user)

        cols = [
            'name', 'phone', 'alt_phone', 'email', 'project', 'source', 'campaign', 'adset', 'ad_name',
            'requirement', 'budget_min', 'budget_max', 'preferred_location', 'city', 'address', 'purpose', 'budget_bucket',
            'lead_date', 'overall_status',
            'telecaller_code', 'telecaller_status', 'telecaller_remarks', 'stm_code', 'stm_status', 'stm_remarks',
            'sv_scheduled_date', 'sv_visited_date', 'sv_status', 'sv_referred_by_code', 'sv_remarks',
            'closure_date', 'closure_status', 'unit_no', 'unit_type', 'booking_amount', 'total_amount', 'closure_remarks',
        ]
        if uploader_is_stm:
            cols = [c for c in cols if c not in ('telecaller_code', 'telecaller_status', 'telecaller_remarks')]
        # Display-only header text — the parser normalises spaces/case back to the
        # canonical snake_case key (see _imp_canon_key), so this is purely cosmetic.
        HEADER_LABELS = {
            'name': 'Name', 'phone': 'Phone', 'alt_phone': 'Alt Phone', 'email': 'Email',
            'project': 'Project', 'source': 'Source', 'campaign': 'Campaign Name', 'adset': 'Ad Set',
            'ad_name': 'Ad Name', 'requirement': 'Requirement', 'budget_min': 'Budget Min',
            'budget_max': 'Budget Max', 'preferred_location': 'Preferred Location', 'city': 'City',
            'address': 'Address', 'purpose': 'Purpose', 'budget_bucket': 'Budget Bucket',
            'lead_date': 'Lead Date', 'overall_status': 'Overall Status',
            'telecaller_code': 'Telecaller Code', 'telecaller_status': 'Telecaller Status',
            'telecaller_remarks': 'Telecaller Remarks', 'stm_code': 'STM Code', 'stm_status': 'STM Status',
            'stm_remarks': 'STM Remarks', 'sv_scheduled_date': 'SV Scheduled Date',
            'sv_visited_date': 'SV Visited Date', 'sv_status': 'SV Status',
            'sv_referred_by_code': 'SV Referred By Code', 'sv_remarks': 'SV Remarks',
            'closure_date': 'Closure Date', 'closure_status': 'Closure Status', 'unit_no': 'Unit No',
            'unit_type': 'Unit Type', 'booking_amount': 'Booking Amount', 'total_amount': 'Total Amount',
            'closure_remarks': 'Closure Remarks',
        }
        STATUS = {
            'overall_status': 'new,assigned,contacted,not_reachable,warm_transferred,hot,warm,cold,not_interested,sv_scheduled,sv_done,closed,lost',
            'telecaller_status': 'warm,cold,not_interested,not_reachable,callback',
            'stm_status': 'hot,warm,cold,not_interested,sv_scheduled,sv_done,closed',
            'sv_status': 'scheduled,completed,cancelled,no_show',
            'closure_status': 'booked,cancelled,refunded',
            'budget_bucket': 'lt_10l,10_50l,50l_1cr,1_2cr,2_3cr,3_5cr,gt_5cr',
        }
        if uploader_is_stm:
            STATUS.pop('telecaller_status', None)
        # purpose is multi-select (comma-separated) so it can't use the same
        # single-value dropdown as STATUS — documented in the Reference sheet instead.
        PURPOSE_VALUES = 'investment, end_use, other'
        def _role(u):
            return (u['designation'] or u['role'] or '').lower()
        tc_code  = next((u['user_code'] for u in users if 'tele' in _role(u) and u['user_code']), (users[0]['user_code'] if users else ''))
        stm_code = next((u['user_code'] for u in users if any(k in _role(u) for k in ('stm', 'sales', 'manager')) and u['user_code']),
                        (users[1]['user_code'] if len(users) > 1 else (users[0]['user_code'] if users else '')))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Leads'
        ws.append([HEADER_LABELS.get(c, c) for c in cols])
        ex1 = {'name': 'Rahul Sharma', 'phone': '9876543210', 'email': 'rahul@example.com', 'source': (sources[0] if sources else 'meta'), 'campaign': 'Meta - Luxury Homes', 'ad_name': 'Video 2BHK', 'city': 'Ahmedabad', 'purpose': 'end_use', 'budget_bucket': '50l_1cr', 'lead_date': '01-05-2025', 'overall_status': 'new', 'telecaller_code': tc_code, 'telecaller_status': 'callback', 'telecaller_remarks': 'Call back evening'}
        ex2 = {'name': 'Priya Mehta', 'phone': '9988776655', 'email': 'priya@example.com', 'project': (projects[0] if projects else 'Kalrav'), 'source': (sources[0] if sources else 'walk-in'), 'city': 'Vadodara', 'address': '12 Alkapuri Society', 'purpose': 'investment, end_use', 'budget_bucket': '1_2cr', 'lead_date': '02-04-2025', 'overall_status': 'closed', 'telecaller_code': tc_code, 'telecaller_status': 'warm', 'stm_code': stm_code, 'stm_status': 'closed', 'sv_scheduled_date': '05-04-2025', 'sv_visited_date': '06-04-2025', 'sv_status': 'completed', 'sv_remarks': 'Liked plot A-12', 'closure_date': '08-04-2025', 'closure_status': 'booked', 'unit_no': 'A-12', 'unit_type': '2BHK', 'booking_amount': 200000, 'total_amount': 5000000, 'closure_remarks': 'Token received'}
        for ex in (ex1, ex2):
            ws.append([ex.get(c, '') for c in cols])

        last_col = get_column_letter(len(cols))
        table = Table(displayName='LeadsImport', ref='A1:%s3' % last_col)
        table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
        ws.add_table(table)
        ws.freeze_panes = 'A2'
        for i, c in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(26, max(12, len(HEADER_LABELS.get(c, c)) + 3))

        def col_of(name):
            return get_column_letter(cols.index(name) + 1)
        red, purple, white = PatternFill('solid', fgColor='C62828'), PatternFill('solid', fgColor='7C3AED'), Font(bold=True, color='FFFFFF')
        for f in ('name', 'phone'):
            ws['%s1' % col_of(f)].fill = red
            ws['%s1' % col_of(f)].font = white
        for f in ('closure_date', 'closure_status', 'unit_no', 'unit_type', 'booking_amount', 'total_amount', 'closure_remarks'):
            ws['%s1' % col_of(f)].fill = purple
            ws['%s1' % col_of(f)].font = white

        lists = wb.create_sheet('Lists')
        lists.sheet_state = 'hidden'
        for i, n in enumerate(projects, start=1):
            lists['A%d' % i] = n
        for i, n in enumerate(sources, start=1):
            lists['B%d' % i] = n

        MAXROW = 1000
        def add_dv(name, formula):
            dv = DataValidation(type='list', formula1=formula, allow_blank=True, showErrorMessage=True, errorStyle='warning')
            ws.add_data_validation(dv)
            dv.add('%s2:%s%d' % (col_of(name), col_of(name), MAXROW))
        for field, vals in STATUS.items():
            add_dv(field, '"%s"' % vals)
        if projects:
            add_dv('project', 'Lists!$A$1:$A$%d' % len(projects))
        if sources:
            add_dv('source', 'Lists!$B$1:$B$%d' % len(sources))

        ref = wb.create_sheet('Reference — codes & values')
        ref.append([
            '— TEAM — put this code in the STM Code / SV Referred By Code columns —' if uploader_is_stm else
            '— TEAM — put this code in the Telecaller Code / STM Code / SV Referred By Code columns —',
        ])
        ref.append(['User Code', 'Name', 'Role / Designation', 'Phone'])
        ref['A2'].font = Font(bold=True)
        for u in users:
            ref.append([u['user_code'] or '—', u['name'], (u['designation'] or u['role'] or ''), u['phone'] or ''])
        ref.append([])
        ref.append(['— ALLOWED VALUES (the Leads sheet has dropdowns for these) —'])
        for k, v in STATUS.items():
            ref.append([HEADER_LABELS.get(k, k), v.replace(',', ', ')])
        ref.append(['Purpose (multi-select — separate multiple with a comma)', PURPOSE_VALUES])
        ref.append([])
        ref.append(['— NOTES —'])
        ref.append(['Header colours: RED = required (name, phone). PURPLE = closure columns.'])
        ref.append(['Dates: dd-mm-yyyy. project/source are matched by name. Leave a cell blank to skip.'])
        ref.append(['Fill any sv_* column to create a Site Visit; fill closure_date to create a Closure.'])
        ref.append(['purpose accepts multiple values in one cell, e.g. "investment, end_use".'])
        ref.column_dimensions['A'].width = 24
        ref.column_dimensions['B'].width = 62

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="vistara_pipeline_import_template.xlsx"'
        return resp


# ── Reports ───────────────────────────────────────────────────────────────────
class ReportsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Count, Sum, Q

        user      = request.user
        leads_qs  = scope_to_company(Lead.objects.all(), user)
        sv_qs     = scope_to_company(SiteVisit.objects.all(), user, 'lead__company')
        closure_qs = scope_to_company(Closure.objects.all(), user, 'company')
        company_id = request.query_params.get('company_id')
        if company_id and is_platform_admin(user):
            leads_qs   = leads_qs.filter(company_id=company_id)
            sv_qs      = sv_qs.filter(lead__company_id=company_id)
            closure_qs = closure_qs.filter(company_id=company_id)

        # Optional date window — bounds the aggregate scans. No default, so the
        # existing all-time behaviour is unchanged unless the client sends dates.
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        if date_from:
            leads_qs   = leads_qs.filter(created_at__date__gte=date_from)
            sv_qs      = sv_qs.filter(created_at__date__gte=date_from)
            closure_qs = closure_qs.filter(closure_date__gte=date_from)
        if date_to:
            leads_qs   = leads_qs.filter(created_at__date__lte=date_to)
            sv_qs      = sv_qs.filter(created_at__date__lte=date_to)
            closure_qs = closure_qs.filter(closure_date__lte=date_to)

        # Hierarchy scope: managers (anyone with reports below them) get a team report
        # over their subtree; leaf users get a personal report. Admins/top heads see all.
        if _sees_all_company(user, request):
            team_view = True
        else:
            _ids = _visible_user_ids(user)
            leads_qs   = leads_qs.filter(Q(stm__in=_ids) | Q(telecaller__in=_ids))
            sv_qs      = sv_qs.filter(Q(stm__in=_ids) | Q(referred_by_telecaller__in=_ids))
            closure_qs = closure_qs.filter(Q(stm__in=_ids) | Q(referred_by_telecaller__in=_ids))
            team_view  = len(_ids) > 1  # has at least one subordinate → manager view

        def get_campaigns():
            return list(
                leads_qs.exclude(meta_campaign_name='')
                .values('meta_campaign_name')
                .annotate(
                    total=Count('id'),
                    warm=Count('id', filter=Q(status__in=['warm_transferred', 'sv_scheduled', 'sv_done', 'closed'])),
                    sv=Count('id', filter=Q(status__in=['sv_done', 'closed'])),
                    closed=Count('id', filter=Q(status='closed')),
                )
                .order_by('-total')[:20]
            )

        def get_telecallers():
            return list(
                leads_qs.exclude(telecaller__isnull=True)
                .values('telecaller__id', 'telecaller__name')
                .annotate(
                    total=Count('id'),
                    warm=Count('id', filter=Q(telecaller_status='warm')),
                    transferred=Count('id', filter=Q(status='warm_transferred')),
                    sv=Count('id', filter=Q(status__in=['sv_done', 'closed'])),
                )
                .order_by('-total')
            )

        def get_stms():
            return list(
                leads_qs.exclude(stm__isnull=True)
                .values('stm__id', 'stm__name')
                .annotate(
                    total=Count('id'),
                    hot=Count('id', filter=Q(stm_status='hot')),
                    sv_scheduled=Count('id', filter=Q(stm_status='sv_scheduled')),
                    sv_done=Count('id', filter=Q(stm_status__in=['sv_done'])),
                    closed=Count('id', filter=Q(status='closed')),
                )
                .order_by('-total')
            )

        def get_summary():
            # Amounts are encrypted at rest → can't SQL-Sum; sum in Python. Revenue is
            # the FULL closure value (total_amount = final amount), falling back to
            # booking_amount (plot basic) for older closures with no total.
            cnt = closure_qs.count()
            total = sum((c.total_amount or c.booking_amount or 0) for c in closure_qs.only('id', 'booking_amount', 'total_amount'))
            return {
                'total_sv':       sv_qs.count(),
                'completed_sv':   sv_qs.filter(status='completed').count(),
                'total_closures': cnt,
                'total_revenue':  float(total or 0),
                'meta_leads':     leads_qs.exclude(meta_campaign_name='').count(),
            }

        def get_closures():
            return closure_qs.select_related('lead', 'project', 'stm', 'referred_by_telecaller').order_by('-closure_date')[:20]

        # Run sequentially. These are indexed aggregates (fast); the previous
        # ThreadPoolExecutor opened 5 DB connections per request and didn't close
        # them in the worker threads — a connection leak that, with the pooled
        # endpoint + multiple gunicorn workers, risked exhausting Neon.
        return Response({
            # Team-performance tables are management-only; personal reports omit them.
            'team_view':   team_view,
            'campaigns':   get_campaigns()   if team_view else [],
            'telecallers': get_telecallers() if team_view else [],
            'stms':        get_stms()        if team_view else [],
            'closures':    ClosureSerializer(get_closures(), many=True).data,
            'summary':     get_summary(),
        })


class MyTeamView(APIView):
    """Everyone reporting under the requester (their org subtree), with lead/closure
    counts — powers the manager 'My Team' view. Returns [] for users with no reports."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        # Honour the admin "Viewing Company" filter (?company_id) for platform admins.
        company = _resolve_company(request)
        module = (request.query_params.get('module') or '').strip()  # department/module org chart
        scope  = request.query_params.get('scope')                   # 'all' → full company org
        admin_view = request.query_params.get('admin_view') == '1'
        ids = _visible_user_ids(user) - {user.id}   # subtree, excluding self
        is_admin = _sees_all_company(user, request, include_manager_role=False)

        def _full_company():
            # Everyone in a reporting relationship + all Managers (leadership shows
            # even before anyone reports to them); standalone users stay out.
            return list(
                User.objects.filter(company=company, is_active=True)
                .filter(
                    Q(reporting_manager__isnull=False)
                    | Q(subordinates__isnull=False)
                    | Q(role__in=MANAGER_ROLES)
                )
                .distinct().select_related('reporting_manager').order_by('name')
            )

        if is_admin and module:
            # Department/module org chart — users assigned to this module.
            all_users = (User.objects.filter(company=company, is_active=True)
                         .select_related('reporting_manager').order_by('name'))
            members = [u for u in all_users
                       if module in (u.modules or []) or module in (u.manager_modules or [])]
            ids = {u.id for u in members}
        elif is_admin and (scope == 'all' or admin_view or not ids):
            # Full company org (User Management / admin default).
            members = _full_company()
            ids = {u.id for u in members}
        elif not ids:
            return Response([])
        else:
            members = list(
                User.objects.filter(id__in=ids, company=company)
                .select_related('reporting_manager').order_by('name')
            )
        # Admins never appear in the org chart — it reflects the operational hierarchy.
        members = [m for m in members
                   if getattr(m, 'role', '') != 'Admin' and not getattr(m, 'is_staff', False)]
        ids = {m.id for m in members}
        # Owned-lead counts (as STM or telecaller) and closure counts, in a few aggregates.
        lead_counts, closure_counts = {}, {}
        for fld in ('stm_id', 'telecaller_id'):
            for row in Lead.objects.filter(company=company, **{f'{fld}__in': ids}).values(fld).annotate(c=Count('id')):
                lead_counts[row[fld]] = lead_counts.get(row[fld], 0) + row['c']
        for fld in ('stm_id', 'referred_by_telecaller_id'):
            for row in Closure.objects.filter(company=company, **{f'{fld}__in': ids}).values(fld).annotate(c=Count('id')):
                closure_counts[row[fld]] = closure_counts.get(row[fld], 0) + row['c']
        data = [{
            'id':                u.id,
            'name':              u.name,
            'user_code':         u.user_code,
            'designation':       u.designation,
            'role':              u.role,
            'phone':             u.phone,
            'email':             u.email,
            'reporting_manager':    u.reporting_manager.name if u.reporting_manager_id else None,
            'reporting_manager_id': u.reporting_manager_id,
            'is_direct_report':     u.reporting_manager_id == user.id,
            'leads':             lead_counts.get(u.id, 0),
            'closures':          closure_counts.get(u.id, 0),
        } for u in members]
        return Response(data)


# ──────────────────────────────────────────────
#  Booking  (native plot booking — replaces the GAS web app for Vistara)
# ──────────────────────────────────────────────

def _loi_enabled(company):
    """LOI / EOI documents are a per-company entitlement, not a platform feature."""
    return bool(getattr(company, 'loi_enabled', False))


def _loi_path(b):
    """GAS-style object path: <Project>/Plot <no> - <Client>/R<rev>_LOI_Plot<no>_<Client>.pdf"""
    import re
    # Also strips &%#+;= — safe as literal filesystem chars, but they break Supabase's
    # signed-URL scheme (an "&" in the path produced a token whose embedded path didn't
    # match the actual object key, failing signature verification on every open —
    # confirmed against a real booking, "PARAG & SAHIL BHAI").
    san = lambda s: (re.sub(r'[\\/:*?"<>|&%#+;=]+', '', str(s or '')).strip() or 'NA')
    proj = san(b.project.name if b.project_id else 'Project')
    # EOI bookings have no plot — fall back to the EOI code held in plot_numbers.
    plot = san(b.plot.number if b.plot_id else (b.plot_numbers or b.area))
    client = san(b.client_name)
    rev = b.revision_no or 0
    return f'{proj}/Plot {plot} - {client}/R{rev}_LOI_Plot{plot}_{client}.pdf'


def _next_eoi_no(company, project_id, prefer='', block=None):
    """Next per-project EOI code. Honours a client-supplied code if it's still free,
    otherwise assigns the next available so numbers never collide.

    Default format (block=None, every pre-existing call site): EOI-1, EOI-2, …
    Block-wise industrial projects pass `block` instead — a block's own running
    number, prefixed with the block letter ('E' -> E1, E2, …) or bare if there's no
    block ('' -> 1, 2, 3…). Scoped to block_industrial projects only, so this never
    changes behaviour for any existing project."""
    prefer = (prefer or '').strip()
    if block is not None:
        import re
        prefix = block or ''
        existing = set(
            Booking.objects.filter(company=company, project_id=project_id)
            .exclude(plot_numbers='').values_list('plot_numbers', flat=True)
        )
        if prefer and prefer not in existing:
            return prefer
        pat = re.compile(rf'^{re.escape(prefix)}(\d+)$')
        used = [int(m.group(1)) for code in existing if (m := pat.match(code))]
        n = (max(used) + 1) if used else 1
        while f'{prefix}{n}' in existing:
            n += 1
        return f'{prefix}{n}'

    existing = set(
        Booking.objects.filter(company=company, project_id=project_id,
                               plot_numbers__istartswith='EOI')
        .values_list('plot_numbers', flat=True)
    )
    if prefer and prefer not in existing:
        return prefer
    n = len(existing) + 1
    while f'EOI-{n}' in existing:
        n += 1
    return f'EOI-{n}'


class BookingListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company = _resolve_company(request)
        qs = Booking.objects.filter(company=company).select_related('project', 'plot', 'stm')
        # Drafts are private scratch work, not something an approver/manager should
        # browse mid-edit — exclude anyone else's draft unconditionally, regardless of
        # which status filter (or none at all, e.g. the "All" tab) is requested. This
        # has to run before every other visibility rule below, since those exist to
        # broaden access (to a whole company, an approver's projects, etc.) and would
        # otherwise leak drafts right back in.
        qs = qs.exclude(Q(status='draft') & ~Q(stm=request.user))
        # Naming someone a project's booking approver is a narrowing statement: they
        # review those projects and no others. It therefore takes precedence over the
        # broad org-tree visibility a Manager may otherwise have (a top-of-tree head
        # like Sachin sees all company data everywhere else, but approves only the
        # projects he is named on). `?mine` is the user's own bookings list, so it is
        # left on the normal scoping or an approver loses sight of their own bookings
        # in projects they don't approve. Real admins are exempt entirely.
        approver_project_ids = [] if _is_hard_admin(request.user) else _approver_project_ids(request.user, company)
        # A Channel-Partner-sourced booking is gated by its own approver list, so
        # someone named a CP approver (but not a regular one) still needs to see
        # those bookings without gaining visibility into the project's other ones.
        cp_approver_project_ids = [] if _is_hard_admin(request.user) else _cp_approver_project_ids(request.user, company)
        # A booking counts as Channel-Partner-sourced either through its lead
        # (channel_partner FK OR Lead.source = "Channel Partner", same as
        # cp_lead_q) or its own free-text Source field — see
        # _is_cp_sourced_booking, mirrored here in query form.
        is_cp_booking_q = cp_lead_q(prefix='lead__') | Q(source__iexact='channel partner')
        if (approver_project_ids or cp_approver_project_ids) and not request.query_params.get('mine'):
            qs = qs.filter(
                (Q(project_id__in=approver_project_ids) & ~is_cp_booking_q)
                | (Q(project_id__in=cp_approver_project_ids) & is_cp_booking_q)
            )
        elif not _sees_all_company(request.user, request, include_manager_role=False):
            qs = qs.filter(stm__in=_visible_user_ids(request.user))
        if request.query_params.get('mine'):           # "My Bookings" — only this user's
            qs = qs.filter(stm=request.user)
        qs = _drop_superseded_revisions(qs)
        if request.query_params.get('closure'):
            qs = qs.filter(closure_id=request.query_params['closure'])
        if request.query_params.get('plot'):
            qs = qs.filter(plot_id=request.query_params['plot'])
        if request.query_params.get('status'):
            qs = qs.filter(status=request.query_params['status'])
        if request.query_params.get('cp_only') == 'true':
            qs = qs.filter(cp_lead_q(prefix='lead__'))
        return Response(BookingSerializer(qs, many=True).data)

    def post(self, request):
        company = _resolve_company(request)
        data = request.data

        # Reject an oversized signed LOI before any booking/lead/plot side effects run —
        # base64 inflates the raw file by ~1/3, so compare against the encoded length.
        lf_check = data.get('loi_file')
        if isinstance(lf_check, dict) and lf_check.get('data'):
            max_b64_len = int(settings.MAX_UPLOAD_FILE_MB * 1024 * 1024 * 4 / 3)
            if len(lf_check['data']) > max_b64_len:
                return Response(
                    {'detail': f'File too large (max {settings.MAX_UPLOAD_FILE_MB} MB).'},
                    status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                )

        # A booking against a project that HAS plots mapped must name one. Without
        # this the API accepted a booking with no plot, so nothing was reserved and
        # the unit map kept showing the unit as available — and because the display
        # falls back to the area, an 80,000 sq.ft parcel rendered as "Unit 80000".
        #
        # EOIs are exempt by design: they are raised before a unit is chosen. A
        # project with no plots mapped is left alone too — land sold by area (the
        # industrial projects) has no unit list to choose from, and refusing would
        # stop those sales outright.
        if not data.get('eoi'):
            proj_id = data.get('project')
            has_plot = bool(data.get('plot')) or bool(
                [x for x in (data.get('plot_ids') or []) if str(x).isdigit()])
            if proj_id and not has_plot and Plot.objects.filter(project_id=proj_id).exists():
                return Response(
                    {'detail': 'Select a unit for this booking — this project has units mapped.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Guard against duplicate submissions: a plot shouldn't have more than one
        # active (pending/approved) booking at a time. Traced real production
        # duplicates (same client/plot/amount, 10-30s apart) to a user resubmitting
        # after an unclear success state — this is the authoritative backend check
        # regardless of the exact client-side cause. Revisions (revision_of)
        # legitimately reuse the same plot, so they're excluded, as are EOIs
        # (no real plot reserved yet).
        if not data.get('revision_of') and not data.get('eoi'):
            requested_plot_ids = set()
            raw_plot_ids = data.get('plot_ids')
            if isinstance(raw_plot_ids, list) and raw_plot_ids:
                requested_plot_ids = {int(x) for x in raw_plot_ids if str(x).isdigit()}
            elif data.get('plot') and str(data['plot']).isdigit():
                requested_plot_ids = {int(data['plot'])}
            if requested_plot_ids:
                active = Booking.objects.filter(company=company, status__in=['pending', 'approved'])
                for b in active.only('id', 'plot_id', 'plot_ids', 'client_name', 'status'):
                    b_plot_ids = set(b.plot_ids or [])
                    if b.plot_id:
                        b_plot_ids.add(b.plot_id)
                    if b_plot_ids & requested_plot_ids:
                        return Response(
                            {'detail': f'This plot already has a {b.status} booking for {b.client_name} (#{b.id}).'},
                            status=status.HTTP_409_CONFLICT,
                        )
                # A plot soft-held by a DIFFERENT rep (selected on the plot-map picker
                # via PlotHoldView, not yet submitted) or already sold blocks submission
                # too — closes the gap where someone bypasses the picker's lock (stale
                # page, direct API call) and submits anyway. select_for_update so this
                # resolves consistently against a PlotHoldView call racing at the same
                # instant.
                with transaction.atomic():
                    for p in Plot.objects.select_for_update().filter(pk__in=requested_plot_ids):
                        ok = p.status == 'available' or (p.status == 'hold' and p.held_by_id == request.user.id)
                        if not ok:
                            return Response(
                                {'detail': f'Plot {p.number} is no longer available — it may have just been selected or booked by another salesperson.'},
                                status=status.HTTP_409_CONFLICT,
                            )

        # Resolve or create the lead (Book Unit flow types a new client; Record Closure
        # passes an existing lead).
        lead_id = data.get('lead') or None
        if not lead_id and (data.get('client_name') or '').strip():
            src = None
            sname = (data.get('source') or '').strip()
            if sname:
                src = LeadSource.objects.filter(company=company, name__iexact=sname).first()
            lead = Lead.objects.create(
                company=company, name=data.get('client_name', '').strip(),
                phone=(data.get('phone') or '').strip(), status='new',
                project_id=data.get('project') or None, source=src,
            )
            lead_id = lead.id

        # Revision of an existing (sold) booking — carries the prior lead, bumps the
        # revision number, and leaves the plot/closure untouched until approved.
        prior = None
        rev_of = data.get('revision_of')
        if rev_of:
            prior = Booking.objects.filter(id=rev_of, company=company).first()
            if prior:
                lead_id = prior.lead_id

        # Submitting a saved draft promotes that same row instead of creating a new
        # Booking — otherwise the draft would be left behind as an orphaned duplicate.
        draft = None
        if data.get('draft_id'):
            draft = Booking.objects.filter(id=data['draft_id'], company=company,
                                            stm=request.user, status='draft').first()

        ser = BookingSerializer(draft, data=data, partial=True) if draft else BookingSerializer(data=data)
        ser.is_valid(raise_exception=True)
        if prior:
            extra = dict(revision_no=prior.revision_no + 1, closure=prior.closure,
                         revision_of=prior,
                         approval_status='REVISION R%d PENDING' % (prior.revision_no + 1))
            # A revision inherits from its parent only what it does not supply itself.
            # `plot` used to be inherited unconditionally, which overwrote a unit the
            # revision had just chosen — and when the parent was an EOI holding no
            # plot, it overwrote that choice with nothing. Carrying the unit and area
            # across matters just as much: without them a revision of an EOI on a
            # project with no plots mapped came out blank and displayed its area as
            # though it were a plot number.
            chose_plot = bool(data.get('plot')) or bool(
                [x for x in (data.get('plot_ids') or []) if str(x).isdigit()])
            if not chose_plot:
                extra['plot'] = prior.plot
                if not str(data.get('plot_numbers') or '').strip() and prior.plot_numbers:
                    extra['plot_numbers'] = prior.plot_numbers
            if not str(data.get('area') or '').strip() and prior.area:
                extra['area'] = prior.area
        else:
            extra = dict(revision_no=0, approval_status='PENDING')
        booking = ser.save(company=company, stm=request.user, lead_id=lead_id, status='pending', **extra)

        # Multi-plot: resolve ALL selected plots. `plot` stays the primary (first);
        # plot_ids holds every selected id and plot_numbers is the comma display.
        pids = data.get('plot_ids')
        if isinstance(pids, list) and pids:
            pids = [int(x) for x in pids if str(x).isdigit()]
        elif prior and prior.plot_ids:
            pids = list(prior.plot_ids)
        elif booking.plot_id:
            pids = [booking.plot_id]
        else:
            pids = []
        if pids:
            num_map = dict(Plot.objects.filter(id__in=pids).values_list('id', 'number'))
            booking.plot_ids = pids
            booking.plot_numbers = ', '.join(num_map[p] for p in pids if p in num_map)
            if not booking.plot_id:
                booking.plot_id = pids[0]
            booking.save(update_fields=['plot_ids', 'plot_numbers', 'plot'])

        # EOI (Expression of Interest) — a booking on a project that has no plots yet
        # (raised before govt approvals). No plot is reserved; the sequential per-project
        # EOI code (EOI-1, EOI-2, …) is stored in plot_numbers so the LOI renders as an EOI.
        if data.get('eoi'):
            if prior:
                # Revising an EOI keeps the same EOI code (EOI-20 stays EOI-20).
                booking.plot_numbers = prior.plot_numbers
            else:
                # Block-prefixed numbering only for block-wise industrial projects.
                eoi_block = data.get('eoi_block') if getattr(booking.project, 'block_industrial', False) else None
                booking.plot_numbers = _next_eoi_no(company, booking.project_id,
                                                     prefer=(data.get('eoi_no') or ''), block=eoi_block)
            booking.save(update_fields=['plot_numbers'])

        # A booking on a project with no units mapped, submitted without a unit and
        # without the EOI flag, would otherwise carry no identity at all — the list
        # then falls back to its area, which is how "80,000 sq.ft" once rendered as
        # "Unit 80000". There is no other identity available on such a project, so it
        # is numbered as the EOI it effectively is. Projects that DO have units are
        # already refused above unless one is named.
        if (not data.get('eoi') and not booking.plot_numbers and not booking.plot_id
                and booking.project_id
                and not Plot.objects.filter(project_id=booking.project_id).exists()):
            booking.plot_numbers = _next_eoi_no(company, booking.project_id)
            booking.save(update_fields=['plot_numbers'])

        # Signed LOI (sent as base64 {name,type,data}). Stored GAS-style:
        # <Project>/Plot <no> - <Client>/R<rev>_LOI_Plot<no>_<Client>.pdf
        lf = data.get('loi_file')
        if isinstance(lf, dict) and lf.get('data') and not _loi_enabled(company):
            return Response(
                {'detail': 'LOI / EOI documents are not enabled for this company.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        if isinstance(lf, dict) and lf.get('data'):
            import base64
            from django.core.files.base import ContentFile
            try:
                booking.loi_document.save(_loi_path(booking),
                                          ContentFile(base64.b64decode(lf['data'])), save=True)
            except Exception:
                # The file may already be in storage (e.g. the storage POST succeeded but the
                # model save failed, or the request timed out). Persist the deterministic path
                # so the signed LOI isn't orphaned/invisible, and surface the error in logs.
                import logging
                logging.getLogger(__name__).exception('LOI document save failed for booking %s', booking.id)
                try:
                    booking.loi_document.name = _loi_path(booking)
                    booking.save(update_fields=['loi_document'])
                except Exception:
                    logging.getLogger(__name__).exception('LOI document path relink failed for booking %s', booking.id)

        if not prior:
            # New booking: reserve ALL selected plots. The Closure is mirrored into
            # My Conversions on APPROVAL (see BookingActionView) — so a booking that
            # is still pending approval does NOT appear as a booked closure.
            # held_by/held_at are cleared here — this is now a hard hold backed by a
            # real pending Booking, not the picker's soft hold, so it never auto-expires.
            if pids:
                Plot.objects.filter(id__in=pids).update(status='hold', held_by=None, held_at=None)

        # Notify the admin-selected approvers (managers) via push.
        _notify_booking_approvers(company, booking, request.user)
        # Accounts & Finance follow the money from the moment it is submitted, not
        # only once it clears — a revised LOI changes the figure they are tracking.
        _notify_accounts_booking(
            company, booking, 'booking_approval',
            'Booking revised — approval pending' if booking.revision_no else 'New booking — approval pending',
            'awaiting approval')

        return Response(BookingSerializer(booking).data, status=status.HTTP_201_CREATED)


class BookingDraftView(APIView):
    """Save an in-progress booking as a draft — same payload shape as
    BookingListCreateView.post, but with none of its completeness requirements (no
    signed LOI, no 100%-installment check — those are the frontend's job to enforce
    only when calling Submit, not Save). Lets a rep persist partially-filled work so
    closing the browser mid-flow doesn't lose it. Pass `id` to update an existing
    draft in place rather than creating a new row on every Save click."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        company = _resolve_company(request)
        data = request.data

        draft = None
        if data.get('id'):
            draft = Booking.objects.filter(id=data['id'], company=company,
                                            stm=request.user, status='draft').first()
            if not draft:
                return Response({'detail': 'Draft not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Resolve or create the lead — reuse the draft's existing lead on repeat
        # Saves instead of minting a new one every time the rep clicks Save.
        lead_id = data.get('lead') or (draft.lead_id if draft else None)
        if not lead_id and (data.get('client_name') or '').strip():
            src = None
            sname = (data.get('source') or '').strip()
            if sname:
                src = LeadSource.objects.filter(company=company, name__iexact=sname).first()
            lead = Lead.objects.create(
                company=company, name=data.get('client_name', '').strip(),
                phone=(data.get('phone') or '').strip(), status='new',
                project_id=data.get('project') or None, source=src,
            )
            lead_id = lead.id

        ser = BookingSerializer(draft, data=data, partial=True) if draft else BookingSerializer(data=data)
        ser.is_valid(raise_exception=True)
        booking = ser.save(company=company, stm=request.user, lead_id=lead_id, status='draft')

        # Resolve selected plots the same way BookingListCreateView.post does.
        pids = data.get('plot_ids')
        if isinstance(pids, list) and pids:
            pids = [int(x) for x in pids if str(x).isdigit()]
        elif booking.plot_id:
            pids = [booking.plot_id]
        else:
            pids = []

        plot_conflicts = []
        if pids:
            num_map = dict(Plot.objects.filter(id__in=pids).values_list('id', 'number'))
            booking.plot_ids = pids
            booking.plot_numbers = ', '.join(num_map[p] for p in pids if p in num_map)
            if not booking.plot_id:
                booking.plot_id = pids[0]
            booking.save(update_fields=['plot_ids', 'plot_numbers', 'plot'])

            # Claim any plot that's free; never fail the whole save over one that
            # isn't — losing typed data is worse than a stale plot reference. Flag
            # the conflict instead so the frontend can warn without discarding anything.
            with transaction.atomic():
                for plot in Plot.objects.select_for_update().filter(pk__in=pids):
                    if plot.status == 'available':
                        plot.status, plot.held_by, plot.held_at = 'hold', request.user, timezone.now()
                        plot.save(update_fields=['status', 'held_by', 'held_at'])
                    elif not (plot.status == 'hold' and plot.held_by_id == request.user.id):
                        plot_conflicts.append({'id': plot.id, 'number': plot.number})

        # Signed LOI, if one happens to already be attached — same handling as the
        # real submit path, kept for forward compatibility even though drafts don't
        # require it.
        lf = data.get('loi_file')
        if isinstance(lf, dict) and lf.get('data') and _loi_enabled(booking.company):
            max_b64_len = int(settings.MAX_UPLOAD_FILE_MB * 1024 * 1024 * 4 / 3)
            if len(lf['data']) <= max_b64_len:
                import base64
                from django.core.files.base import ContentFile
                try:
                    booking.loi_document.save(_loi_path(booking),
                                              ContentFile(base64.b64decode(lf['data'])), save=True)
                except Exception:
                    logging.getLogger(__name__).exception('LOI document save failed for draft %s', booking.id)

        resp = BookingSerializer(booking).data
        resp['plot_conflicts'] = plot_conflicts
        return Response(resp, status=status.HTTP_200_OK)


class BookingDiscardDraftView(APIView):
    """Discard a saved draft — releases any plots it still holds and deletes the row.
    Irreversible. The drafter can discard their own; a real Admin can discard anyone's;
    a Manager can discard one belonging to an STM in their own reporting chain, not
    just any Manager company-wide (e.g. from the plot map, where a drafted unit's name
    is visible to the whole team even though the draft's own details aren't)."""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        company = _resolve_company(request)
        try:
            b = Booking.objects.get(pk=pk, company=company, status='draft')
        except Booking.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        # _visible_user_ids already includes the requester themselves, so this alone
        # covers "the drafter", "an admin" is the only other unconditional case, and a
        # Manager only clears it when the drafting STM is actually in their own
        # reporting subtree — not any Manager company-wide.
        if not _is_hard_admin(request.user) and b.stm_id not in _visible_user_ids(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        pids = b.plot_ids or ([b.plot_id] if b.plot_id else [])
        if pids:
            Plot.objects.filter(id__in=pids).update(status='available', held_by=None, held_at=None)
        b.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BookingNextEOIView(APIView):
    """Preview the next per-project EOI code so the form + LOI can show it before submit."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company = _resolve_company(request)
        pid = request.query_params.get('project')
        if not pid:
            return Response({'detail': 'project is required'}, status=status.HTTP_400_BAD_REQUEST)
        # Block-prefixed numbering only applies to block-wise industrial projects — a
        # stray `block` param on any other project is ignored so nothing else changes.
        project = Project.objects.filter(id=pid).only('block_industrial').first()
        block = request.query_params.get('block') if (project and project.block_industrial) else None
        return Response({'eoi_no': _next_eoi_no(company, pid, block=block)})


def _drop_superseded_revisions(qs):
    """Hide bookings that a later revision has replaced, so a deal appears once.

    Revising a booking creates a NEW row carrying revision_no + 1, and approving that
    revision leaves the original approved as well — so both were listed and the project
    totals counted the deal twice. Only the latest revision should stand.

    `revision_of` is the authoritative link and is followed first. It has only been
    recorded since the field was added, so two older facts still stand in for it, and
    neither works alone:

      * they share a closure — which survives an EOI being converted to an LOI, where
        the unit is renumbered from "EOI-2" to a real plot number; but
      * a revision is issued its own closure, so a plain revise leaves the closures
        different while the unit stays the same.

    So rows are connected if they share EITHER a closure OR a (project, phone, unit),
    and the connections are followed transitively — a chain that was revised twice and
    then converted still resolves to one deal.

    A group is only collapsed when it actually contains a revision, so two ordinary
    bookings that happen to share a key are left alone; rejected rows are excluded,
    belonging in the Rejected tab rather than folded into a live chain. Ties on
    revision_no fall to the newest row, which happens where a booking was revised twice
    from the same parent.
    """
    rows = [r for r in qs.values('id', 'project_id', 'phone', 'plot_numbers', 'plot__number',
                                 'area', 'revision_no', 'status', 'closure_id', 'revision_of_id')
            if r['status'] != 'rejected']

    parent = {r['id']: r['id'] for r in rows}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[max(ra, rb)] = min(ra, rb)

    present = {r['id'] for r in rows}
    seen = {}
    for r in rows:
        # The recorded parent, where there is one — exact, no inference needed.
        if r['revision_of_id'] and r['revision_of_id'] in present:
            union(r['revision_of_id'], r['id'])
        keys = []
        if r['closure_id']:
            keys.append(('closure', r['closure_id']))
        unit = (r['plot_numbers'] or r['plot__number'] or r['area'] or '').strip()
        if unit:
            keys.append(('unit', r['project_id'], (r['phone'] or '').strip(), unit))
        for k in keys:
            if k in seen:
                union(seen[k], r['id'])
            else:
                seen[k] = r['id']

    groups = {}
    for r in rows:
        groups.setdefault(find(r['id']), []).append(r)
    drop = set()
    for g in groups.values():
        if len(g) < 2 or not any((x['revision_no'] or 0) > 0 for x in g):
            continue
        keep = max(g, key=lambda x: ((x['revision_no'] or 0), x['id']))
        drop.update(x['id'] for x in g if x['id'] != keep['id'])
    return qs.exclude(id__in=drop) if drop else qs


def _can_view_all_bookings(user):
    """Whole-company booking visibility: company-wide viewers (admin/staff/dept head)
    plus the Accounts & Finance department (read-only review of LOIs/EOIs). The
    Manager role is excluded — bookings stay scoped by approver assignment."""
    if _sees_all_company(user, include_manager_role=False):
        return True
    mods = [str(m).lower() for m in (getattr(user, 'modules', None) or [])]
    return any('account' in m or 'finance' in m for m in mods)


class BookingAllView(APIView):
    """Read-only: ALL company bookings (LOI + EOI) for authorised viewers — used by the
    Accounts & Finance module to review booking / LOI / EOI details. No create or edit."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _can_view_all_bookings(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        company = _resolve_company(request)
        qs = _drop_superseded_revisions(
            Booking.objects.filter(company=company).select_related('project', 'plot', 'stm')
        ).order_by('-created_at')
        return Response(BookingSerializer(qs[:1000], many=True).data)


def _notify_closure_cancellation(stm, project_obj, company, unit, client, amount, canceller, extra_data=None):
    """Notify STM + approver/manager chain when a closure is cancelled.
    Accepts pre-extracted primitive values so it is safe to call after the closure is deleted."""
    try:
        from notifications import notify, reporting_chain
        canceller_name = getattr(canceller, 'name', '')
        canceller_id   = getattr(canceller, 'id', None)
        stm_id         = getattr(stm, 'id', None)
        project_name   = getattr(project_obj, 'name', '') or ''
        data           = extra_data or {}

        # 1. Notify the STM (skip if they are the one cancelling).
        if stm and stm_id != canceller_id:
            notify(
                stm, 'booking_cancelled',
                'Booking Cancelled',
                f'{client} · {project_name} Unit {unit} has been cancelled.',
                data,
            )

        # 2. Notify project approvers → STM's reporting chain → all managers/admins (first non-empty).
        approver_ids = (getattr(project_obj, 'booking_approvers', None) or [])
        recipients = list(User.objects.filter(id__in=approver_ids, company=company, is_active=True)) if approver_ids else []
        if not recipients and stm:
            recipients = reporting_chain(stm)
        if not recipients:
            recipients = list(
                User.objects.filter(company=company, is_active=True)
                .filter(Q(role__in=MANAGER_ROLES) | Q(is_staff=True))
            )
        seen = set()
        for u in recipients:
            if u and u.id not in seen and u.id != canceller_id and u.id != stm_id:
                seen.add(u.id)
                notify(
                    u, 'booking_cancelled',
                    'Booking Cancelled',
                    f'{client} · {project_name} Unit {unit} · ₹{amount} — cancelled by {canceller_name}',
                    data,
                )
    except Exception:
        import logging
        logging.getLogger(__name__).exception('_notify_closure_cancellation failed')


def _notify_booking_approvers(company, booking, submitter):
    try:
        from notifications import notify, reporting_chain
        # A Channel-Partner-sourced booking routes to the project's CP approver
        # list instead of its regular one — same split the approve/reject
        # endpoint and the list scoping enforce (see _can_approve_booking).
        is_cp = _is_cp_sourced_booking(booking.lead_id, booking.source)
        can_approve_fn = _can_approve_cp_project if is_cp else _can_approve_project
        # 1) Per-project configured approvers (most precise).
        approver_field = 'cp_booking_approvers' if is_cp else 'booking_approvers'
        ids = (getattr(booking.project, approver_field, None) if booking.project_id else None) or []
        recipients = list(User.objects.filter(id__in=ids, company=company, is_active=True)) if ids else []
        # 2) Fallback: the submitting STM's reporting-manager chain.
        if not recipients and booking.stm_id:
            recipients = reporting_chain(booking.stm)
        # 3) Last resort: every manager/admin in the company (so it's never silent).
        #    role='Admin' is included explicitly — a company admin need not be is_staff,
        #    and without this the last resort skipped exactly the people who can always
        #    approve.
        if not recipients:
            recipients = list(User.objects.filter(company=company, is_active=True)
                              .filter(Q(role__in=MANAGER_ROLES) | Q(is_staff=True)))
        # Never notify the person who submitted it; de-dup.
        sub_id = getattr(submitter, 'id', None)
        recipients = [u for u in recipients if u and u.id != sub_id]
        # This is a request to *act*, so it must reach only people who actually can:
        # the same authority the approve/reject endpoint enforces. Without this, the
        # tier-2/3 fallbacks would ask a manager to approve a project they are not an
        # approver for -- they would tap the notification and get a 403 -- and would
        # ask non-managers, who cannot approve at all.
        if booking.project_id:
            recipients = [
                u for u in recipients
                if is_admin_or_manager(u) and can_approve_fn(u, booking.project, company)
            ]
            # Never let the filter silence the request entirely: if no one qualifies,
            # fall back to the company admins, who can approve any project.
            if not recipients:
                recipients = [u for u in User.objects.filter(company=company, is_active=True)
                              if _is_hard_admin(u) and u.id != sub_id]
        if not recipients:
            return
        unit = booking.plot_numbers or (booking.plot.number if booking.plot_id else booking.area)
        rev = (' (R%d)' % booking.revision_no) if booking.revision_no else ''
        # The rep who sold the unit should hear about it even when someone else
        # submitted the revision on their behalf. A revision is saved with the
        # submitter as its stm, so the original rep is the one on the parent booking.
        owner = booking.stm
        if booking.revision_of_id and booking.revision_of and booking.revision_of.stm_id:
            owner = booking.revision_of.stm
        if owner and owner.id != sub_id and owner.id not in {u.id for u in recipients}:
            recipients = recipients + [owner]
        title = 'Booking approval needed%s' % rev
        msg = '%s · %s Unit %s · ₹%s — by %s' % (
            booking.client_name or '—', booking.project.name if booking.project_id else '',
            unit, int(booking.final_amount or 0), getattr(submitter, 'name', ''),
        )
        seen = set()
        for u in recipients:
            if u.id not in seen:
                seen.add(u.id)
                notify(u, 'booking_approval', title, msg, {'booking_id': booking.id})
    except Exception:
        pass


def _notify_accounts_booking(company, booking, ntype, title, suffix):
    """Tell the Accounts & Finance managers about a booking event."""
    unit = booking.plot_numbers or (booking.plot.number if booking.plot_id else booking.area)
    rev = (' (R%d)' % booking.revision_no) if booking.revision_no else ''
    _notify_accounts_managers(
        company, ntype, '%s%s' % (title, rev),
        '%s · %s Unit %s · Rs %s — %s' % (
            booking.client_name or 'Booking',
            booking.project.name if booking.project_id else '',
            unit, int(booking.final_amount or 0), suffix,
        ),
        {'booking_id': booking.id},
    )


def _notify_accounts_managers(company, ntype, title, body, data=None):
    """Notify every active manager of the Accounts & Finance module (managers have
    'Accounts & Finance' in their manager_modules). Best-effort — never raises."""
    try:
        from notifications import notify
        recipients = [
            u for u in User.objects.filter(company=company, is_active=True)
            if 'Accounts & Finance' in (getattr(u, 'manager_modules', None) or [])
        ]
        seen = set()
        for u in recipients:
            if u.id not in seen:
                seen.add(u.id)
                notify(u, ntype, title, body, data or {})
    except Exception:
        import logging
        logging.getLogger(__name__).exception('_notify_accounts_managers failed')


def _ensure_lead_and_site_visit_for_booking(b):
    """On a booking's first approval, guarantee it exists in the pipeline as a lead
    with a completed site visit dated on the booking date.

    Two flows land here. A closure recorded from a lead already has the lead but may
    have no visit — a walk-in that booked without one ever being logged. A unit booked
    directly has a lead created at submission time, but likewise no visit. Either way
    the sale is real and the visit demonstrably happened, so the pipeline should say so
    rather than showing a closure that came from nowhere.

    Deliberately conservative:
      - Never duplicates. A completed visit already on this lead for this project is
        left exactly as it is, which is the normal Record-Closure-from-a-visit path.
      - Only fabricates a lead when there is a name or phone to build one from.
      - Attributes to the booking's STM, and to the lead's telecaller where there is
        one, so the visit counts for the same people the closure does.
    """
    if not b.booking_date:
        return None, None                     # nothing to date the visit by

    lead_id = b.lead_id
    if not lead_id:
        name  = (b.client_name or '').strip()
        phone = (b.phone or '').strip()
        if not (name or phone):
            return None, None                 # no identity to build a lead from
        # Match on the number first. The same client often already exists as a lead —
        # they were called, or they booked a second unit — and creating another record
        # would split one person's history across two leads. Phone is encrypted, so the
        # lookup goes through the blind index, which normalises to the last ten digits.
        existing = None
        key = phone_blind_index(phone) if phone else ''
        if key:
            existing = (Lead.objects.filter(company_id=b.company_id, phone_key=key)
                        .order_by('id').first())
        if existing:
            lead_id = existing.id
            # Attach the sale to them without overwriting a working history: only fill
            # the STM in, and only when nobody is on it.
            fields = {}
            if not existing.stm_id and b.stm_id:
                fields['stm_id'] = b.stm_id
            if fields:
                Lead.objects.filter(pk=lead_id).update(**fields)
        else:
            lead = Lead.objects.create(
                company_id=b.company_id, name=name, phone=phone, status='closed',
                project_id=b.project_id, stm=b.stm, stm_status='closed',
            )
            lead_id = lead.id
        Booking.objects.filter(pk=b.pk).update(lead_id=lead_id)
        b.lead_id = lead_id

    # The guard is the booking DATE, not merely "this lead has been on a visit".
    # A repeat buyer visited once per unit they bought, so a visit already logged on
    # some other day belongs to the other sale and is left alone while this booking
    # still gets its own. Only a visit already sitting on this booking's date means
    # there is nothing to add.
    already = SiteVisit.objects.filter(
        lead_id=lead_id, project_id=b.project_id, status='completed',
        visited_at__date=b.booking_date).exists()
    if already:
        return lead_id, None

    # booking_date is a date; visits are timestamped, so anchor it at midday local
    # time — a plain midnight can land on the previous day once rendered in another
    # timezone, which would put the visit before the booking it came from.
    visited = datetime.combine(b.booking_date, dt_time(12, 0))
    if timezone.is_naive(visited):
        visited = timezone.make_aware(visited, timezone.get_current_timezone())

    tc_id = Lead.objects.filter(pk=lead_id).values_list('telecaller_id', flat=True).first()
    sv = SiteVisit.objects.create(
        lead_id=lead_id, project_id=b.project_id, stm=b.stm,
        referred_by_telecaller_id=tc_id or None,
        scheduled_at=visited, visited_at=visited,
        status='completed', outcome='hot',
        remarks='Recorded automatically from booking #%s on approval.' % b.pk,
    )
    return lead_id, sv.id


class BookingActionView(APIView):
    """Approve / reject a pending booking (approver = admin or manager)."""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        company = _resolve_company(request)
        try:
            b = Booking.objects.get(pk=pk, company=company)
        except Booking.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        # Enforced here and not merely hidden in the list, or a manager could approve
        # another project's booking straight through the API. Two cases, mirroring the
        # list scoping above: someone named on any project is confined to those projects
        # (a project they don't approve is off limits even if it names nobody), and
        # someone named nowhere is blocked from projects that do name approvers.
        # Real admins are exempt.
        if not _can_approve_booking(request.user, b.project_id, b.project, b.lead_id, company, b.source):
            return Response(
                {'detail': 'You are not a booking approver for this project.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        action = request.data.get('action')
        is_rev = b.revision_no and b.revision_no > 0

        if action == 'approve':
            _pids = b.plot_ids or ([b.plot_id] if b.plot_id else [])
            if _pids:
                Plot.objects.filter(id__in=_pids).update(status='sold')
            b.status = 'sold'
            b.approval_status = ('REVISION R%d APPROVED' % b.revision_no) if is_rev else 'APPROVED'
            b.approved_at = timezone.now()
            if b.closure_id:
                # Existing closure (revision / re-approval) → just sync the amounts.
                b.save(update_fields=['status', 'approval_status', 'approved_at'])
                Closure.objects.filter(id=b.closure_id).update(
                    booking_amount=b.plot_basic or None, total_amount=b.final_amount or None)
            else:
                # First approval of a new booking → mirror it into My Conversions now.
                if b.lead_id:
                    Lead.objects.filter(id=b.lead_id).update(stm=b.stm, stm_status='closed')
                closure = Closure.objects.create(
                    company_id=b.company_id, lead_id=b.lead_id, project_id=b.project_id, stm=b.stm,
                    client_name=b.client_name or '', client_phone=b.phone or '',
                    status='booked', closure_date=b.booking_date or timezone.now().date(),
                    unit_no=(b.plot_numbers or (b.plot.number if b.plot_id else b.area)),
                    unit_type=b.villa_type or b.bunglow_type or '',
                    booking_amount=b.plot_basic or None, total_amount=b.final_amount or None,
                )
                b.closure = closure
                b.save(update_fields=['status', 'approval_status', 'approved_at', 'closure'])
                # The sale is now real, so make sure the pipeline shows how it got
                # here: a lead, and a completed site visit dated on the booking date.
                try:
                    _lid, _svid = _ensure_lead_and_site_visit_for_booking(b)
                    if _svid and not closure.lead_id and _lid:
                        Closure.objects.filter(pk=closure.pk).update(lead_id=_lid)
                except Exception:
                    # Never let this block an approval — the booking is what matters.
                    logger.exception('Could not back-fill lead/site visit for booking %s', b.pk)
            # Notify the STM (approved) and — on a fresh closure — their manager chain.
            from notifications import notify, notify_many, reporting_chain
            _unit = (b.plot_numbers or (b.plot.number if b.plot_id else b.area))
            _rev = (' (R%d)' % b.revision_no) if is_rev else ''
            if b.stm:
                notify(b.stm, 'booking_approved', 'Booking Approved%s' % _rev,
                       f'{b.client_name or "Your booking"} · Unit {_unit} was approved.', {'booking_id': b.id})
                if not is_rev:
                    notify_many(reporting_chain(b.stm), 'closure', 'New Closure',
                                f'{b.stm.name} closed {b.client_name or "a unit"} · Unit {_unit} · ₹{int(b.final_amount or 0)}',
                                {'booking_id': b.id})
            # Notify Accounts & Finance managers of every approved booking.
            _notify_accounts_managers(
                company, 'booking_approved', 'Booking Approved%s' % _rev,
                f'{b.client_name or "Booking"} · {b.project.name if b.project_id else ""} Unit {_unit} · ₹{int(b.final_amount or 0)} — approved',
                {'booking_id': b.id},
            )
        elif action == 'reject':
            b.status = 'rejected'
            b.approval_status = ('REVISION R%d REJECTED' % b.revision_no) if is_rev else 'REJECTED'
            # Remove the rejected signed LOI PDF from Supabase storage.
            if b.loi_document:
                try: b.loi_document.delete(save=False)
                except Exception: pass
            b.save(update_fields=['status', 'approval_status', 'loi_document'])
            if not is_rev:
                _pids = b.plot_ids or ([b.plot_id] if b.plot_id else [])
                if _pids:
                    Plot.objects.filter(id__in=_pids).update(status='available')
                if b.closure_id:
                    Closure.objects.filter(id=b.closure_id).delete()
            from notifications import notify
            _unit = (b.plot_numbers or (b.plot.number if b.plot_id else b.area))
            _rev = (' (R%d)' % b.revision_no) if is_rev else ''
            if b.stm:
                _notify_accounts_booking(company, b, 'booking_rejected', 'Booking Rejected', 'rejected')
                notify(b.stm, 'booking_rejected', 'Booking Rejected%s' % _rev,
                       f'{b.client_name or "Your booking"} · Unit {_unit} was rejected.', {'booking_id': b.id})
        else:
            return Response({'detail': 'action must be approve or reject.'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(BookingSerializer(b).data)


class BookingLOIUrlView(APIView):
    """Returns a short-lived signed URL for a booking's confidential LOI PDF.
    Authorised viewers only (admin/manager or the booking's STM). The bucket is
    private, so this signed URL is the *only* way to open the document."""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        b = scope_to_company(Booking.objects.all(), request.user).filter(pk=pk).first()
        if not b or not b.loi_document:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if not _loi_enabled(b.company):
            return Response(
                {'detail': 'LOI / EOI documents are not enabled for this company.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        if not (is_admin_or_manager(request.user) or b.stm_id == request.user.id):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        from sales.supabase_storage import create_signed_url
        url = create_signed_url(b.loi_document.name, expires_in=120)
        if not url:
            # Local dev (FileSystem storage) fallback.
            try:
                url = request.build_absolute_uri(b.loi_document.url)
            except Exception:
                url = None
        if not url:
            return Response({'detail': 'LOI unavailable.'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'url': url})


# Largest media file accepted by MediaUploadView. Architects' floor-plan PDFs run
# well past the old 25 MB. Keep the web/app pickers' own limits in step with this —
# they check client-side purely to fail fast before the upload starts.
MEDIA_UPLOAD_MAX_MB = 100


class MediaUploadView(APIView):
    """Authenticated media upload to the public erp-media bucket via the service-role
    key. Lets the frontend stop using the anon key for writes (so anon INSERT can be
    revoked in Supabase). Returns {url, path}."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import time, random, string
        from sales.supabase_storage import upload_public
        f = request.FILES.get('file')
        if not f:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
        if f.size and f.size > MEDIA_UPLOAD_MAX_MB * 1024 * 1024:
            return Response({'detail': f'File too large (max {MEDIA_UPLOAD_MAX_MB} MB).'}, status=status.HTTP_400_BAD_REQUEST)
        folder = (request.data.get('folder') or 'erp/media').strip('/')
        ext = (f.name.rsplit('.', 1)[-1].lower() if '.' in (f.name or '') else 'bin')[:10]
        rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=8))
        path = f'{folder}/{int(time.time() * 1000)}_{rand}.{ext}'
        try:
            url = upload_public(f.read(), path, f.content_type or 'application/octet-stream')
        except Exception as e:
            return Response({'detail': str(e)[:200]}, status=status.HTTP_502_BAD_GATEWAY)
        if not url:
            return Response({'detail': 'Storage not configured.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        return Response({'url': url, 'path': path})


class MediaDeleteView(APIView):
    """Delete a media object from erp-media via the service-role key (anon can't)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from sales.supabase_storage import delete_object
        path = request.data.get('path')
        if not path:
            return Response({'detail': 'path required.'}, status=status.HTTP_400_BAD_REQUEST)
        delete_object(path)
        return Response({'ok': True})


class ClosureCancelView(APIView):
    """Cancel a closure: deletes the closure, frees the plot(s), removes the
    signed LOI PDFs from Supabase, and marks the related booking(s) cancelled."""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        closure = scope_to_company(
            Closure.objects.filter(pk=pk).select_related('stm', 'project', 'lead'),
            request.user, 'company').first()
        if not closure:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        # Only an approver (admin/manager) may cancel a booking — same authority that
        # approves/rejects it. The owning STM can no longer self-cancel.
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Only an approver can cancel a booking.'}, status=status.HTTP_403_FORBIDDEN)
        company = _resolve_company(request)
        # ...and only for a project they actually approve. Cancelling undoes an approval,
        # frees the plots and deletes the signed LOI, so it cannot be laxer than approving.
        # A closure has no Source field of its own — pull it from the booking that
        # was approved into this closure, if any, so a CP-sourced deal still routes
        # to the CP approvers at cancel time too.
        booking_source = Booking.objects.filter(closure_id=closure.pk).values_list('source', flat=True).first()
        if not _can_approve_booking(request.user, closure.project_id, closure.project, closure.lead_id, company, booking_source):
            return Response({'detail': 'You are not a booking approver for this project.'},
                            status=status.HTTP_403_FORBIDDEN)

        # Extract all notification data BEFORE deletion (closure.pk becomes None after delete).
        notif_stm      = closure.stm
        notif_project  = closure.project
        notif_unit     = (closure.unit_type + ' ' + (closure.unit_no or '')).strip() or '—'
        notif_client   = getattr(closure.lead, 'name', None) or closure.client_name or '—'
        notif_amount   = int(closure.total_amount or 0)
        notif_extra    = {'closure_id': closure.pk}

        linked_booking = Booking.objects.filter(closure=closure).only('id').first()
        if linked_booking:
            notif_extra['booking_id'] = linked_booking.pk

        for b in Booking.objects.filter(closure=closure):
            if b.loi_document:
                try: b.loi_document.delete(save=False)
                except Exception: pass
            _pids = b.plot_ids or ([b.plot_id] if b.plot_id else [])
            if _pids:
                Plot.objects.filter(id__in=_pids).update(status='available')
            b.status = 'rejected'
            b.approval_status = 'CANCELLED'
            b.save(update_fields=['status', 'approval_status', 'loi_document'])
        if closure.lead_id:
            Lead.objects.filter(id=closure.lead_id).update(stm_status='')
        closure.delete()

        _notify_closure_cancellation(
            notif_stm, notif_project, company,
            notif_unit, notif_client, notif_amount,
            request.user, extra_data=notif_extra,
        )
        # Notify Accounts & Finance managers of every cancelled booking.
        _notify_accounts_managers(
            company, 'booking_cancelled', 'Booking Cancelled',
            f'{notif_client} · {getattr(notif_project, "name", "") or ""} Unit {notif_unit} · ₹{notif_amount} — cancelled by {getattr(request.user, "name", "")}',
            notif_extra,
        )
        return Response({'detail': 'Closure cancelled.'})


# ──────────────────────────────────────────────
#  Meta Lead Ads Webhook
# ──────────────────────────────────────────────

def _fetch_meta_lead_data(leadgen_id, page_access_token):
    """Call Meta Graph API to get lead field data and ad info."""
    try:
        url = f'https://graph.facebook.com/v19.0/{leadgen_id}'
        r = http_requests.get(url, params={
            'access_token': page_access_token,
            # form_id decides project routing — fetch it authoritatively here so we
            # don't depend on the webhook payload always including it.
            'fields': 'field_data,ad_id,ad_name,form_id',
        }, timeout=10)
        if r.status_code == 200:
            return r.json()
    except Exception:
        logger.exception('Meta: failed to fetch lead data for leadgen_id=%s', leadgen_id)
    return None


def _fetch_ad_campaign_info(ad_id, page_access_token):
    """Given an ad_id, fetch campaign name and adset name from Meta Graph API."""
    if not ad_id:
        return '', ''
    try:
        url = f'https://graph.facebook.com/v19.0/{ad_id}'
        r = http_requests.get(url, params={
            'access_token': page_access_token,
            'fields': 'campaign{name},adset{name}',
        }, timeout=10)
        if r.status_code == 200:
            data = r.json()
            campaign_name = (data.get('campaign') or {}).get('name', '')
            adset_name    = (data.get('adset') or {}).get('name', '')
            return campaign_name, adset_name
    except Exception:
        logger.exception('Meta: failed to fetch campaign info for ad_id=%s', ad_id)
    return '', ''


def _create_lead_from_meta(field_data, config, campaign_name='', adset_name='', ad_name='', form_id=''):
    """Parse Meta field_data list and create a Lead."""
    fields = {f['name']: f['values'][0] for f in field_data if f.get('values') and f.get('name')}
    name  = fields.get('full_name') or fields.get('name') or (fields.get('first_name', '') + ' ' + fields.get('last_name', '')).strip()
    phone = (fields.get('phone_number') or fields.get('phone') or '').strip()[:20]
    email = fields.get('email', '')[:254]
    if not name and not phone:
        return None

    # Resolve project: form mapping takes priority over default
    project = config.default_project
    if form_id:
        mapping = MetaFormMapping.objects.filter(form_id=form_id).select_related('project').first()
        if mapping:
            project = mapping.project
            MetaFormMapping.objects.filter(pk=mapping.pk).update(total_leads=mapping.total_leads + 1)

    # Tenant for the incoming lead: project's company → config's company
    company = (project.company if project and project.company_id else None) or config.company
    if company is None:
        return None  # Can't attribute to a tenant — drop rather than leak globally.

    source, _ = LeadSource.objects.get_or_create(
        company=company, name='meta', defaults={'is_active': True},
    )

    # Duplicate detection using last 10 digits, scoped to this company
    clean = ''.join(c for c in phone if c.isdigit())[-10:]
    existing = (
        Lead.objects.filter(company=company, phone_key=phone_blind_index(clean)).first()
        if clean else None
    )
    if existing:
        existing.duplicate_count = (existing.duplicate_count or 0) + 1
        existing.save(update_fields=['duplicate_count'])

    lead = Lead.objects.create(
        company=company,
        name=(name or 'Meta Lead')[:200],
        phone=phone,
        email=email,
        source=source,
        project=project,
        meta_campaign_name=campaign_name[:200] if campaign_name else '',
        meta_adset_name=adset_name[:200] if adset_name else '',
        meta_ad_name=ad_name[:200] if ad_name else '',
        meta_form_id=str(form_id or '')[:100],
        status='new',
        is_duplicate=bool(existing),
        duplicate_of=existing if existing else None,
    )
    MetaWebhookConfig.objects.filter(pk=config.pk).update(
        total_leads_received=config.total_leads_received + 1,
        last_lead_at=timezone.now(),
        is_active=True,
    )
    _record_lead_created(lead)  # source = 'meta'
    # Auto-assign the live lead to an available telecaller (window-gated).
    _run_distribution(company, 'telecaller')
    return lead


class MetaWebhookView(APIView):
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        """Meta webhook verification challenge."""
        mode      = request.GET.get('hub.mode')
        token     = request.GET.get('hub.verify_token')
        challenge = request.GET.get('hub.challenge')
        # Match any company's verify token (each tenant has its own config).
        if mode == 'subscribe' and token and MetaWebhookConfig.objects.filter(verify_token=token).exists():
            return HttpResponse(challenge, content_type='text/plain')
        return HttpResponse(status=403)

    def _config_for_page(self, page_id):
        """Find the tenant config that owns the given Meta page id."""
        configs = list(MetaWebhookConfig.objects.filter(page_access_token__gt=''))
        if page_id:
            for cfg in configs:
                for p in (cfg.pages_data or []):
                    if str(p.get('page_id')) == str(page_id):
                        return cfg
        return configs[0] if configs else None

    @staticmethod
    def _signature_ok(request, app_secret):
        """Verify Meta's X-Hub-Signature-256 over the raw request body.

        Meta signs every delivery with HMAC-SHA256 keyed on the app secret. Without
        this the endpoint is an open door: anyone who learns a page id could post a
        payload and make the ERP call the Graph API with that page's token.

        A config with no app_secret is not rejected -- that would silently drop real
        leads for a tenant mid-setup -- but it is logged so the gap is visible.
        """
        if not app_secret:
            logger.warning('Meta webhook: no app_secret configured — delivery accepted unverified')
            return True
        header = request.headers.get('X-Hub-Signature-256', '')
        if not header.startswith('sha256='):
            return False
        import hashlib, hmac
        expected = hmac.new(app_secret.encode(), request.body, hashlib.sha256).hexdigest()
        return hmac.compare_digest(expected, header[len('sha256='):])

    def post(self, request):
        """Receive lead notification from Meta."""
        try:
            data = request.data
            if data.get('object') != 'page':
                return Response({'ok': True})
            for entry in data.get('entry', []):
                config = self._config_for_page(entry.get('id'))
                if not config:
                    continue
                if not self._signature_ok(request, config.app_secret):
                    logger.warning('Meta webhook: bad signature for page %s — ignored', entry.get('id'))
                    continue
                for change in entry.get('changes', []):
                    if change.get('field') == 'leadgen':
                        val        = change.get('value', {})
                        leadgen_id = val.get('leadgen_id')
                        campaign   = val.get('campaign_name', '') or ''
                        adset      = val.get('adset_name', '') or val.get('adgroup_name', '') or ''
                        ad         = val.get('ad_name', '') or ''
                        form_id    = str(val.get('form_id', '') or '')
                        if leadgen_id:
                            meta_data = _fetch_meta_lead_data(leadgen_id, config.page_access_token)
                            if meta_data and meta_data.get('field_data'):
                                ad    = meta_data.get('ad_name') or ad
                                ad_id = meta_data.get('ad_id')
                                # Prefer the form_id from the Graph lead object; the
                                # webhook payload doesn't always include it.
                                form_id = str(meta_data.get('form_id') or form_id or '')
                                if ad_id and not campaign and not adset:
                                    campaign, adset = _fetch_ad_campaign_info(ad_id, config.page_access_token)
                                _create_lead_from_meta(meta_data['field_data'], config, campaign, adset, ad, form_id)
        except Exception:
            logger.exception('Meta webhook: unhandled error processing payload')
        return Response({'ok': True})


class MetaWebhookConfigView(APIView):
    permission_classes = [IsAuthenticated]

    def _ensure_config(self, request):
        company = _resolve_company(request)
        config, created = MetaWebhookConfig.objects.get_or_create(
            company=company,
            defaults={'verify_token': secrets.token_urlsafe(32)},
        )
        if not config.verify_token:
            config.verify_token = secrets.token_urlsafe(32)
            config.save(update_fields=['verify_token'])
        return config

    def _fetch_pages_and_forms(self, pat):
        """Fetch all subscribed pages and their lead forms from Meta API."""
        pages_data, subscribed = [], []
        try:
            pages_r = http_requests.get(
                'https://graph.facebook.com/v19.0/me/accounts',
                params={'access_token': pat, 'limit': 50}, timeout=10
            )
            if pages_r.status_code == 200:
                for page in pages_r.json().get('data', []):
                    page_token = page.get('access_token')
                    page_id    = page.get('id')
                    page_name  = page.get('name', page_id)
                    if not page_token or not page_id:
                        continue
                    subscribed.append(page_name)
                    forms = []
                    try:
                        forms_r = http_requests.get(
                            f'https://graph.facebook.com/v19.0/{page_id}/leadgen_forms',
                            params={'access_token': page_token, 'fields': 'id,name', 'limit': 50},
                            timeout=10
                        )
                        if forms_r.status_code == 200:
                            forms = [{'id': f['id'], 'name': f.get('name', '')}
                                     for f in forms_r.json().get('data', [])]
                    except Exception:
                        logger.exception('Meta: failed to fetch forms for page_id=%s', page_id)
                    pages_data.append({'page_id': page_id, 'page_name': page_name, 'forms': forms})
        except Exception:
            logger.exception('Meta: failed to fetch pages list')
        return subscribed, pages_data

    def get(self, request):
        config = self._ensure_config(request)
        # Auto-refresh pages/forms if stale (older than 2 hours) or never fetched
        if config.page_access_token:
            stale = (
                not config.pages_refreshed_at or
                (timezone.now() - config.pages_refreshed_at).total_seconds() > 7200
            )
            if stale:
                subscribed, pages_data = self._fetch_pages_and_forms(config.page_access_token)
                if pages_data:
                    config.subscribed_pages  = subscribed
                    config.pages_data        = pages_data
                    config.pages_refreshed_at = timezone.now()
                    config.save(update_fields=['subscribed_pages', 'pages_data', 'pages_refreshed_at'])
        projects = list(
            scope_to_company(Project.objects.filter(is_active=True), request.user).values('id', 'name')
        )
        # Lead count per Meta form (mapped or not) so the UI can flag forms that are
        # bringing in leads but aren't yet routed to a project.
        form_lead_counts = {
            row['meta_form_id']: row['c']
            for row in scope_to_company(Lead.objects.exclude(meta_form_id=''), request.user)
                        .values('meta_form_id').annotate(c=Count('id'))
        }
        return Response({
            'verify_token':         config.verify_token,
            'page_access_token':    config.page_access_token,
            # Whether a secret is stored, never the secret itself.
            'app_secret_set':       bool(config.app_secret),
            'default_project_id':   config.default_project_id,
            'is_active':            config.is_active,
            'total_leads_received': config.total_leads_received,
            'last_lead_at':         config.last_lead_at,
            'subscribed_pages':     config.subscribed_pages or [],
            'pages_data':           config.pages_data or [],
            'form_lead_counts':     form_lead_counts,
            'projects':             projects,
        })

    def post(self, request):
        config = self._ensure_config(request)
        action = request.data.get('action')
        if action == 'debug_forms':
            pat = config.page_access_token
            debug = {}
            pages_r = http_requests.get('https://graph.facebook.com/v19.0/me/accounts',
                                        params={'access_token': pat, 'limit': 50}, timeout=10)
            debug['accounts_status'] = pages_r.status_code
            debug['pages'] = []
            if pages_r.status_code == 200:
                for page in pages_r.json().get('data', []):
                    page_id = page.get('id')
                    page_name = page.get('name', page_id)
                    page_tok = page.get('access_token')
                    forms_r = http_requests.get(
                        f'https://graph.facebook.com/v19.0/{page_id}/leadgen_forms',
                        params={'access_token': page_tok, 'fields': 'id,name', 'limit': 50}, timeout=10)
                    debug['pages'].append({
                        'page': page_name,
                        'page_id': page_id,
                        'forms_status': forms_r.status_code,
                        'forms_response': forms_r.json(),
                    })
            else:
                debug['accounts_error'] = pages_r.json()
            return Response(debug)
        if action == 'regenerate_token':
            config.verify_token = secrets.token_urlsafe(32)
            config.save(update_fields=['verify_token'])
            return Response({'verify_token': config.verify_token})
        if action == 'save':
            pat = request.data.get('page_access_token', '').strip()
            pid = request.data.get('default_project_id')
            if pid and not _project_in_scope(request, pid):
                return Response({'detail': 'Invalid project for your company.'}, status=400)
            config.page_access_token = pat
            # Optional but strongly recommended: without it deliveries can't be
            # verified. Only overwrite when a value is supplied, so saving other
            # settings doesn't wipe a secret already stored.
            secret = str(request.data.get('app_secret', '') or '').strip()
            if secret:
                config.app_secret = secret
            config.default_project_id = pid if pid else None
            config.is_active = bool(pat)
            config.save(update_fields=['page_access_token', 'app_secret',
                                       'default_project_id', 'is_active'])
            # Subscribe app to all accessible pages' leadgen events
            subscribed, failed, pages_data = [], [], []
            if pat:
                try:
                    pages_r = http_requests.get(
                        'https://graph.facebook.com/v19.0/me/accounts',
                        params={'access_token': pat, 'limit': 50}, timeout=10
                    )
                    if pages_r.status_code == 200:
                        for page in pages_r.json().get('data', []):
                            page_token = page.get('access_token')
                            page_id    = page.get('id')
                            page_name  = page.get('name', page_id)
                            if not page_token or not page_id:
                                continue
                            sub_r = http_requests.post(
                                f'https://graph.facebook.com/v19.0/{page_id}/subscribed_apps',
                                params={'access_token': page_token,
                                        'subscribed_fields': 'leadgen'}, timeout=10
                            )
                            if sub_r.status_code == 200 and sub_r.json().get('success'):
                                subscribed.append(page_name)
                            else:
                                failed.append(page_name)
                except Exception:
                    logger.exception('Meta: failed to subscribe pages to app')
            _, pages_data = self._fetch_pages_and_forms(pat) if pat else ([], [])
            config.subscribed_pages   = subscribed
            config.pages_data         = pages_data
            config.pages_refreshed_at = timezone.now()
            config.save(update_fields=['subscribed_pages', 'pages_data', 'pages_refreshed_at'])
            return Response({'ok': True, 'is_active': config.is_active,
                             'subscribed_pages': subscribed, 'failed_pages': failed,
                             'pages_data': pages_data})
        return Response({'detail': 'Unknown action'}, status=400)


def _backfill_form_mapping(company, form_id, project, page_access_token=None):
    """Assign `project` to existing UNMAPPED leads that belong to this form, so a
    mapping added/fixed after leads arrived also fixes those leads. Two passes:
      1) leads already tagged with this form_id (stored on the lead);
      2) best-effort — leads with no/blank project that match (by phone) a lead in
         this form on Meta, covering leads that arrived before form_id was stored
         or without a form_id in the webhook payload.
    Returns the number of leads updated."""
    fid = str(form_id)
    n = Lead.objects.filter(company=company, project__isnull=True, meta_form_id=fid).update(project=project)
    if page_access_token:
        try:
            import urllib.request, json as _json
            phones, url, pages = set(), (
                f'https://graph.facebook.com/v19.0/{fid}/leads?fields=field_data&limit=200&access_token={page_access_token}'), 0
            while url and pages < 6:
                d = _json.load(urllib.request.urlopen(url, timeout=25))
                for r in d.get('data', []):
                    for f in r.get('field_data', []):
                        if 'phone' in (f.get('name', '').lower()):
                            digits = ''.join(c for c in (f.get('values') or [''])[0] if c.isdigit())[-10:]
                            if len(digits) >= 10:
                                phones.add(digits)
                url = d.get('paging', {}).get('next'); pages += 1
            for digits in phones:
                # endswith (not a (^|\D)…$ boundary regex): a +91-prefixed number like
                # +919510188522 has its 10-digit core preceded by the '1' of +91, so a
                # \D boundary never matches. Last-10 endswith matches the same number.
                n += Lead.objects.filter(
                    company=company, project__isnull=True,
                    phone_key=phone_blind_index(digits),
                ).update(project=project, meta_form_id=fid)
        except Exception:
            logger.exception('Meta backfill failed for form_id=%s', fid)
    return n


class MetaFormMappingView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company = _resolve_company(request)
        mappings = MetaFormMapping.objects.select_related('project').filter(
            company=company
        ).order_by('-created_at')
        return Response([{
            'id':          m.id,
            'form_id':     m.form_id,
            'form_name':   m.form_name,
            'project_id':  m.project_id,
            'project_name':m.project.name,
            'total_leads': m.total_leads,
        } for m in mappings])

    def post(self, request):
        form_id   = request.data.get('form_id', '').strip()
        form_name = request.data.get('form_name', '').strip()
        project_id = request.data.get('project_id')
        if not form_id or not project_id:
            return Response({'detail': 'form_id and project_id are required.'}, status=400)
        company = _resolve_company(request)
        try:
            project = Project.objects.filter(company=company).get(pk=project_id)
        except Project.DoesNotExist:
            return Response({'detail': 'Project not found.'}, status=404)
        mapping, created = MetaFormMapping.objects.update_or_create(
            form_id=form_id,
            defaults={'form_name': form_name, 'project': project, 'company': project.company},
        )
        # Retroactively map existing unmapped leads from this form.
        cfg = MetaWebhookConfig.objects.filter(company=company).first()
        backfilled = _backfill_form_mapping(
            company, form_id, project, cfg.page_access_token if cfg else None)
        return Response({
            'id': mapping.id, 'form_id': mapping.form_id,
            'form_name': mapping.form_name, 'project_id': mapping.project_id,
            'project_name': mapping.project.name, 'total_leads': mapping.total_leads,
            'backfilled': backfilled,
        }, status=201 if created else 200)

    def delete(self, request):
        mid = request.data.get('id')
        MetaFormMapping.objects.filter(pk=mid, company=_resolve_company(request)).delete()
        return Response({'ok': True})


# ── User Project Assignments ──────────────────────────────────────────────────
class UserProjectAssignmentView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user_id = request.query_params.get('user_id')
        if not user_id:
            return Response({'detail': 'user_id required.'}, status=400)
        assigned = scope_to_company(
            UserProjectAssignment.objects.filter(user_id=user_id),
            request.user, 'user__company',
        ).values_list('project_id', flat=True)
        return Response(list(assigned))

    def post(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=403)
        user_id     = request.data.get('user_id')
        project_ids = request.data.get('project_ids', [])
        try:
            user = User.objects.get(pk=user_id, company=request.user.company)
        except User.DoesNotExist:
            return Response({'detail': 'User not found.'}, status=404)
        # Only allow assigning projects that belong to the requester's company.
        valid_ids = list(
            scope_to_company(Project.objects.filter(pk__in=project_ids), request.user)
            .values_list('id', flat=True)
        )
        UserProjectAssignment.objects.filter(user=user).delete()
        UserProjectAssignment.objects.bulk_create([
            UserProjectAssignment(user=user, project_id=pid) for pid in valid_ids
        ], ignore_conflicts=True)
        return Response({'user_id': user_id, 'project_ids': valid_ids})


# ── Bulk Plot Creation ────────────────────────────────────────────────────────
class PlotBulkCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=403)
        project_id = request.data.get('project_id')
        plots_data = request.data.get('plots', [])
        try:
            project = scope_to_company(Project.objects.all(), request.user).get(pk=project_id)
        except Project.DoesNotExist:
            return Response({'detail': 'Project not found.'}, status=404)
        # Tower floor-builder sends the whole unit record (floor, areas, facing, price),
        # not just a number — carry every field through instead of dropping them.
        def _floor(v):
            try: return int(v)
            except (TypeError, ValueError): return None
        plots = [
            Plot(
                project=project,
                number=p.get('number', ''),
                cluster_type=p.get('cluster_type', ''),
                size=p.get('size', '') or '',
                construction_area=p.get('construction_area', '') or '',
                terrace_area=p.get('terrace_area', '') or '',
                facing=p.get('facing', '') or '',
                price=p.get('price', '') or '',
                notes=p.get('notes', '') or '',
                floor=_floor(p.get('floor')),
                status='available',
            )
            for p in plots_data
            if p.get('number')
        ]
        # Re-running the builder for one floor must not blow up on units that already
        # exist (project+number is unique) — skip the clashes, report what landed.
        created = Plot.objects.bulk_create(plots, ignore_conflicts=True)
        n = Plot.objects.filter(project=project, number__in=[p.number for p in plots]).count()
        return Response({'created': len(plots), 'existing_total': n}, status=201)


class PlotBulkDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=403)
        project_id = request.data.get('project_id')
        if not project_id:
            return Response({'detail': 'project_id is required.'}, status=400)
        if not _project_in_scope(request, project_id):
            return Response({'detail': 'Project not found.'}, status=404)
        deleted, _ = Plot.objects.filter(project_id=project_id).delete()
        Project.objects.filter(pk=project_id).update(total_plots=0)
        return Response({'deleted': deleted})


class PlotRenameTypeView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not is_admin_or_manager(request.user):
            return Response({'detail': 'Permission denied.'}, status=403)
        project_id = request.data.get('project_id')
        old_name   = request.data.get('old_name', '').strip()
        new_name   = request.data.get('new_name', '').strip()
        if not project_id or not old_name or not new_name:
            return Response({'detail': 'project_id, old_name and new_name are required.'}, status=400)
        if not _project_in_scope(request, project_id):
            return Response({'detail': 'Project not found.'}, status=404)
        updated = Plot.objects.filter(project_id=project_id, cluster_type=old_name).update(cluster_type=new_name)
        return Response({'updated': updated})


class SalesDataResetView(APIView):
    """Admin-only, company-scoped: wipe TRIAL transactional data (leads + their
    history/follow-ups/site-visits/closures, bookings, distribution log,
    availability, notifications) and reset all plots to 'available'. KEEPS setup:
    company, users, projects, plot definitions, lead sources, team/distribution
    config. POST requires confirm='DELETE'. GET returns current counts."""
    permission_classes = [IsAuthenticated]

    def _is_admin(self, user):
        return bool(
            getattr(user, 'is_staff', False) or getattr(user, 'role', '') == 'Admin' or is_platform_admin(user)
            or 'Sales' in (getattr(user, 'admin_modules', None) or [])
        )

    def _counts(self, co):
        from accounts.models import Notification
        return {
            'leads':            Lead.objects.filter(company=co).count(),
            'follow_ups':       FollowUp.objects.filter(lead__company=co).count(),
            'site_visits':      SiteVisit.objects.filter(lead__company=co).count(),
            'bookings':         Booking.objects.filter(company=co).count(),
            'cancelled_bookings': Booking.objects.filter(company=co, approval_status='CANCELLED').count(),
            'closures':         Closure.objects.filter(company=co).count(),
            'lead_history':     LeadStatusHistory.objects.filter(lead__company=co).count(),
            'distribution_log': DistributionLog.objects.filter(company=co).count(),
            'availability':     UserAvailability.objects.filter(user__company=co).count(),
            'notifications':    Notification.objects.filter(recipient__company=co).count(),
            'plots_to_reset':   Plot.objects.filter(project__company=co).exclude(status='available').count(),
        }

    def get(self, request):
        if not self._is_admin(request.user):
            return Response({'detail': 'Admin only.'}, status=status.HTTP_403_FORBIDDEN)
        return Response(self._counts(_resolve_company(request)))

    def post(self, request):
        if not self._is_admin(request.user):
            return Response({'detail': 'Admin only.'}, status=status.HTTP_403_FORBIDDEN)
        if (request.data.get('confirm') or '') != 'DELETE':
            return Response({'detail': 'Type DELETE to confirm.'}, status=status.HTTP_400_BAD_REQUEST)
        # A second gate the app itself does not hold: the reset key lives in the
        # server environment, so a signed-in admin — or anyone who takes over an
        # admin session — still cannot wipe the company's data without it.
        #
        # Fails closed on purpose. If DATA_RESET_KEY is unset the reset is refused
        # outright rather than silently falling back to the DELETE box, because a
        # missing key must never mean "no protection" on something irreversible.
        expected = (os.getenv('DATA_RESET_KEY') or '').strip()
        if not expected:
            return Response(
                {'detail': 'Data reset is disabled: no DATA_RESET_KEY is configured on the server.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        supplied = str(request.data.get('reset_key') or '').strip()
        if not hmac.compare_digest(supplied, expected):
            logger.warning('Data reset refused: bad key from user %s', getattr(request.user, 'id', None))
            return Response({'detail': 'Incorrect reset key.'}, status=status.HTTP_403_FORBIDDEN)
        co = _resolve_company(request)
        before = self._counts(co)
        with_attendance = bool(request.data.get('with_attendance'))
        with_loi        = bool(request.data.get('with_loi_files'))

        # Which categories to clear. Defaults to ALL (legacy behaviour) when the
        # client doesn't send an explicit selection.
        all_keys = ['bookings', 'cancelled_bookings', 'closures', 'site_visits', 'follow_ups', 'lead_history',
                    'distribution_log', 'availability', 'notifications', 'leads', 'plots_to_reset']
        raw = request.data.get('targets')
        if isinstance(raw, list) and raw:
            targets = [k for k in all_keys if k in raw]
        else:
            targets = list(all_keys)
        sel = set(targets)
        # Deleting leads cascades their children in the DB — reflect that in the summary.
        # Closures are deliberately NOT in that set: they own their company FK and only
        # SET_NULL their lead, so clearing leads leaves the conversion history (and the
        # bookings that point at it) intact.
        cascades = {'site_visits', 'follow_ups', 'lead_history'}
        effective = set(sel) | (cascades if 'leads' in sel else set())
        # An approved booking mirrors itself into a Closure; drop those alongside the
        # bookings so the two counts can never disagree. Standalone closures (imported
        # or recorded by hand, with no booking) are only removed by ticking Closures.
        booking_closures = Closure.objects.filter(
            id__in=Booking.objects.filter(company=co).exclude(closure=None).values('closure_id')
        ) if 'bookings' in sel else Closure.objects.none()
        n_booking_closures = booking_closures.count()

        # Optionally purge confidential LOI PDFs from Supabase before deleting bookings.
        if with_loi and 'bookings' in sel:
            for b in Booking.objects.filter(company=co).exclude(loi_document=''):
                try: b.loi_document.delete(save=False)
                except Exception: pass

        from django.db import transaction
        from accounts.models import Notification
        with transaction.atomic():
            if 'bookings' in sel:
                Closure.objects.filter(id__in=list(booking_closures.values_list('id', flat=True))).delete()
                Booking.objects.filter(company=co).delete()
            # Purge only the cancelled booking records (the CANCELLED log entries).
            if 'cancelled_bookings' in sel and 'bookings' not in sel:
                Booking.objects.filter(company=co, approval_status='CANCELLED').delete()
            if 'closures' in sel:         Closure.objects.filter(company=co).delete()
            if 'site_visits' in sel:      SiteVisit.objects.filter(lead__company=co).delete()
            if 'follow_ups' in sel:       FollowUp.objects.filter(lead__company=co).delete()
            if 'lead_history' in sel:     LeadStatusHistory.objects.filter(lead__company=co).delete()
            if 'distribution_log' in sel: DistributionLog.objects.filter(company=co).delete()
            if 'availability' in sel:     UserAvailability.objects.filter(user__company=co).delete()
            if 'notifications' in sel:    Notification.objects.filter(recipient__company=co).delete()
            if 'leads' in sel:            Lead.objects.filter(company=co).delete()  # cascades children
            if 'plots_to_reset' in sel:   Plot.objects.filter(project__company=co).exclude(status='available').update(status='available')
            if with_attendance:
                from attendance.models import AttendanceRecord, LeaveApplication, LeaveTransaction, LeaveBalance
                AttendanceRecord.objects.filter(user__company=co).delete()
                LeaveApplication.objects.filter(user__company=co).delete()
                LeaveTransaction.objects.filter(user__company=co).delete()
                LeaveBalance.objects.filter(user__company=co).delete()
        deleted = {k: v for k, v in before.items() if k in effective}
        if n_booking_closures and 'closures' not in effective:
            deleted['closures'] = n_booking_closures  # the booking-mirrored ones only
        return Response({'detail': 'Trial data cleared.', 'deleted': deleted, 'targets': sorted(effective)})


class BackupSettingsView(APIView):
    """Platform-super-user-only: view/update the automatic backup schedule."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not is_platform_admin(request.user):
            return Response({'detail': 'Super admin only.'}, status=status.HTTP_403_FORBIDDEN)
        settings_row, _ = BackupSettings.objects.get_or_create(pk=1)
        return Response(BackupSettingsSerializer(settings_row).data)

    def patch(self, request):
        if not is_platform_admin(request.user):
            return Response({'detail': 'Super admin only.'}, status=status.HTTP_403_FORBIDDEN)
        settings_row, _ = BackupSettings.objects.get_or_create(pk=1)
        ser = BackupSettingsSerializer(settings_row, data=request.data, partial=True)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        ser.save(updated_by=request.user)
        return Response(ser.data)


class BackupListView(APIView):
    """Platform-super-user-only: recent backup history."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not is_platform_admin(request.user):
            return Response({'detail': 'Super admin only.'}, status=status.HTTP_403_FORBIDDEN)
        records = BackupRecord.objects.select_related('triggered_by')[:50]
        return Response(BackupRecordSerializer(records, many=True).data)


class BackupRunNowView(APIView):
    """Platform-super-user-only: trigger a backup immediately, outside the schedule."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not is_platform_admin(request.user):
            return Response({'detail': 'Super admin only.'}, status=status.HTTP_403_FORBIDDEN)
        from .backup_service import run_backup
        record = run_backup(triggered_by=request.user)
        return Response(BackupRecordSerializer(record).data,
                         status=status.HTTP_201_CREATED if record.status == 'success' else status.HTTP_502_BAD_GATEWAY)


class BackupDownloadView(APIView):
    """Platform-super-user-only: a short-lived signed URL for one backup file."""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        if not is_platform_admin(request.user):
            return Response({'detail': 'Super admin only.'}, status=status.HTTP_403_FORBIDDEN)
        record = BackupRecord.objects.filter(pk=pk, status='success').first()
        if not record or not record.file_path:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        from .backup_storage import signed_backup_url
        url = signed_backup_url(record.file_path)
        if not url:
            return Response({'detail': 'Could not generate a download link.'}, status=status.HTTP_502_BAD_GATEWAY)
        return Response({'url': url})


# ─────────────────────────────────────────────────────────────────────────────
# Lead transfer: one STM hands a lead to another, held for approval
# ─────────────────────────────────────────────────────────────────────────────
def _transfer_approver_ids(company, project_id):
    """Who may sign off a transfer of a lead on this project.

    The same people who approve that project's bookings — the list an admin sets in
    Booking & Approvals. Deliberately not "any manager": a lead moving between reps
    changes whose numbers it lands in, so it needs the same named authority a booking
    does. A project with nobody named leaves only real admins, which is the same rule
    _can_approve_project applies.
    """
    ids = set()
    for p in Project.objects.filter(company=company).only('id', 'booking_approvers'):
        if project_id and p.id != project_id:
            continue
        ids.update(p.booking_approvers or [])
    return sorted(ids)


class LeadTransferListCreateView(APIView):
    """GET  — transfers this user can see (their own requests, plus the queue they approve).
    POST — request a transfer of one of your leads to another STM."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company = _resolve_company(request)
        qs = (LeadTransfer.objects.filter(company=company)
              .select_related('lead', 'project', 'from_stm', 'to_stm', 'requested_by', 'decided_by'))
        if not _is_hard_admin(request.user):
            approver_pids = _approver_project_ids(request.user, company)
            qs = qs.filter(
                Q(requested_by=request.user) | Q(from_stm=request.user) | Q(to_stm=request.user)
                | Q(project_id__in=approver_pids)
            )
        if request.query_params.get('status'):
            qs = qs.filter(status=request.query_params['status'])
        return Response(LeadTransferSerializer(qs[:200], many=True).data)

    def post(self, request):
        company = _resolve_company(request)
        lead_id = request.data.get('lead')
        to_stm_id = request.data.get('to_stm')
        if not lead_id or not to_stm_id:
            return Response({'detail': 'Lead and destination STM are required.'},
                            status=status.HTTP_400_BAD_REQUEST)
        lead = Lead.objects.filter(pk=lead_id, company=company).first()
        if not lead:
            return Response({'detail': 'Lead not found.'}, status=status.HTTP_404_NOT_FOUND)
        # Only the rep holding the lead may hand it on (admins may act for them).
        if not _is_hard_admin(request.user) and lead.stm_id != request.user.id:
            return Response({'detail': 'This lead is not assigned to you.'},
                            status=status.HTTP_403_FORBIDDEN)
        if str(to_stm_id) == str(lead.stm_id):
            return Response({'detail': 'That is already the assigned STM.'},
                            status=status.HTTP_400_BAD_REQUEST)
        to_stm = User.objects.filter(pk=to_stm_id, company=company, is_active=True).first()
        if not to_stm:
            return Response({'detail': 'Destination STM not found in your company.'},
                            status=status.HTTP_404_NOT_FOUND)
        if LeadTransfer.objects.filter(lead=lead, status='pending').exists():
            return Response({'detail': 'A transfer for this lead is already awaiting approval.'},
                            status=status.HTTP_409_CONFLICT)

        t = LeadTransfer.objects.create(
            company=company, lead=lead, project_id=lead.project_id,
            from_stm_id=lead.stm_id, to_stm=to_stm, requested_by=request.user,
            reason=(request.data.get('reason') or '').strip(),
        )
        _notify_transfer_requested(t)
        return Response(LeadTransferSerializer(t).data, status=status.HTTP_201_CREATED)


class LeadTransferActionView(APIView):
    """Approve or reject a pending transfer. Approval is what actually moves the lead."""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        company = _resolve_company(request)
        t = LeadTransfer.objects.filter(pk=pk, company=company).select_related('lead').first()
        if not t:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        action = request.data.get('action')

        # The requester may withdraw their own request; everything else needs an approver.
        if action == 'cancel':
            if t.requested_by_id != request.user.id and not _is_hard_admin(request.user):
                return Response({'detail': 'Only the requester can withdraw this.'},
                                status=status.HTTP_403_FORBIDDEN)
        elif not _can_approve_project(request.user, t.project_id, company):
            return Response({'detail': 'You are not an approver for this project.'},
                            status=status.HTTP_403_FORBIDDEN)

        if t.status != 'pending':
            return Response({'detail': 'This request has already been %s.' % t.status},
                            status=status.HTTP_409_CONFLICT)
        if action not in ('approve', 'reject', 'cancel'):
            return Response({'detail': 'Unknown action.'}, status=status.HTTP_400_BAD_REQUEST)

        note = (request.data.get('note') or '').strip()
        if action == 'approve':
            with transaction.atomic():
                # Re-read under a lock: the lead may have moved since the request was
                # raised, and the record should say what it actually moved from.
                lead = Lead.objects.select_for_update().get(pk=t.lead_id)
                t.from_stm_id = lead.stm_id
                lead.stm_id = t.to_stm_id
                lead.stm_assigned_at = timezone.now()
                lead.save(update_fields=['stm', 'stm_assigned_at'])
                t.status = 'approved'
                t.decided_by = request.user
                t.decided_at = timezone.now()
                t.decision_note = note
                t.save(update_fields=['status', 'decided_by', 'decided_at', 'decision_note',
                                      'from_stm', 'updated_at'])
        else:
            t.status = 'rejected' if action == 'reject' else 'cancelled'
            t.decided_by = request.user
            t.decided_at = timezone.now()
            t.decision_note = note
            t.save(update_fields=['status', 'decided_by', 'decided_at', 'decision_note', 'updated_at'])

        _notify_transfer_decided(t)
        return Response(LeadTransferSerializer(t).data)


def _notify_transfer_requested(t):
    """Tell the approvers there is something to look at, and the receiving STM it is coming."""
    try:
        from notifications import notify, notify_many
        who = User.objects.filter(id__in=_transfer_approver_ids(t.company, t.project_id))
        body = '%s wants to transfer %s to %s' % (
            getattr(t.requested_by, 'name', 'An STM') or 'An STM',
            getattr(t.lead, 'name', 'a lead') or 'a lead',
            getattr(t.to_stm, 'name', 'another STM') or 'another STM')
        notify_many(who, 'lead_transfer_requested', 'Lead Transfer Request', body,
                    data={'transfer': t.id, 'lead': t.lead_id})
        if t.to_stm:
            notify(t.to_stm, 'lead_transfer_requested', 'Lead Coming Your Way',
                   '%s has asked to transfer %s to you — awaiting approval.' % (
                       getattr(t.requested_by, 'name', 'An STM') or 'An STM',
                       getattr(t.lead, 'name', 'a lead') or 'a lead'),
                   data={'transfer': t.id, 'lead': t.lead_id})
    except Exception:
        logger.exception('Could not notify for lead transfer %s', t.pk)


def _notify_transfer_decided(t):
    """Tell both reps the outcome — the one losing the lead and the one gaining it."""
    try:
        from notifications import notify
        verb = {'approved': 'approved', 'rejected': 'rejected', 'cancelled': 'withdrawn'}[t.status]
        lead_name = getattr(t.lead, 'name', 'a lead') or 'a lead'
        for person in {t.requested_by_id: t.requested_by, t.from_stm_id: t.from_stm,
                       t.to_stm_id: t.to_stm}.values():
            if person and person.id != t.decided_by_id:
                notify(person, 'lead_transfer_%s' % t.status, 'Lead Transfer %s' % verb.title(),
                       'The transfer of %s was %s.' % (lead_name, verb),
                       data={'transfer': t.id, 'lead': t.lead_id})
    except Exception:
        logger.exception('Could not notify the outcome of lead transfer %s', t.pk)
